#!/usr/bin/env python3
"""
ExportaGarmin: tus datos de Garmin, ordenados y preparados para la IA.

Descarga datos de salud, forma física y actividades mediante la biblioteca
python-garminconnect y los guarda como texto con bloques JSON. El modo completo
conserva las respuestas originales de la API; el compacto crea una exportación
semántica y privada para NotebookLM, ChatGPT, Claude y otras IA.

No necesita una clave oficial. La biblioteca utiliza el mismo acceso SSO que la
web de Garmin. Los tokens se guardan localmente durante aproximadamente un año.

Incluye perfil, salud diaria, actividades, composición corporal, métricas de
entrenamiento, objetivos, tendencias, Golf, equipamiento, planes, hidratación,
nutrición y salud femenina.

Usa --all para exportar el historial completo. Consulta README.md para la
instalación y el uso.
"""

import argparse
import json
import logging
import math
import os
import re
import sys
import threading
import time
import unicodedata
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import date, datetime, timedelta, timezone
from getpass import getpass
from pathlib import Path
from typing import Any, Optional
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError

from garth.exc import GarthHTTPError

from garminconnect import (
    Garmin,
    GarminConnectAuthenticationError,
    GarminConnectConnectionError,
    GarminConnectTooManyRequestsError,
)

from training_analysis import (
    SCHEMA_VERSION,
    activity_catalog_entry,
    atomic_write_json,
    atomic_write_text,
    build_quality_report,
    build_report_extensions,
    is_personal_data_key,
    load_local_json,
    load_or_create_reference_secret,
    normalise_journal,
    normalise_race_context,
    private_reference,
    privacy_audit,
    render_xlsx,
)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("garmin_export")
logging.getLogger("garminconnect").setLevel(logging.WARNING)


# ---------------------------------------------------------------------------
# Regulador adaptativo: reacciona a los 429 y es seguro entre hilos.
# ---------------------------------------------------------------------------
class RateLimiter:
    """Regula llamadas y solo reduce el ritmo cuando Garmin lo solicita."""

    def __init__(self, base_delay: float = 0.15):
        self.base_delay = base_delay
        self.current_delay = base_delay
        self.call_count = 0
        self.last_call = 0.0
        self.blocked_until = 0.0
        self.consecutive_ok = 0
        self._lock = threading.Lock()

    def wait(self):
        preventive_pause_pending = False
        while True:
            with self._lock:
                now = time.monotonic()
                if preventive_pause_pending:
                    wait_for = self.blocked_until - now
                    if wait_for <= 0:
                        self.last_call = now
                        return
                else:
                    wait_for = max(
                        self.blocked_until - now,
                        self.current_delay - (now - self.last_call),
                    )
                if wait_for <= 0:
                    self.last_call = now
                    self.call_count += 1
                    if self.call_count % 250 == 0:
                        log.info(
                            "  Pausa de seguridad después de "
                            f"{self.call_count} llamadas a la API..."
                        )
                        self.blocked_until = max(
                            self.blocked_until,
                            now + 2,
                        )
                        preventive_pause_pending = True
                    else:
                        return
            time.sleep(max(wait_for, 0.01))

    def on_success(self):
        with self._lock:
            self.consecutive_ok += 1
            if self.consecutive_ok > 10 and self.current_delay > self.base_delay:
                self.current_delay = max(self.base_delay, self.current_delay * 0.9)

    def on_rate_limit(self):
        with self._lock:
            self.consecutive_ok = 0
            self.current_delay = min(self.current_delay * 2, 10.0)
            self.blocked_until = max(
                self.blocked_until,
                time.monotonic() + 60,
            )
            log.warning(
                "  Límite de Garmin alcanzado: nueva espera "
                f"{self.current_delay:.1f}s; pausando 60s todos los hilos..."
            )

    def on_error(self):
        with self._lock:
            self.consecutive_ok = 0
            self.current_delay = min(self.current_delay * 1.2, 5.0)


_limiter = RateLimiter()
_safe_call_failure_handler = None
_safe_call_failure_handler_lock = threading.Lock()
_safe_call_state = threading.local()


def _set_safe_call_failure_handler(handler) -> None:
    """Asocia temporalmente fallos saneados con la sección que se exporta."""
    global _safe_call_failure_handler
    with _safe_call_failure_handler_lock:
        _safe_call_failure_handler = handler


def _report_safe_call_failure(endpoint: str, exc: Exception) -> None:
    _safe_call_state.failed = True
    with _safe_call_failure_handler_lock:
        handler = _safe_call_failure_handler
    if handler is not None:
        try:
            handler(endpoint, _safe_exception_reason(exc))
        except Exception:
            log.debug(
                "No se pudo registrar una incidencia de endpoint; "
                "se ha omitido el detalle para proteger la salida."
            )


def _safe_call_with_status(fn, *args, label: str = "", **kwargs):
    """Devuelve el valor y si la llamada terminó sin un fallo técnico."""
    result = safe_call(fn, *args, label=label, **kwargs)
    return result, not bool(getattr(_safe_call_state, "failed", False))


def _exception_http_status(exc: Exception) -> Optional[int]:
    """Obtiene un código HTTP sin serializar la excepción ni su URL."""
    response = getattr(exc, "response", None)
    status = getattr(response, "status_code", None)
    if status is None:
        wrapped_error = getattr(exc, "error", None)
        status = getattr(
            getattr(wrapped_error, "response", None),
            "status_code",
            None,
        )
    if status is None and getattr(exc, "__cause__", None) is not None:
        status = getattr(
            getattr(exc.__cause__, "response", None),
            "status_code",
            None,
        )
    return status if isinstance(status, int) else None


def _safe_exception_reason(exc: Exception) -> str:
    """Describe un fallo sin incluir mensajes que puedan contener datos."""
    class_name = re.sub(
        r"[^A-Za-z0-9_.-]",
        "",
        type(exc).__name__,
    )[:80] or "Error"
    status = _exception_http_status(exc)
    return f"{class_name} (HTTP {status})" if status is not None else class_name


def _safe_endpoint_name(fn: Any) -> str:
    """Devuelve únicamente el nombre técnico controlado del método llamado."""
    name = getattr(fn, "__name__", "") or "endpoint"
    return re.sub(r"[^A-Za-z0-9_]", "", str(name))[:80] or "endpoint"


def safe_call(fn, *args, label: str = "", **kwargs) -> Optional[Any]:
    """Llama a Garmin con regulación adaptativa y control de errores."""
    endpoint = _safe_endpoint_name(fn)
    _safe_call_state.failed = False
    _limiter.wait()
    try:
        result = fn(*args, **kwargs)
        _limiter.on_success()
        return result
    except GarminConnectTooManyRequestsError:
        _limiter.on_rate_limit()
        _limiter.wait()
        try:
            result = fn(*args, **kwargs)
            _limiter.on_success()
            return result
        except Exception as e:
            log.warning(
                f"  El reintento de {endpoint} ha fallado: "
                f"{_safe_exception_reason(e)}"
            )
            _limiter.on_error()
            _report_safe_call_failure(endpoint, e)
            return None
    except GarthHTTPError as e:
        status = _exception_http_status(e)
        if status == 429:
            _limiter.on_rate_limit()
            _limiter.wait()
            try:
                result = fn(*args, **kwargs)
                _limiter.on_success()
                return result
            except Exception as retry_error:
                log.warning(
                    f"  El reintento de {endpoint} ha fallado: "
                    f"{_safe_exception_reason(retry_error)}"
                )
                _limiter.on_error()
                _report_safe_call_failure(endpoint, retry_error)
                return None
        if status in (400, 404):
            log.debug(f"  {endpoint} no está disponible ({status})")
        else:
            log.warning(
                f"  Fallo de {endpoint}: {_safe_exception_reason(e)}"
            )
            _limiter.on_error()
            _report_safe_call_failure(endpoint, e)
        return None
    except Exception as e:
        log.warning(
            f"  Fallo de {endpoint}: {_safe_exception_reason(e)}"
        )
        _limiter.on_error()
        _report_safe_call_failure(endpoint, e)
        return None


# ---------------------------------------------------------------------------
# Autenticación
# ---------------------------------------------------------------------------
def _load_env_file():
    """Carga solo variables Garmin desde el .env situado junto al script."""
    env_path = Path(__file__).resolve().parent / ".env"
    allowed = {"GARMIN_EMAIL", "GARMIN_PASSWORD"}
    if env_path.exists():
        for line in env_path.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if line and not line.startswith("#") and "=" in line:
                key, _, val = line.partition("=")
                key, val = key.strip(), val.strip().strip("\"'")
                if key in allowed and val:
                    os.environ.setdefault(key, val)
        return True
    return False


def _friendly_login_error(exc: Exception) -> str:
    """Extrae un mensaje breve y comprensible de una excepción de acceso.

    Los errores originales de garth incluyen la URL SSO completa. Se elimina
    ese ruido para mostrar una indicación útil.
    """
    msg = str(exc)
    # GarthHTTPError envuelve HTTPError; recuperar su código de estado.
    status = _exception_http_status(exc)

    # Comprobar también códigos mencionados dentro del mensaje.
    if status is None:
        for code in ("401", "403", "429"):
            if code in msg:
                status = int(code)
                break

    if status == 401:
        return "Correo o contraseña incorrectos (401)."
    if status == 403:
        return "Acceso denegado (403). La cuenta puede estar bloqueada; intenta entrar primero en connect.garmin.com."
    if status == 429:
        return "Demasiados intentos de inicio de sesión (429). Espera unos minutos y vuelve a intentarlo."
    if "authentication" in msg.lower() or "unauthorized" in msg.lower():
        return "Correo o contraseña incorrectos."
    if "connection" in msg.lower() or "timeout" in msg.lower():
        return "No se pudo contactar con Garmin. Comprueba la conexión a Internet."

    return (
        "Garmin no pudo completar el inicio de sesión "
        f"({_safe_exception_reason(exc)})."
    )


def _print_login_error(exc: Exception, attempt: int, max_attempts: int):
    friendly = _friendly_login_error(exc)
    log.error(f"El inicio de sesión ha fallado: {friendly}")
    log.debug(f"Tipo de fallo de acceso: {_safe_exception_reason(exc)}")
    if attempt < max_attempts:
        remaining = max_attempts - attempt
        suffix = "intento" if remaining == 1 else "intentos"
        print(f"\n  Quedan {remaining} {suffix}. Vuelve a intentarlo:\n")


def _persist_auth_tokens(garmin: Garmin, tokenstore_path: Path):
    """Guarda tokens con clientes antiguos y actuales de garminconnect."""
    auth_client = getattr(garmin, "garth", None) or getattr(garmin, "client", None)
    if auth_client is None or not hasattr(auth_client, "dump"):
        raise RuntimeError("Versión de garminconnect no compatible: no se encontró la función para guardar tokens")
    auth_client.dump(str(tokenstore_path))


def authenticate(
    tokenstore: str,
    *,
    use_credential_environment: bool = True,
    interactive: bool = True,
    force_login: bool = False,
) -> Garmin:
    """Autentica en Garmin Connect.

    Flujo:
      1. Probar tokens guardados, normalmente válidos alrededor de un año.
      2. Si está permitido, cargar .env o variables de credenciales.
      3. Si está permitido, preguntar de forma interactiva y guardar tokens.

    El lanzador gráfico desactiva el paso 2 para que un perfil nunca herede
    credenciales de otra persona. Durante una exportación también desactiva el
    paso 3: una sesión caducada debe renovarse expresamente desde el asistente.
    """
    tokenstore_path = Path(tokenstore).expanduser()

    # Paso 1: probar tokens guardados sin solicitar credenciales.
    if tokenstore_path.exists() and not force_login:
        try:
            garmin = Garmin()
            garmin.login(str(tokenstore_path))
            log.info("Sesión iniciada con los tokens guardados")
            return garmin
        except (FileNotFoundError, GarthHTTPError, GarminConnectAuthenticationError,
                GarminConnectConnectionError) as e:
            log.info(f"Los tokens guardados han caducado o no son válidos ({type(e).__name__}); hace falta iniciar sesión")
        except Exception as e:
            log.info(
                "No se pudieron cargar los tokens "
                f"({_safe_exception_reason(e)}); hace falta iniciar sesión"
            )

    if not interactive:
        raise RuntimeError(
            "No hay una sesión válida para este perfil. "
            "Abre el asistente y pulsa «Iniciar sesión»."
        )

    # Paso 2: cargar credenciales configuradas solo cuando se ha autorizado.
    if use_credential_environment:
        _load_env_file()
        email = os.getenv("GARMIN_EMAIL")
        password = os.getenv("GARMIN_PASSWORD")
    else:
        email = None
        password = None

    if not email or not password:
        print()
        print()
        print("  Inicio de sesión de Garmin Connect")
        print("  ----------------------------------")
        print("  Las credenciales se envían directamente al acceso seguro de Garmin")
        print("  (el mismo que utiliza la página web y la aplicación).")
        print("  Los tokens de sesión se guardan localmente durante aproximadamente un año.")
        print("  La contraseña no se almacena en el proyecto.")
        print("  No compartas tu contraseña, código MFA ni tokens con nadie.")
        print()
        print()
        if not email:
            email = input("  Correo de Garmin: ").strip()
        if not password:
            password = getpass("  Contraseña de Garmin: ")

    if not email or not password:
        log.error("El correo y la contraseña son obligatorios")
        sys.exit(1)

    log.info("Iniciando sesión en Garmin Connect...")
    garmin = Garmin(email=email, password=password, is_cn=False, return_on_mfa=True)

    max_attempts = 3
    for attempt in range(1, max_attempts + 1):
        try:
            result1, result2 = garmin.login()
            break
        except GarminConnectAuthenticationError as e:
            _print_login_error(e, attempt, max_attempts)
            if attempt == max_attempts:
                sys.exit(1)
            # Permitir que se vuelvan a escribir para el siguiente intento.
            email = input("  Correo de Garmin: ").strip()
            password = getpass("  Contraseña de Garmin: ")
            garmin = Garmin(email=email, password=password, is_cn=False, return_on_mfa=True)
        except GarminConnectConnectionError as e:
            log.error("Error de conexión: no se pudo contactar con Garmin. Comprueba Internet.")
            log.debug(f"Tipo de fallo: {_safe_exception_reason(e)}")
            sys.exit(1)
        except (GarthHTTPError, Exception) as e:
            _print_login_error(e, attempt, max_attempts)
            if attempt == max_attempts:
                sys.exit(1)
            email = input("  Correo de Garmin: ").strip()
            password = getpass("  Contraseña de Garmin: ")
            garmin = Garmin(email=email, password=password, is_cn=False, return_on_mfa=True)

    if result1 == "needs_mfa":
        print()
        mfa_code = input("  Código MFA/2FA de tu aplicación de autenticación: ").strip()
        garmin.resume_login(result2, mfa_code)

    # Guardar tokens para las siguientes ejecuciones.
    tokenstore_path.mkdir(parents=True, exist_ok=True)
    _persist_auth_tokens(garmin, tokenstore_path)
    log.info("Sesión iniciada: tokens guardados en la carpeta de sesión elegida")
    log.info("   (Las próximas ejecuciones utilizarán los tokens automáticamente)")
    return garmin



# ---------------------------------------------------------------------------
# Funciones auxiliares
# ---------------------------------------------------------------------------

# Opciones globales activadas por los argumentos de consola.
_compact_mode = False
_split_mode = False
_update_mode = False
_SPLIT_WORD_LIMIT = 480000  # Margen bajo el límite de 500.000 de NotebookLM.
_COMPACT_SCHEMA_VERSION = SCHEMA_VERSION
_RECENT_ACTIVITY_REFRESH_DAYS = 14
_GEAR_REFRESH_DAYS = 7


def _word_count(text: str) -> int:
    """Calcula rápidamente un número aproximado de palabras."""
    return len(text.split())


def _strip_empty(data, preserve_list_nulls=False):
    """Quita vacíos sin desalinear las matrices posicionales de ``samples``."""
    if isinstance(data, dict):
        cleaned = {}
        for k, v in data.items():
            v = _strip_empty(v, preserve_list_nulls or k == "samples")
            if v is None or v == "" or v == [] or v == {}:
                continue
            cleaned[k] = v
        return cleaned
    elif isinstance(data, list):
        cleaned = [
            _strip_empty(item, preserve_list_nulls)
            for item in data
        ]
        if preserve_list_nulls:
            return cleaned
        return [
            item for item in cleaned
            if item is not None and item != "" and item != [] and item != {}
        ]
    return data


def _downsample_timeseries(data, key_fields=None, max_points=24):
    """Reduce series de alta frecuencia a resúmenes horarios.

    Admite listas de objetos con timestamps y matrices posicionales de Garmin.
    Divide en grupos, promedia valores numéricos y conserva el primer timestamp.
    """
    if not isinstance(data, list) or len(data) <= max_points:
        return data
    if not data:
        return data

    bucket_size = max(1, len(data) // max_points)

    # Matrices posicionales: [[timestamp, valor], ...].
    if isinstance(data[0], (list, tuple)):
        result = []
        for i in range(0, len(data), bucket_size):
            bucket = data[i:i + bucket_size]
            if not bucket:
                continue
            merged = list(bucket[0])  # Conservar el timestamp de la primera fila.
            for col in range(1, len(merged)):
                if isinstance(merged[col], (int, float)) and merged[col] is not True and merged[col] is not False:
                    vals = [row[col] for row in bucket
                            if len(row) > col
                            and isinstance(row[col], (int, float))
                            and row[col] is not True and row[col] is not False]
                    if vals:
                        merged[col] = round(sum(vals) / len(vals), 1)
            result.append(merged)
        return result

    # Listas de objetos.
    if not isinstance(data[0], dict):
        return data

    result = []
    for i in range(0, len(data), bucket_size):
        bucket = data[i:i + bucket_size]
        if not bucket:
            continue
        merged = dict(bucket[0])  # Conservar timestamps y etiquetas iniciales.
        # Promediar los campos numéricos del grupo.
        for k in merged:
            if isinstance(merged[k], (int, float)) and merged[k] is not True and merged[k] is not False:
                vals = [row[k] for row in bucket if isinstance(row.get(k), (int, float))
                        and row[k] is not True and row[k] is not False]
                if vals:
                    merged[k] = round(sum(vals) / len(vals), 1)
        result.append(merged)
    return result


def _compact_daily(data):
    """Reduce un día de salud para el modo compacto.

    Resume series frecuentes de pulso, estrés, sueño y respiración.
    """
    if not isinstance(data, dict):
        return data

    # Claves conocidas que contienen series de alta frecuencia.
    timeseries_keys = {"heart_rate", "stress", "sleep", "respiration",
                       "hrv", "body_battery", "bb_events"}

    compacted = {}
    for k, v in data.items():
        if k in timeseries_keys and isinstance(v, dict):
            # Muchas respuestas contienen la lista dentro de un objeto.
            inner = {}
            for ik, iv in v.items():
                if isinstance(iv, list) and len(iv) > 24:
                    inner[ik] = _downsample_timeseries(iv)
                else:
                    inner[ik] = iv
            compacted[k] = inner
        elif k in timeseries_keys and isinstance(v, list) and len(v) > 24:
            compacted[k] = _downsample_timeseries(v)
        else:
            compacted[k] = v
    return compacted


def _pick(sources, *keys):
    """Devuelve el primer valor no nulo para cualquiera de las claves."""
    if isinstance(sources, dict):
        sources = [sources]
    for source in sources or []:
        if not isinstance(source, dict):
            continue
        for key in keys:
            value = source.get(key)
            if value is not None:
                return value
    return None


def _number(value):
    """Devuelve un número real, excluyendo booleanos y cadenas numéricas."""
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return value
    return None


def _rounded(value, digits=1):
    value = _number(value)
    if value is None:
        return None
    return round(value, digits)


def _pace_from_speed(speed):
    """Convierte m/s a s/km sin suponer ninguna otra unidad."""
    speed = _number(speed)
    if speed is None or speed <= 0:
        return None
    return round(1000.0 / speed, 1)


_MIN_REASONABLE_EPOCH_MS = 946684800000   # 2000-01-01T00:00:00Z
_MAX_REASONABLE_EPOCH_MS = 4102444800000  # 2100-01-01T00:00:00Z


def epoch_ms_to_iso(value, timezone_name="Europe/Madrid"):
    """Convierte explícitamente un Unix epoch en milisegundos a ISO local."""
    if (
        not isinstance(value, (int, float))
        or isinstance(value, bool)
        or (isinstance(value, float) and not math.isfinite(value))
        or not (_MIN_REASONABLE_EPOCH_MS <= value <= _MAX_REASONABLE_EPOCH_MS)
    ):
        return None
    try:
        utc_datetime = datetime.fromtimestamp(
            value / 1000.0,
            tz=timezone.utc,
        )
        local_datetime = utc_datetime.astimezone(ZoneInfo(timezone_name))
    except (OSError, OverflowError, ValueError, ZoneInfoNotFoundError):
        return None
    return local_datetime.isoformat(timespec="seconds")


def _normalise_sleep_need(raw_minutes):
    """Garmin entrega ``sleepNeed.actual`` en minutos."""
    value = _number(raw_minutes)
    if value is None:
        return {}
    if 0 <= value <= 24 * 60:
        return {"sleep_need_s": round(value * 60)}
    return {
        "sleep_need_raw": value,
        "sleep_need_source_unit": "minutes",
        "sleep_need_warning": "out_of_range",
    }


def _normalise_lactate_speed(raw_speed, source_unit="unknown"):
    """Normaliza velocidad de umbral solo para unidades de origen conocidas."""
    raw = _number(raw_speed)
    if raw is None:
        return {}

    factor = None
    if source_unit in {"m/s", "meters_per_second"}:
        factor = 1.0
    elif source_unit == "garmin_tenths_m_s":
        # La respuesta de lactate-threshold usa décimas de m/s. Se conserva
        # siempre el valor de origen para que la conversión sea auditable.
        factor = 10.0

    result = {
        "speed_raw": raw,
        "speed_source_unit": source_unit,
    }
    if factor is None:
        return result

    speed_m_s = raw * factor
    pace = 1000.0 / speed_m_s if speed_m_s > 0 else None
    if not (1.0 <= speed_m_s <= 12.0 and pace is not None and 80 <= pace <= 1000):
        result["speed_warning"] = "normalised_value_out_of_range"
        return result

    result.update({
        "speed_source_factor_to_m_s": factor,
        "speed_m_s": round(speed_m_s, 4),
        "pace_s_per_km": round(pace, 1),
    })
    return result


def _normalise_temperature(raw_temperature, source_unit):
    """Devuelve grados Celsius solo cuando la unidad de origen es conocida."""
    value = _number(raw_temperature)
    if value is None:
        return {}
    if source_unit == "celsius":
        return {
            "temperature_c": value,
            "temperature_source_unit": "celsius",
        }
    if source_unit == "fahrenheit":
        return {
            "temperature_c": round((value - 32.0) * 5.0 / 9.0, 1),
            "temperature_raw": value,
            "temperature_source_unit": "fahrenheit",
        }
    return {
        "temperature_raw": value,
        "temperature_source_unit": source_unit or "unknown",
    }


_FEELING_CATEGORIES = {
    0: "very_weak",
    25: "weak",
    50: "normal",
    75: "strong",
    100: "very_strong",
}


def _normalise_self_evaluation(dto):
    """Normaliza la autoevaluación de Garmin sin confundir ausencia con cero."""
    if not isinstance(dto, dict):
        return None
    rpe_raw = _number(dto.get("directWorkoutRpe"))
    feeling_raw = _number(dto.get("directWorkoutFeel"))

    rpe = None
    if (
        rpe_raw is not None
        and 10 <= rpe_raw <= 100
        and rpe_raw % 10 == 0
    ):
        rpe = rpe_raw / 10.0

    feeling = (
        _FEELING_CATEGORIES.get(feeling_raw)
        if feeling_raw is not None
        else None
    )
    # Garmin omite estos campos si no se evaluó. Un par explícito 0/0 puede
    # ser un valor predeterminado de clientes antiguos y no cuenta como real.
    evaluated = rpe is not None or (
        feeling_raw in {25, 50, 75, 100}
    )
    if not evaluated:
        return None
    return _strip_empty({
        "perceived_exertion_raw": rpe_raw,
        "perceived_exertion_1_10": rpe,
        "feeling_raw": feeling_raw,
        "feeling": feeling,
    })


def _as_list(value):
    if isinstance(value, list):
        return value
    if value is None:
        return []
    return [value]


_COMPACT_PRIVATE_KEYS = {
    "accesscontrolruledto",
    "activityid",
    "activityuuid",
    "applicationkey",
    "deviceid",
    "displayname",
    "fullname",
    "gearid",
    "lat",
    "link",
    "lng",
    "lon",
    "ownerdisplayname",
    "ownerfullname",
    "ownerid",
    "primaryactivitytracker",
    "primarytrainingdevice",
    "profileid",
    "profilepk",
    "publicdisplayname",
    "recordeddevices",
    "registereddevices",
    "serialnumber",
    "email",
    "username",
    "unitid",
    "userid",
    "userpk",
    "userprofileid",
    "userprofilepk",
    "userprofilenumber",
    "uuid",
    "devicename",
    "devicemodel",
    "devices",
    "deviceweights",
}

_COMPACT_PRIVATE_PARTS = (
    "cookie",
    "latitude",
    "longitude",
    "password",
    "polyline",
    "profileimage",
    "imageurl",
    "location",
    "photourl",
    "url",
    "token",
)

_COMPACT_FREE_TEXT_KEYS = {
    "activityname",
    "comment",
    "comments",
    "description",
    "displaylabel",
    "goaldescription",
    "goalname",
    "label",
    "name",
    "note",
    "notes",
    "title",
}

def _sanitize_compact(data, remove_free_text=False):
    """Elimina identidad, dispositivos y ubicaciones del compacto."""
    if isinstance(data, dict):
        cleaned = {}
        for key, value in data.items():
            normal_key = re.sub(r"[^a-z0-9]", "", str(key).lower())
            if normal_key in _COMPACT_PRIVATE_KEYS:
                continue
            if normal_key.endswith("id") or normal_key.endswith("uuid"):
                continue
            if any(part in normal_key for part in _COMPACT_PRIVATE_PARTS):
                continue
            if remove_free_text and normal_key in _COMPACT_FREE_TEXT_KEYS:
                continue
            cleaned[key] = _sanitize_compact(
                value,
                remove_free_text=remove_free_text,
            )
        return cleaned
    if isinstance(data, list):
        return [
            _sanitize_compact(item, remove_free_text=remove_free_text)
            for item in data
        ]
    return data


_OPTIONAL_ACTIVITY_TEXT_KEYS = {
    "comment",
    "comments",
    "description",
    "note",
    "notes",
    "starttime",
    "starttimegmt",
    "starttimelocal",
    "endtime",
    "endtimegmt",
    "endtimelocal",
}


def _activity_source_data(
    data,
    include_free_text=False,
    parents=(),
):
    """Conserva la fuente deportiva y retira identidad e identificadores."""
    if isinstance(data, dict):
        cleaned = {}
        for key, value in data.items():
            if str(key).startswith("_"):
                continue
            normal = re.sub(r"[^a-z0-9]", "", str(key).lower())
            if not include_free_text and normal in _OPTIONAL_ACTIVITY_TEXT_KEYS:
                continue
            if is_personal_data_key(key, parents):
                continue
            cleaned[key] = _activity_source_data(
                value,
                include_free_text,
                (*parents, normal),
            )
        return cleaned
    if isinstance(data, list):
        return [
            _activity_source_data(item, include_free_text, parents)
            for item in data
        ]
    return data


_NORMALISED_ACTIVITY_SOURCE_KEYS = {
    "activitydescription",
    "activityid",
    "activityname",
    "activitytrainingload",
    "aerobictrainingeffect",
    "anaerobictrainingeffect",
    "averagegradeadjustedspeed",
    "averagehr",
    "averagepower",
    "averageruncadence",
    "averagerunningcadenceinstepsperminute",
    "averagespeed",
    "averagetemperature",
    "avggradeadjustedspeed",
    "avgpower",
    "calories",
    "description",
    "directworkoutfeel",
    "directworkoutrpe",
    "distance",
    "duration",
    "elapsedduration",
    "endelevation",
    "endlat",
    "endlatitude",
    "endlon",
    "endlongitude",
    "endaltitude",
    "endingelevation",
    "elevationgain",
    "elevationloss",
    "eventtypekey",
    "gradeadjustedspeed",
    "groundcontacttime",
    "maxelevation",
    "maxhr",
    "maximumcadenceinstepsperminute",
    "maximumelevation",
    "maxpower",
    "maxruncadence",
    "maxrunningcadenceinstepsperminute",
    "maxspeed",
    "maxtemperature",
    "minaltitude",
    "minelevation",
    "minimumelevation",
    "movingduration",
    "name",
    "normalizedpower",
    "normpower",
    "note",
    "notes",
    "startelevation",
    "startlat",
    "startlatitude",
    "startlon",
    "startlongitude",
    "startaltitude",
    "startingelevation",
    "starttimelocal",
    "stridelength",
    "title",
    "trainingeffect",
    "totalascent",
    "totaldescent",
    "verticaloscillation",
    "waterestimated",
}

_NORMALISED_LAP_SOURCE_KEYS = {
    "ascent",
    "averagegradeadjustedspeed",
    "averagehr",
    "averagemovingspeed",
    "averagepower",
    "averageruncadence",
    "averagespeed",
    "avggradeadjustedspeed",
    "descent",
    "distance",
    "duration",
    "elapsedduration",
    "endelevation",
    "endaltitude",
    "endingelevation",
    "elevationgain",
    "elevationloss",
    "gradeadjustedspeed",
    "groundcontacttime",
    "intensitytype",
    "lapindex",
    "maxelevation",
    "maxhr",
    "maximumcadenceinstepsperminute",
    "maximumelevation",
    "maxpower",
    "maxspeed",
    "messageindex",
    "minaltitude",
    "minelevation",
    "minimumelevation",
    "movingduration",
    "startelevation",
    "startaltitude",
    "startingelevation",
    "starttimelocal",
    "stridelength",
    "totalascent",
    "totaldescent",
    "type",
    "verticaloscillation",
}

_NORMALISED_ZONE_SOURCE_KEYS = {
    "lowboundary",
    "seconds",
    "secsinzone",
    "zone",
    "zonelowboundary",
    "zonenumber",
}

_NORMALISED_GEAR_SOURCE_KEYS = {
    "brand",
    "custommakemodel",
    "displayname",
    "distance",
    "gearid",
    "gearmakename",
    "gearmodelname",
    "gearname",
    "gearpk",
    "gearstatusname",
    "geartype",
    "geartypename",
    "gearuuid",
    "id",
    "make",
    "makemodel",
    "manufacturer",
    "model",
    "name",
    "productname",
    "retired",
    "retiredind",
    "status",
    "totaldistance",
    "totaldistancemeters",
    "type",
    "uuid",
}


def _source_extras(data, normalised_keys):
    """Conserva solo campos que no tienen ya una representación semántica."""
    if not isinstance(data, dict):
        return {}
    return {
        key: value
        for key, value in data.items()
        if re.sub(r"[^a-z0-9]", "", str(key).lower()) not in normalised_keys
    }


def _activity_summary_extras(data):
    """Retira del resumen los valores ya normalizados sin perder novedades."""
    if not isinstance(data, dict):
        return {}
    result = {}
    for key, value in data.items():
        normal = re.sub(r"[^a-z0-9]", "", str(key).lower())
        if normal in {"activitytype", "activitytypedto"}:
            extras = _source_extras(value, {"key", "name", "typekey"})
            if extras:
                result[key] = extras
            continue
        if normal == "eventtype":
            extras = _source_extras(value, {"key", "name", "typekey"})
            if extras:
                result[key] = extras
            continue
        if normal in _NORMALISED_ACTIVITY_SOURCE_KEYS:
            continue
        result[key] = value
    return result


def _activity_detail_extras(data):
    """Retira resumen, tipo y geometría que ya tienen salida normalizada."""
    if not isinstance(data, dict):
        return {}
    result = {}
    for key, value in data.items():
        normal = re.sub(r"[^a-z0-9]", "", str(key).lower())
        if normal == "summarydto":
            extras = _activity_summary_extras(value)
            if extras:
                result[key] = extras
            continue
        if normal == "activitytypedto":
            extras = _source_extras(value, {"key", "name", "typekey"})
            if extras:
                result[key] = extras
            continue
        if (
            "polyline" in normal
            or normal in {"encodedpath", "geometry", "map", "route", "track"}
        ):
            continue
        result[key] = value
    return result


def _row_collection_extras(data, row_key, normalised_keys):
    """Reduce una colección conservando únicamente extras por fila."""
    if not isinstance(data, dict):
        return {}
    result = {}
    target = re.sub(r"[^a-z0-9]", "", row_key.lower())
    for key, value in data.items():
        normal = re.sub(r"[^a-z0-9]", "", str(key).lower())
        if normal != target:
            result[key] = value
            continue
        if not isinstance(value, list):
            continue
        extras = [
            _source_extras(row, normalised_keys)
            for row in value
            if isinstance(row, dict)
        ]
        extras = [row for row in extras if row]
        if extras:
            result[key] = extras
    return result


def _zone_extras(zones):
    return [
        extras
        for zone in _as_list(zones)
        if isinstance(zone, dict)
        for extras in [_source_extras(zone, _NORMALISED_ZONE_SOURCE_KEYS)]
        if extras
    ]


def _gear_extras(items):
    return [
        extras
        for item in _as_list(items)
        if isinstance(item, dict)
        for extras in [_source_extras(item, _NORMALISED_GEAR_SOURCE_KEYS)]
        if extras
    ]


def _unmapped_activity_series(details, include_free_text=False):
    """Conserva solo columnas temporales que el esquema aún no reconoce."""
    if not isinstance(details, dict):
        return None
    descriptors = details.get("metricDescriptors")
    rows = details.get("activityDetailMetrics")
    if not isinstance(descriptors, list) or not isinstance(rows, list):
        return None

    selected = []
    output_descriptors = []
    for descriptor in descriptors:
        if not isinstance(descriptor, dict):
            continue
        raw_key = descriptor.get("key")
        if not isinstance(raw_key, str) or raw_key in _ACTIVITY_SERIES_KEYS:
            continue
        normal = re.sub(r"[^a-z0-9]", "", raw_key.lower())
        if is_personal_data_key(raw_key, ("details", "metricDescriptors")):
            continue
        if (
            not include_free_text
            and (
                "timestamp" in normal
                or normal.endswith("timegmt")
                or normal.endswith("timelocal")
            )
        ):
            continue
        index = descriptor.get("metricsIndex")
        if not isinstance(index, int) or index < 0:
            continue
        unit = descriptor.get("unit", {})
        selected.append(index)
        output_descriptors.append(_strip_empty({
            "source_field": raw_key,
            "source_unit": unit.get("key") if isinstance(unit, dict) else None,
            "source_factor": unit.get("factor") if isinstance(unit, dict) else None,
        }))

    samples = []
    for row in rows:
        metrics = row.get("metrics") if isinstance(row, dict) else None
        if not isinstance(metrics, list):
            continue
        if any(index >= len(metrics) for index in selected):
            continue
        samples.append([metrics[index] for index in selected])
    if not output_descriptors or not samples:
        return None
    return {
        "metric_descriptors": output_descriptors,
        "samples": samples,
    }


def _activity_unmapped_sport_data(
    activity_data,
    include_series=False,
    include_free_text=False,
):
    """Crea un delta deportivo sin repetir la representación normalizada."""
    if not isinstance(activity_data, dict):
        return None
    residual = {}

    summary = _activity_summary_extras(activity_data.get("summary"))
    if summary:
        residual["summary"] = summary

    detail = _activity_detail_extras(activity_data.get("detail"))
    if detail:
        residual["detail"] = detail

    splits = _row_collection_extras(
        activity_data.get("splits"),
        "lapDTOs",
        _NORMALISED_LAP_SOURCE_KEYS,
    )
    if splits:
        residual["splits"] = splits

    typed_splits = _row_collection_extras(
        activity_data.get("typed_splits"),
        "splits",
        _NORMALISED_LAP_SOURCE_KEYS,
    )
    if typed_splits:
        residual["typed_splits"] = typed_splits

    weather = _source_extras(
        activity_data.get("weather"),
        {"temp", "relativehumidity"},
    )
    if weather:
        residual["weather"] = weather

    hr_zones = _zone_extras(activity_data.get("hr_zones"))
    if hr_zones:
        residual["hr_zones"] = hr_zones

    power_zones = _zone_extras(activity_data.get("power_zones"))
    if power_zones:
        residual["power_zones"] = power_zones

    gear = _gear_extras(activity_data.get("gear"))
    if gear:
        residual["gear"] = gear

    if include_series:
        details = activity_data.get("details")
        temporal_series = _unmapped_activity_series(
            details,
            include_free_text=include_free_text,
        )
        details_extras = _source_extras(
            details,
            {"activitydetailmetrics", "metricdescriptors"},
        )
        if temporal_series or details_extras:
            residual["details"] = _strip_empty({
                "unmapped_activity_series": temporal_series,
                "metadata": details_extras,
            })

    known_sections = {
        "detail",
        "details",
        "gear",
        "hr_zones",
        "power_zones",
        "splits",
        "summary",
        "typed_splits",
        "weather",
    }
    for key, value in activity_data.items():
        if key not in known_sections:
            residual[key] = value

    return _strip_empty(_activity_source_data(
        residual,
        include_free_text=include_free_text,
    ))


def _source_metrics(data, *terms):
    """Devuelve campos originales relevantes sin inventar unidades."""
    if not isinstance(data, dict):
        return {}
    result = {}
    for key, value in data.items():
        normal = re.sub(r"[^a-z0-9]", "", str(key).lower())
        if any(term in normal for term in terms):
            result[str(key)] = value
    return result


def _route_geometry(data):
    """Recoge geometría de ruta con su ruta de origen, sin modificar valores."""
    found = {}

    def visit(value, path=""):
        if not isinstance(value, dict):
            return
        for key, nested in value.items():
            key_path = f"{path}.{key}".strip(".")
            normal = re.sub(r"[^a-z0-9]", "", str(key).lower())
            if (
                "polyline" in normal
                or normal in {"encodedpath", "geometry", "map", "route", "track"}
            ):
                found[key_path] = nested
            elif isinstance(nested, dict):
                visit(nested, key_path)

    visit(data)
    return found


def _compact_profile(data, period_end, timezone_name=None):
    """Crea el perfil reducido y privado del compacto semántico."""
    user_profile = data.get("user_profile") if isinstance(data, dict) else {}
    user_data = user_profile.get("userData", {}) if isinstance(user_profile, dict) else {}
    settings = data.get("profile_settings", {}) if isinstance(data, dict) else {}
    devices = data.get("devices", []) if isinstance(data, dict) else []
    primary_data = data.get("primary_device", {}) if isinstance(data, dict) else {}

    birth_date = _pick(user_data, "birthDate")
    age_years = None
    if isinstance(birth_date, str):
        try:
            born = date.fromisoformat(birth_date[:10])
            age_years = period_end.year - born.year
            if (period_end.month, period_end.day) < (born.month, born.day):
                age_years -= 1
        except ValueError:
            pass

    primary_device_id = None
    if isinstance(primary_data, dict):
        primary = primary_data.get("PrimaryTrainingDevice")
        if isinstance(primary, dict):
            primary_device_id = primary.get("deviceId")

    primary_watch = None
    if isinstance(devices, list):
        selected = None
        for device in devices:
            if not isinstance(device, dict):
                continue
            if primary_device_id is not None and device.get("deviceId") == primary_device_id:
                selected = device
                break
            if selected is None and (device.get("primary") or device.get("isPrimaryUser")):
                selected = device
        if selected is None and devices:
            selected = devices[0] if isinstance(devices[0], dict) else None
        if selected:
            primary_watch = _pick(
                selected,
                "productDisplayName",
                "productName",
            )

    return _strip_empty({
        "sex": str(_pick(user_data, "gender") or "").lower() or None,
        "age_years_at_period_end": age_years,
        "height_cm": _rounded(_pick(user_data, "height"), 1),
        "measurement_system": _pick(
            [settings, user_data, data],
            "measurementSystem",
            "unit_system",
        ),
        "timezone": timezone_name or _pick(settings, "timeZone"),
        "primary_watch": primary_watch,
    })


def _compact_sleep(
    sleep_data,
    timezone_name="Europe/Madrid",
    day_string=None,
    quality_callback=None,
):
    if not isinstance(sleep_data, dict):
        return None
    dto = sleep_data.get("dailySleepDTO")
    if not isinstance(dto, dict):
        return None

    scores = dto.get("sleepScores", {})
    overall = scores.get("overall", {}) if isinstance(scores, dict) else {}
    sleep_need = dto.get("sleepNeed", {})
    total_sleep = _number(dto.get("sleepTimeSeconds"))
    awake_sleep = _number(dto.get("awakeSleepSeconds"))
    start_epoch_ms = dto.get("sleepStartTimestampLocal")
    end_epoch_ms = dto.get("sleepEndTimestampLocal")
    start_local = epoch_ms_to_iso(start_epoch_ms, timezone_name)
    end_local = epoch_ms_to_iso(end_epoch_ms, timezone_name)
    garmin_date = day_string or dto.get("calendarDate") or "fecha desconocida"

    for field_name, raw_value, converted in (
        ("sleep_start_local", start_epoch_ms, start_local),
        ("sleep_end_local", end_epoch_ms, end_local),
    ):
        if raw_value is not None and converted is None and quality_callback:
            quality_callback(
                "temporal_warnings",
                f"{field_name} contiene un epoch en milisegundos inválido para "
                f"{garmin_date}; el campo se omitió.",
            )

    if start_local and end_local:
        sleep_window_s = (end_epoch_ms - start_epoch_ms) / 1000.0
        if sleep_window_s <= 0 and quality_callback:
            quality_callback(
                "temporal_warnings",
                f"La ventana de sueño no tiene orden temporal válido para {garmin_date}.",
            )
        measured_window_s = (
            total_sleep + awake_sleep
            if total_sleep is not None and awake_sleep is not None
            else None
        )
        if (
            sleep_window_s > 0
            and measured_window_s is not None
            and abs(sleep_window_s - measured_window_s) > 300
            and quality_callback
        ):
            quality_callback(
                "temporal_warnings",
                f"La ventana de sueño difiere más de cinco minutos de "
                f"total_sleep_s + awake_s para {garmin_date}.",
            )

    normalised_need = _normalise_sleep_need(
        sleep_need.get("actual") if isinstance(sleep_need, dict) else None
    )
    result = {
        "sleep_start_local": start_local,
        "sleep_end_local": end_local,
        "total_sleep_s": total_sleep,
        "awake_s": awake_sleep,
        "light_sleep_s": _number(dto.get("lightSleepSeconds")),
        "deep_sleep_s": _number(dto.get("deepSleepSeconds")),
        "rem_sleep_s": _number(dto.get("remSleepSeconds")),
        "unmeasurable_sleep_s": _number(dto.get("unmeasurableSleepSeconds")),
        "nap_time_s": _number(dto.get("napTimeSeconds")),
        "sleep_score": _number(overall.get("value")) if isinstance(overall, dict) else None,
        "sleep_score_qualifier": (
            overall.get("qualifierKey") if isinstance(overall, dict) else None
        ),
        "average_sleep_heart_rate_bpm": _number(dto.get("avgHeartRate")),
        "average_sleep_stress": _number(dto.get("avgSleepStress")),
        "average_sleep_spo2_pct": _number(
            _pick(dto, "averageSpO2Value", "averageSpO2HRSleep")
        ),
        "valid_sleep": total_sleep is not None and total_sleep > 0,
    }
    result.update(normalised_need)
    return _strip_empty(result)


def _debug_health_payload(kind, day_string, payload, source):
    """Registra forma y ausencia de payloads sin imprimir datos personales."""
    top_keys = sorted(payload.keys()) if isinstance(payload, dict) else []
    if kind == "sleep":
        dto = payload.get("dailySleepDTO") if isinstance(payload, dict) else None
        alignment = payload.get("sleepAlignment") if isinstance(payload, dict) else None
        has_real_data = (
            isinstance(dto, dict)
            and (_number(dto.get("sleepTimeSeconds")) or 0) > 0
        )
        reason = (
            alignment.get("alignmentStatus")
            if isinstance(alignment, dict)
            else None
        )
        method = "get_sleep_data"
    else:
        has_real_data = (
            isinstance(payload, dict)
            and isinstance(payload.get("hrvSummary"), dict)
        )
        reason = "missing_hrv_summary" if not has_real_data else "available"
        method = "get_hrv_data"
    log.debug(
        "Diagnóstico %s: method=%s date=%s source=%s has_data=%s "
        "top_keys=%s reason=%s",
        kind,
        method,
        day_string,
        source,
        has_real_data,
        top_keys,
        reason or "not_reported",
    )


def _compact_hrv(hrv_data):
    if not isinstance(hrv_data, dict):
        return None
    summary = hrv_data.get("hrvSummary")
    if not isinstance(summary, dict):
        return None
    baseline = summary.get("baseline", {})
    result = {
        "date": summary.get("calendarDate"),
        "overnight_average_ms": _number(summary.get("lastNightAvg")),
        "highest_five_min_average_ms": _number(summary.get("lastNight5MinHigh")),
        "weekly_average_ms": _number(summary.get("weeklyAvg")),
        "baseline_balanced_low_ms": (
            _number(baseline.get("balancedLow"))
            if isinstance(baseline, dict)
            else None
        ),
        "baseline_balanced_high_ms": (
            _number(baseline.get("balancedUpper"))
            if isinstance(baseline, dict)
            else None
        ),
        "status": summary.get("status"),
    }
    return _strip_empty(result)


def _compact_lifestyle_entries(lifestyle, include_free_text=False):
    """Conserva solo categorías deportivas seguras y explícitamente registradas."""
    if not isinstance(lifestyle, dict):
        return []
    safe_behaviours = {
        "illness": "illness",
        "sickness": "illness",
        "enfermedad": "illness",
        "maladie": "illness",
        "krankheit": "illness",
        "doenca": "illness",
        "injury": "injury",
        "lesion": "injury",
        "blessure": "injury",
        "verletzung": "injury",
        "lesao": "injury",
        "travel": "travel",
        "viaje": "travel",
        "voyage": "travel",
        "reise": "travel",
        "viagem": "travel",
        "alcohol": "alcohol",
        "alcool": "alcohol",
        "caffeine": "caffeine",
        "cafeina": "caffeine",
        "cafeine": "caffeine",
        "koffein": "caffeine",
        "nap": "nap",
        "siesta": "nap",
        "sieste": "nap",
        "nickerchen": "nap",
        "soneca": "nap",
    }
    result = []
    for item in _as_list(lifestyle.get("dailyLogsReport")):
        if not isinstance(item, dict):
            continue
        status = _pick(item, "logStatus", "status", "selected")
        amount = _pick(item, "quantity", "amount", "value", "measurementValue")
        note = _pick(item, "note", "notes", "comment")
        explicitly_logged = bool(status) or _number(amount) is not None or bool(note)
        if not explicitly_logged:
            continue
        behaviour_name = str(
            _pick(item, "key", "typeKey", "name") or ""
        )
        normal_name = unicodedata.normalize(
            "NFKD",
            behaviour_name.casefold(),
        )
        normal_name = "".join(
            character for character in normal_name
            if not unicodedata.combining(character)
        )
        normal_name = re.sub(r"[^a-z0-9]+", " ", normal_name).strip()
        safe_name = safe_behaviours.get(normal_name)
        if safe_name is None:
            continue
        result.append(_strip_empty({
            "date": item.get("calendarDate"),
            "behaviour": safe_name,
            "status": status,
            "amount": amount,
            "note": note if include_free_text else None,
        }))
    return result


def _compact_daily_record(
    day_string,
    data,
    timezone_name="Europe/Madrid",
    quality_callback=None,
    include_free_text=False,
):
    """Crea un registro diario semántico sin series horarias raw."""
    if not isinstance(data, dict):
        return {"date": day_string}
    summary = data.get("summary", {})
    heart_rate = data.get("heart_rate", {})
    stress = data.get("stress", {})
    spo2 = data.get("spo2", {})
    respiration = data.get("respiration", {})
    intensity = data.get("intensity_min", {})
    body_battery_list = _as_list(data.get("body_battery"))
    body_battery = (
        body_battery_list[0]
        if body_battery_list and isinstance(body_battery_list[0], dict)
        else {}
    )

    record = {
        "date": day_string,
        "steps": _number(_pick(summary, "totalSteps")),
        "distance_m": _number(
            _pick(summary, "totalDistanceMeters", "wellnessDistanceMeters")
        ),
        "active_calories_kcal": _number(
            _pick(summary, "activeKilocalories", "wellnessActiveKilocalories")
        ),
        "total_calories_kcal": _number(_pick(summary, "totalKilocalories")),
        "resting_heart_rate_bpm": _number(
            _pick([summary, heart_rate], "restingHeartRate")
        ),
        "seven_day_resting_heart_rate_bpm": _number(
            _pick([summary, heart_rate], "lastSevenDaysAvgRestingHeartRate")
        ),
        "average_stress": _number(
            _pick([summary, stress], "averageStressLevel", "avgStressLevel")
        ),
        "maximum_stress": _number(
            _pick([summary, stress], "maxStressLevel")
        ),
        "body_battery_high": _number(
            _pick(summary, "bodyBatteryHighestValue")
        ),
        "body_battery_low": _number(
            _pick(summary, "bodyBatteryLowestValue")
        ),
        "body_battery_charged": _number(
            _pick([summary, body_battery], "bodyBatteryChargedValue", "charged")
        ),
        "body_battery_drained": _number(
            _pick([summary, body_battery], "bodyBatteryDrainedValue", "drained")
        ),
        "average_spo2_pct": _number(
            _pick([spo2, summary], "averageSpO2", "averageSpo2")
        ),
        "lowest_spo2_pct": _number(
            _pick([spo2, summary], "lowestSpO2", "lowestSpo2")
        ),
        "average_waking_respiration_brpm": _number(
            _pick(
                [respiration, summary],
                "avgWakingRespirationValue",
            )
        ),
        "average_sleep_respiration_brpm": _number(
            _pick(respiration, "avgSleepRespirationValue")
        ),
        "moderate_intensity_minutes": _number(
            _pick([intensity, summary], "moderateMinutes", "moderateIntensityMinutes")
        ),
        "vigorous_intensity_minutes": _number(
            _pick([intensity, summary], "vigorousMinutes", "vigorousIntensityMinutes")
        ),
        "sleep": _compact_sleep(
            data.get("sleep"),
            timezone_name=timezone_name,
            day_string=day_string,
            quality_callback=quality_callback,
        ),
        "hrv": _compact_hrv(data.get("hrv")),
        "lifestyle_logs": _compact_lifestyle_entries(
            data.get("lifestyle"),
            include_free_text=include_free_text,
        ),
    }
    return _strip_empty(record)


def _normalise_zones(zones, boundary_name):
    result = []
    for zone in _as_list(zones):
        if not isinstance(zone, dict):
            continue
        result.append(_strip_empty({
            "zone": _number(_pick(zone, "zoneNumber", "zone")),
            "duration_s": _number(_pick(zone, "secsInZone", "seconds")),
            boundary_name: _number(
                _pick(zone, "zoneLowBoundary", "lowBoundary")
            ),
        }))
    return [zone for zone in result if zone]


def _mark_partial_last_lap(laps):
    """Marca solo una última vuelta claramente truncada respecto a las anteriores."""
    if not isinstance(laps, list) or len(laps) < 4:
        return laps
    last = laps[-1]
    previous = laps[:-1]
    distances = [
        _number(lap.get("distance_m"))
        for lap in previous
        if isinstance(lap, dict)
    ]
    if len(distances) < 3 or any(value is None or value <= 0 for value in distances):
        return laps
    ordered = sorted(distances)
    middle = len(ordered) // 2
    median = (
        ordered[middle]
        if len(ordered) % 2
        else (ordered[middle - 1] + ordered[middle]) / 2.0
    )
    if median <= 0 or any(abs(value - median) / median > 0.10 for value in distances):
        return laps
    last_distance = _number(last.get("distance_m"))
    previous_types = {lap.get("step_type") for lap in previous}
    same_structure = len(previous_types) == 1 and last.get("step_type") in previous_types
    if (
        last_distance is not None
        and 0 < last_distance < median * 0.5
        and same_structure
    ):
        last["partial_lap"] = True
    return laps


def _normalise_laps(activity_data, include_exact_times=False):
    splits = activity_data.get("splits", {}) if isinstance(activity_data, dict) else {}
    lap_rows = splits.get("lapDTOs") if isinstance(splits, dict) else None
    source_name = "splits.lapDTOs"
    if not isinstance(lap_rows, list) or not lap_rows:
        typed = activity_data.get("typed_splits", {}) if isinstance(activity_data, dict) else {}
        lap_rows = typed.get("splits") if isinstance(typed, dict) else []
        source_name = "typed_splits.splits"

    result = []
    for index, lap in enumerate(_as_list(lap_rows), 1):
        if not isinstance(lap, dict):
            continue
        speed = _pick(lap, "averageMovingSpeed", "averageSpeed")
        max_speed = _pick(lap, "maxSpeed")
        gap_speed = _pick(
            lap,
            "averageGradeAdjustedSpeed",
            "avgGradeAdjustedSpeed",
            "gradeAdjustedSpeed",
        )
        elevation_gain = _number(_pick(lap, "elevationGain", "totalAscent", "ascent"))
        elevation_loss = _number(_pick(lap, "elevationLoss", "totalDescent", "descent"))
        start_elevation = _number(
            _pick(lap, "startElevation", "startingElevation", "startAltitude")
        )
        end_elevation = _number(
            _pick(lap, "endElevation", "endingElevation", "endAltitude")
        )
        elevation_net = (
            end_elevation - start_elevation
            if start_elevation is not None and end_elevation is not None
            else (
                elevation_gain - elevation_loss
                if elevation_gain is not None and elevation_loss is not None
                else None
            )
        )
        result.append(_strip_empty({
            "lap_index": _number(_pick(lap, "lapIndex", "messageIndex")) or index,
            "lap_type": _pick(lap, "type"),
            "step_type": _pick(lap, "intensityType"),
            "start_time_local": (
                _pick(lap, "startTimeLocal") if include_exact_times else None
            ),
            "distance_m": _number(_pick(lap, "distance")),
            "duration_s": _number(_pick(lap, "duration")),
            "elapsed_duration_s": _number(_pick(lap, "elapsedDuration")),
            "moving_duration_s": _number(_pick(lap, "movingDuration")),
            "average_pace_s_per_km": _pace_from_speed(speed),
            "best_pace_s_per_km": _pace_from_speed(max_speed),
            "average_heart_rate_bpm": _number(_pick(lap, "averageHR")),
            "maximum_heart_rate_bpm": _number(_pick(lap, "maxHR")),
            "average_power_w": _number(_pick(lap, "averagePower")),
            "maximum_power_w": _number(_pick(lap, "maxPower")),
            "average_cadence_spm": _number(_pick(lap, "averageRunCadence")),
            "average_stride_length_cm": _number(_pick(lap, "strideLength")),
            "average_ground_contact_time_ms": _number(
                _pick(lap, "groundContactTime")
            ),
            "average_vertical_oscillation_cm": _number(
                _pick(lap, "verticalOscillation")
            ),
            "elevation_gain_m": elevation_gain,
            "elevation_loss_m": elevation_loss,
            "elevation_net_change_m": elevation_net,
            "elevation_net_change_method": (
                "end_minus_start"
                if start_elevation is not None and end_elevation is not None
                else "ascent_minus_descent"
                if elevation_net is not None
                else None
            ),
            "start_elevation_m": start_elevation,
            "end_elevation_m": end_elevation,
            "minimum_elevation_m": _number(
                _pick(lap, "minElevation", "minimumElevation", "minAltitude")
            ),
            "maximum_elevation_m": _number(
                _pick(lap, "maxElevation", "maximumElevation", "maxAltitude")
            ),
            "grade_adjusted_pace_s_per_km": _pace_from_speed(gap_speed),
            "grade_adjusted_pace_source": _source_metrics(
                lap,
                "gradeadjusted",
                "gap",
                "rap",
            ),
        }))
    return _mark_partial_last_lap(result), source_name


_ACTIVITY_SERIES_KEYS = {
    "sumElapsedDuration": "elapsed_duration_raw",
    "sumDuration": "duration_raw",
    "sumDistance": "distance_raw",
    "directHeartRate": "heart_rate_raw",
    "directSpeed": "speed_raw",
    "directPower": "power_raw",
    "directRunCadence": "running_cadence_raw",
    "directDoubleCadence": "cadence_raw",
    "directGroundContactTime": "ground_contact_time_raw",
    "directVerticalOscillation": "vertical_oscillation_raw",
    "directStrideLength": "stride_length_raw",
    "directElevation": "elevation_raw",
    "directEnhancedElevation": "elevation_raw",
    "directGrade": "grade_raw",
    "directGradeAdjustedPace": "grade_adjusted_pace_raw",
    "directGradeAdjustedSpeed": "grade_adjusted_speed_raw",
    "directLatitude": "latitude_deg",
    "directLongitude": "longitude_deg",
    "directTemperature": "temperature_raw",
}
_MAX_ACTIVITY_CHART_POINTS = 100_000


def _compact_activity_series(details, diagnostics=None):
    """Conserva series deportivas aprobadas, incluido el GPS disponible."""
    diagnostics = diagnostics if diagnostics is not None else []
    if not isinstance(details, dict):
        return None
    descriptors = details.get("metricDescriptors")
    rows = details.get("activityDetailMetrics")
    if not isinstance(descriptors, list) or not isinstance(rows, list):
        return None

    selected = []
    output_descriptors = []
    for descriptor in descriptors:
        if not isinstance(descriptor, dict):
            continue
        raw_key = descriptor.get("key")
        if raw_key not in _ACTIVITY_SERIES_KEYS:
            continue
        index = descriptor.get("metricsIndex")
        if not isinstance(index, int) or index < 0:
            diagnostics.append(
                "Se omitió un descriptor temporal con índice no válido."
            )
            continue
        unit = descriptor.get("unit", {})
        selected.append(index)
        output_descriptors.append(_strip_empty({
            "field": _ACTIVITY_SERIES_KEYS[raw_key],
            "source_field": raw_key,
            "source_unit": unit.get("key") if isinstance(unit, dict) else None,
            "source_factor": unit.get("factor") if isinstance(unit, dict) else None,
        }))

    samples = []
    for row_index, row in enumerate(rows):
        metrics = row.get("metrics") if isinstance(row, dict) else None
        if not isinstance(metrics, list):
            diagnostics.append(
                f"La muestra temporal {row_index} no contiene una matriz metrics válida."
            )
            continue
        if any(index >= len(metrics) for index in selected):
            diagnostics.append(
                f"La muestra temporal {row_index} está truncada y se omitió."
            )
            continue
        sample = [metrics[index] for index in selected]
        samples.append(sample)
    if not output_descriptors or not samples:
        return None
    series = {
        "metric_descriptors": output_descriptors,
        "samples": samples,
    }
    expected = len(output_descriptors)
    if any(len(sample) != expected for sample in samples):
        diagnostics.append(
            "La serie temporal se omitió porque contiene muestras desalineadas."
        )
        return None
    return series


def _activity_hr_distribution(details, zones, duration_s, average_hr):
    """Separa pulso válido, zona 0, pulso no clasificado y falta de sensor."""
    result = [dict(zone) for zone in zones]
    classified_s = sum(
        _number(zone.get("duration_s")) or 0
        for zone in result
        if (_number(zone.get("zone")) or 0) >= 1
    )
    active_s = 0.0
    valid_hr_s = 0.0
    measured_below_zone_one_s = 0.0
    source = None
    zone_one = next(
        (zone for zone in result if _number(zone.get("zone")) == 1),
        {},
    )
    zone_one_low = _number(zone_one.get("low_boundary_bpm"))

    descriptors = details.get("metricDescriptors") if isinstance(details, dict) else None
    rows = details.get("activityDetailMetrics") if isinstance(details, dict) else None
    indexes = {}
    if isinstance(descriptors, list):
        for descriptor in descriptors:
            if isinstance(descriptor, dict) and descriptor.get("key") in {
                "sumDuration", "directHeartRate"
            }:
                indexes[descriptor.get("key")] = descriptor.get("metricsIndex")

    duration_index = indexes.get("sumDuration")
    hr_index = indexes.get("directHeartRate")
    if (
        isinstance(rows, list)
        and isinstance(duration_index, int)
        and isinstance(hr_index, int)
    ):
        previous_duration = None
        for row in rows:
            metrics = row.get("metrics") if isinstance(row, dict) else None
            if not isinstance(metrics, list) or duration_index >= len(metrics):
                continue
            current_duration = _number(metrics[duration_index])
            if current_duration is None:
                continue
            if previous_duration is not None:
                delta = current_duration - previous_duration
                if 0 <= delta <= 120:
                    active_s += delta
                    heart_rate = (
                        _number(metrics[hr_index])
                        if hr_index < len(metrics)
                        else None
                    )
                    if heart_rate is not None and heart_rate > 0:
                        valid_hr_s += delta
                        if zone_one_low is not None and heart_rate < zone_one_low:
                            measured_below_zone_one_s += delta
            previous_duration = current_duration
        if active_s > 0:
            source = "activity_series"

    if source is None and _number(average_hr) is not None and _number(duration_s):
        active_s = _number(duration_s)
        valid_hr_s = active_s
        source = "activity_duration_estimate"
    elif source is None and classified_s > 0:
        active_s = classified_s
        valid_hr_s = classified_s
        source = "garmin_zones_only"

    remaining_valid_s = max(0.0, valid_hr_s - classified_s)
    below_zone_one_s = (
        min(measured_below_zone_one_s, remaining_valid_s)
        if source == "activity_series" and zone_one_low is not None
        else 0.0
    )
    valid_unclassified_s = max(
        0.0,
        valid_hr_s - classified_s - below_zone_one_s,
    )

    if valid_hr_s > 0 and source == "activity_series" and zone_one_low is not None:
        result.insert(0, _strip_empty({
            "zone": 0,
            "duration_s": below_zone_one_s,
            "label": "below_zone_1",
            "upper_boundary_bpm": zone_one_low,
        }))
    denominator = sum(_number(zone.get("duration_s")) or 0 for zone in result)
    if denominator:
        for zone in result:
            seconds = _number(zone.get("duration_s")) or 0
            zone["percentage"] = seconds * 100.0 / denominator

    missing_s = max(0.0, active_s - valid_hr_s) if active_s > 0 else None
    paused_s = (
        max(0.0, _number(duration_s) - active_s)
        if _number(duration_s) is not None and active_s > 0
        else None
    )
    return result, _strip_empty({
        "heart_rate_duration_source": source,
        "active_duration_from_series_s": active_s if source == "activity_series" else None,
        "valid_heart_rate_duration_s": valid_hr_s if valid_hr_s > 0 else None,
        "missing_heart_rate_duration_s": missing_s,
        "valid_hr_unclassified_duration_s": (
            valid_unclassified_s if valid_hr_s > 0 else None
        ),
        "paused_duration_s": paused_s,
        "classified_zones_1_5_duration_s": classified_s if classified_s > 0 else None,
        "below_zone_1_duration_s": below_zone_one_s if valid_hr_s > 0 else None,
        "heart_rate_zone_coverage_pct": (
            min(
                100.0,
                (classified_s + below_zone_one_s) * 100.0 / valid_hr_s,
            )
            if valid_hr_s > 0
            else None
        ),
        "garmin_zones_1_5_coverage_pct": (
            min(100.0, classified_s * 100.0 / valid_hr_s)
            if valid_hr_s > 0
            else None
        ),
    })


_GENERIC_GEAR_TEXT = {
    "desconocido",
    "desconocida",
    "ninguno",
    "ninguna",
    "none",
    "not set",
    "other",
    "otra",
    "otro",
    "otros",
    "unknown",
}


def _gear_text(item, *keys, omit_generic=False):
    """Devuelve el primer texto útil de un campo de equipamiento."""
    if not isinstance(item, dict):
        return None
    for key in keys:
        value = item.get(key)
        if not isinstance(value, str):
            continue
        value = " ".join(value.split())
        if not value:
            continue
        normal = unicodedata.normalize("NFKD", value.casefold())
        normal = "".join(
            character
            for character in normal
            if not unicodedata.combining(character)
        )
        normal = re.sub(r"[^a-z0-9]+", " ", normal).strip()
        if omit_generic and normal in _GENERIC_GEAR_TEXT:
            continue
        return value
    return None


def _compact_gear_items(
    items,
    stats_by_id=None,
    reference_secret=None,
    include_free_text=False,
):
    stats_by_id = stats_by_id or {}
    result = []
    for item in _as_list(items):
        if not isinstance(item, dict):
            continue
        gear_id = _pick(item, "uuid", "gearUUID", "gearId", "gearPk", "id")
        stats = stats_by_id.get(str(gear_id), {}) if gear_id is not None else {}
        if not isinstance(stats, dict):
            stats = {}
        explicit_name = _gear_text(
            item,
            "displayName",
            "gearName",
            "name",
            omit_generic=True,
        )
        manufacturer = _gear_text(
            item,
            "gearMakeName",
            "manufacturer",
            "brand",
            "make",
            omit_generic=True,
        )
        custom_model = _gear_text(
            item,
            "customMakeModel",
            omit_generic=True,
        )
        catalog_model = _gear_text(
            item,
            "gearModelName",
            "productName",
            "makeModel",
            "model",
            omit_generic=True,
        )
        model = custom_model or catalog_model
        gear_name = explicit_name
        if gear_name is None:
            gear_name = " ".join(
                value
                for value in (manufacturer, model)
                if value
            ) or None
        gear_name_user_provided = (
            True
            if explicit_name is not None
            else (custom_model is not None if gear_name is not None else None)
        )
        model_user_provided = (
            custom_model is not None if model is not None else None
        )
        reference_material = gear_id
        if reference_material is None:
            reference_material = "|".join(
                str(value or "")
                for value in (
                    model,
                    _pick(item, "gearTypeName", "gearType", "type"),
                    gear_name,
                )
            )
        total_distance = _pick(
            [item, stats],
            "totalDistance",
            "totalDistanceMeters",
            "distance",
        )
        result.append(_strip_empty({
            "gear_ref": private_reference(
                "gear",
                reference_material,
                reference_secret,
            ),
            "gear_name": gear_name,
            "manufacturer": manufacturer,
            "model": model,
            "gear_name_user_provided": gear_name_user_provided,
            "model_user_provided": model_user_provided,
            "type": _pick(item, "gearTypeName", "gearType", "type"),
            "status": _pick(item, "gearStatusName", "status"),
            "total_distance_m": _number(total_distance),
            "retired": _pick(item, "retired", "retiredInd"),
        }))
    return [item for item in result if item]


def _enrich_activity_gear_from_catalog(activities, catalog):
    """Completa asociaciones reducidas con la identidad del catálogo global."""
    catalog_by_reference = {
        item.get("gear_ref"): item
        for item in _as_list(catalog)
        if isinstance(item, dict) and item.get("gear_ref")
    }
    identity_fields = (
        "gear_name",
        "manufacturer",
        "model",
        "gear_name_user_provided",
        "model_user_provided",
        "type",
    )
    for activity in _as_list(activities):
        if not isinstance(activity, dict):
            continue
        for association in _as_list(activity.get("gear")):
            if not isinstance(association, dict):
                continue
            catalog_item = catalog_by_reference.get(
                association.get("gear_ref")
            )
            if not catalog_item:
                continue
            for field in identity_fields:
                if association.get(field) in (None, ""):
                    value = catalog_item.get(field)
                    if value not in (None, ""):
                        association[field] = value
    return activities


def _activity_time_bucket(start_local):
    """Conserva una franja útil sin publicar la hora exacta."""
    if not isinstance(start_local, str):
        return None
    match = re.search(r"[T ](\d{2}):", start_local)
    if not match:
        return None
    hour = int(match.group(1))
    if 5 <= hour < 12:
        return "morning"
    if 12 <= hour < 18:
        return "afternoon"
    if 18 <= hour < 23:
        return "evening"
    return "night"


def _compact_activity(
    activity_data,
    include_series=False,
    quality_callback=None,
    reference_secret=None,
    include_free_text=False,
):
    if not isinstance(activity_data, dict):
        return None
    summary = activity_data.get("summary", {})
    detail = activity_data.get("detail", {})
    dto = detail.get("summaryDTO", {}) if isinstance(detail, dict) else {}
    activity_type = (
        summary.get("activityType", {})
        if isinstance(summary, dict)
        else {}
    )
    if not isinstance(activity_type, dict) and isinstance(detail, dict):
        activity_type = detail.get("activityTypeDTO", {})
    if not isinstance(activity_type, dict):
        activity_type = {}
    sources = [summary, dto]
    speed = _pick(sources, "averageSpeed")
    max_speed = _pick(sources, "maxSpeed")
    start_local = _pick(sources, "startTimeLocal")
    event_type = _pick(sources, "eventType", "eventTypeKey")
    if isinstance(event_type, dict):
        event_type = _pick(event_type, "typeKey", "key", "name")
    laps, lap_source = _normalise_laps(
        activity_data,
        include_exact_times=include_free_text,
    )
    duration_s = _number(_pick(sources, "duration"))
    average_hr = _number(_pick(sources, "averageHR"))
    hr_zones = _normalise_zones(
        activity_data.get("hr_zones"),
        "low_boundary_bpm",
    )
    hr_zones, hr_distribution = _activity_hr_distribution(
        activity_data.get("details"),
        hr_zones,
        duration_s,
        average_hr,
    )

    weather = activity_data.get("weather", {})
    weather = weather if isinstance(weather, dict) else {}
    device_temperature = _pick(sources, "averageTemperature")
    if device_temperature is not None:
        temperature = _normalise_temperature(device_temperature, "celsius")
    else:
        temperature = _normalise_temperature(weather.get("temp"), "fahrenheit")
    maximum_temperature = _normalise_temperature(
        _pick(sources, "maxTemperature"),
        "celsius",
    )

    diagnostics = []
    activity_series = (
        _compact_activity_series(
            activity_data.get("details"),
            diagnostics,
        )
        if include_series
        else None
    )
    if quality_callback:
        for message in diagnostics:
            quality_callback("series_validation_errors", message)

    start_elevation = _number(
        _pick(sources, "startElevation", "startingElevation", "startAltitude")
    )
    end_elevation = _number(
        _pick(sources, "endElevation", "endingElevation", "endAltitude")
    )
    elevation_gain = _number(_pick(sources, "elevationGain", "totalAscent"))
    elevation_loss = _number(_pick(sources, "elevationLoss", "totalDescent"))
    elevation_net = (
        end_elevation - start_elevation
        if start_elevation is not None and end_elevation is not None
        else (
            elevation_gain - elevation_loss
            if elevation_gain is not None and elevation_loss is not None
            else None
        )
    )
    gap_speed = _pick(
        sources,
        "averageGradeAdjustedSpeed",
        "avgGradeAdjustedSpeed",
        "gradeAdjustedSpeed",
    )
    result = {
        "activity_ref": private_reference(
            "activity",
            _pick(summary, "activityId"),
            reference_secret,
        ),
        "date": start_local[:10] if isinstance(start_local, str) else None,
        "start_time_bucket": _activity_time_bucket(start_local),
        "start_time_local": start_local if include_free_text else None,
        "sport": _pick(activity_type, "typeKey"),
        "garmin_event_type": event_type,
        "name": _pick(sources, "activityName", "name", "title"),
        "description": (
            _pick(sources, "description", "activityDescription", "note", "notes")
            if include_free_text
            else None
        ),
        "distance_m": _number(_pick(sources, "distance")),
        "duration_s": duration_s,
        "elapsed_duration_s": _number(_pick(sources, "elapsedDuration")),
        "moving_duration_s": _number(_pick(sources, "movingDuration")),
        "average_pace_s_per_km": _pace_from_speed(speed),
        "best_pace_s_per_km": _pace_from_speed(max_speed),
        "average_speed_m_s": _number(speed),
        "maximum_speed_m_s": _number(max_speed),
        "average_heart_rate_bpm": average_hr,
        "maximum_heart_rate_bpm": _number(_pick(sources, "maxHR")),
        "average_power_w": _number(
            _pick(sources, "avgPower", "averagePower")
        ),
        "maximum_power_w": _number(_pick(sources, "maxPower")),
        "normalized_power_w": _number(
            _pick(sources, "normPower", "normalizedPower")
        ),
        "average_cadence_spm": _number(
            _pick(sources, "averageRunningCadenceInStepsPerMinute", "averageRunCadence")
        ),
        "maximum_cadence_spm": _number(
            _pick(sources, "maxRunningCadenceInStepsPerMinute", "maxRunCadence")
        ),
        "average_stride_length_cm": _number(_pick(sources, "strideLength")),
        "average_ground_contact_time_ms": _number(
            _pick(sources, "groundContactTime")
        ),
        "average_vertical_oscillation_cm": _number(
            _pick(sources, "verticalOscillation")
        ),
        "elevation_gain_m": elevation_gain,
        "elevation_loss_m": elevation_loss,
        "elevation_net_change_m": elevation_net,
        "elevation_net_change_method": (
            "end_minus_start"
            if start_elevation is not None and end_elevation is not None
            else "ascent_minus_descent"
            if elevation_net is not None
            else None
        ),
        "start_elevation_m": start_elevation,
        "end_elevation_m": end_elevation,
        "minimum_elevation_m": _number(
            _pick(sources, "minElevation", "minimumElevation", "minAltitude")
        ),
        "maximum_elevation_m": _number(
            _pick(sources, "maxElevation", "maximumElevation", "maxAltitude")
        ),
        "grade_adjusted_pace_s_per_km": _pace_from_speed(gap_speed),
        "grade_adjusted_pace_source": {
            **_source_metrics(summary, "gradeadjusted", "gap", "rap"),
            **_source_metrics(dto, "gradeadjusted", "gap", "rap"),
        },
        "coordinates": _strip_empty({
            "start": _strip_empty({
                "latitude": _number(_pick(sources, "startLatitude", "startLat")),
                "longitude": _number(_pick(sources, "startLongitude", "startLon")),
            }),
            "end": _strip_empty({
                "latitude": _number(_pick(sources, "endLatitude", "endLat")),
                "longitude": _number(_pick(sources, "endLongitude", "endLon")),
            }),
        }),
        "route_geometry": _activity_source_data(
            _route_geometry({
                key: value
                for key, value in activity_data.items()
                if key != "details"
            }),
            include_free_text=include_free_text,
        ),
        "average_temperature_c": temperature.get("temperature_c"),
        "average_temperature_raw": temperature.get("temperature_raw"),
        "average_temperature_source_unit": temperature.get(
            "temperature_source_unit"
        ),
        "maximum_temperature_c": maximum_temperature.get("temperature_c"),
        "maximum_temperature_raw": maximum_temperature.get("temperature_raw"),
        "maximum_temperature_source_unit": maximum_temperature.get(
            "temperature_source_unit"
        ),
        "humidity_pct": _number(
            _pick(activity_data.get("weather", {}), "relativeHumidity")
        ),
        "aerobic_training_effect": _number(
            _pick(sources, "aerobicTrainingEffect", "trainingEffect")
        ),
        "anaerobic_training_effect": _number(
            _pick(sources, "anaerobicTrainingEffect")
        ),
        "training_load": _number(_pick(sources, "activityTrainingLoad")),
        "calories_kcal": _number(_pick(sources, "calories")),
        "estimated_sweat_loss_ml": _number(_pick(sources, "waterEstimated")),
        "self_evaluation": _normalise_self_evaluation(dto),
        "hr_zones": hr_zones,
        "heart_rate_distribution_quality": hr_distribution,
        "power_zones": _normalise_zones(
            activity_data.get("power_zones"),
            "low_boundary_w",
        ),
        "laps": laps,
        "lap_source": lap_source if laps else None,
        "gear": _compact_gear_items(
            activity_data.get("gear"),
            reference_secret=reference_secret,
            include_free_text=include_free_text,
        ),
        "activity_series": activity_series,
        "unmapped_sport_data": _activity_unmapped_sport_data(
            activity_data,
            include_series=include_series,
            include_free_text=include_free_text,
        ),
    }
    return _strip_empty(result)


def _find_blood_pressure_measurements(data):
    found = []

    def visit(value):
        if isinstance(value, dict):
            systolic = _number(_pick(value, "systolic", "systolicValue"))
            diastolic = _number(_pick(value, "diastolic", "diastolicValue"))
            if systolic is not None and diastolic is not None:
                found.append(_strip_empty({
                    "timestamp": _pick(
                        value,
                        "measurementTimestampLocal",
                        "timestampLocal",
                        "measurementTime",
                        "calendarDate",
                        "date",
                    ),
                    "systolic_mmhg": systolic,
                    "diastolic_mmhg": diastolic,
                    "pulse_bpm": _number(
                        _pick(value, "pulse", "heartRate", "pulseRate")
                    ),
                }))
                return
            for nested in value.values():
                visit(nested)
        elif isinstance(value, list):
            for nested in value:
                visit(nested)

    visit(data)
    unique = []
    seen = set()
    for measurement in found:
        key = (
            measurement.get("timestamp"),
            measurement.get("systolic_mmhg"),
            measurement.get("diastolic_mmhg"),
            measurement.get("pulse_bpm"),
        )
        if key not in seen:
            seen.add(key)
            unique.append(measurement)
    return unique


def _compact_body_composition(data):
    if not isinstance(data, dict):
        return []
    raw_measurements = []
    for chunk in _as_list(data.get("body_comp")):
        if isinstance(chunk, dict):
            raw_measurements.extend(_as_list(chunk.get("dateWeightList")))
    if not raw_measurements:
        for chunk in _as_list(data.get("weigh_ins")):
            if not isinstance(chunk, dict):
                continue
            for daily in _as_list(chunk.get("dailyWeightSummaries")):
                if isinstance(daily, dict):
                    latest = daily.get("latestWeight")
                    if isinstance(latest, dict):
                        raw_measurements.append(latest)

    result = []
    seen = set()
    for measurement in raw_measurements:
        if not isinstance(measurement, dict):
            continue
        timestamp = _pick(
            measurement,
            "timestampGMT",
            "calendarDate",
            "date",
        )
        weight_g = _number(measurement.get("weight"))
        item = _strip_empty({
            "date": (
                timestamp[:10]
                if isinstance(timestamp, str)
                else measurement.get("calendarDate")
            ),
            "weight_kg": (
                round(weight_g / 1000.0, 3)
                if weight_g is not None
                else None
            ),
            "bmi": _number(measurement.get("bmi")),
            "body_fat_pct": _number(measurement.get("bodyFat")),
            "body_water_pct": _number(measurement.get("bodyWater")),
            "muscle_mass_kg": (
                round(measurement["muscleMass"] / 1000.0, 3)
                if _number(measurement.get("muscleMass")) is not None
                else None
            ),
            "bone_mass_kg": (
                round(measurement["boneMass"] / 1000.0, 3)
                if _number(measurement.get("boneMass")) is not None
                else None
            ),
        })
        identity = (item.get("date"), item.get("weight_kg"))
        if item and identity not in seen:
            seen.add(identity)
            result.append(item)
    return result


_EFFECTIVE_DATE_KEYS = {
    "calendardate",
    "date",
    "metriccalendardate",
    "measurementdate",
    "valuedate",
}


def _metric_dates(data):
    """Obtiene fechas efectivas priorizando los campos de fecha de calendario."""
    preferred = []
    fallback = []

    def add_date(value, target):
        if not isinstance(value, str):
            return
        match = re.search(r"\d{4}-\d{2}-\d{2}", value)
        if not match:
            return
        try:
            target.append(date.fromisoformat(match.group(0)))
        except ValueError:
            pass

    def visit(value):
        if isinstance(value, dict):
            for key, nested in value.items():
                normal_key = re.sub(r"[^a-z0-9]", "", str(key).lower())
                if normal_key in _EFFECTIVE_DATE_KEYS:
                    add_date(nested, preferred)
                elif "date" in normal_key or "timestamp" in normal_key:
                    add_date(nested, fallback)
                visit(nested)
        elif isinstance(value, list):
            for nested in value:
                visit(nested)

    visit(data)
    selected = preferred or fallback
    return sorted(set(selected))


def _compact_readiness(value):
    entries = [item for item in _as_list(value) if isinstance(item, dict)]
    if not entries:
        return None
    selected = next(
        (
            item for item in entries
            if item.get("inputContext") == "AFTER_WAKEUP_RESET"
        ),
        entries[0],
    )
    return _strip_empty({
        "date": selected.get("calendarDate"),
        "timestamp_local": selected.get("timestampLocal"),
        "score": _number(selected.get("score")),
        "level": selected.get("level"),
        "acute_load": _number(selected.get("acuteLoad")),
        "recovery_time_raw": _number(selected.get("recoveryTime")),
        "sleep_score": _number(selected.get("sleepScore")),
        "hrv_weekly_average_raw": _number(selected.get("hrvWeeklyAverage")),
        "valid_sleep": selected.get("validSleep"),
    })


def _normalise_training_metric(name, value):
    if value is None:
        return None
    if name in ("training_readiness", "morning_readiness"):
        return _compact_readiness(value)
    if name == "lactate_threshold" and isinstance(value, dict):
        speed_hr = value.get("speed_and_heart_rate", {})
        power = value.get("power", {})
        speed = _pick(speed_hr, "speed")
        result = {
            "date": _pick([speed_hr, power], "calendarDate"),
            "heart_rate_bpm": _number(
                _pick(speed_hr, "heartRate", "heartRateCycling")
            ),
            "power_w": _number(_pick(power, "functionalThresholdPower")),
            "power_to_weight_w_kg": _number(_pick(power, "powerToWeight")),
        }
        result.update(_normalise_lactate_speed(speed, "garmin_tenths_m_s"))
        return _strip_empty(result)
    if name == "cycling_ftp" and isinstance(value, dict):
        return _strip_empty({
            "date": value.get("calendarDate"),
            "power_w": _number(value.get("functionalThresholdPower")),
            "stale": value.get("isStale"),
        })
    if name == "race_predictions" and isinstance(value, dict):
        return _strip_empty({
            "date": value.get("calendarDate"),
            "time_5k_s": _number(value.get("time5K")),
            "time_10k_s": _number(value.get("time10K")),
            "time_half_marathon_s": _number(value.get("timeHalfMarathon")),
            "time_marathon_s": _number(value.get("timeMarathon")),
        })
    if name == "fitness_age" and isinstance(value, dict):
        return _strip_empty({
            "date": value.get("lastUpdated"),
            "fitness_age": _number(value.get("fitnessAge")),
            "chronological_age": _number(value.get("chronologicalAge")),
            "achievable_fitness_age": _number(value.get("achievableFitnessAge")),
        })
    if name == "intensity_min" and isinstance(value, dict):
        return _strip_empty({
            "date": value.get("calendarDate"),
            "moderate_minutes": _number(value.get("moderateMinutes")),
            "vigorous_minutes": _number(value.get("vigorousMinutes")),
            "weekly_total_minutes": _number(value.get("weeklyTotal")),
            "weekly_goal_minutes": _number(value.get("weekGoal")),
        })
    return _strip_empty(
        _sanitize_compact(value, remove_free_text=True)
    )


def _compact_training(data, period_start, period_end):
    result = {
        "period": {
            "start_date": period_start.isoformat(),
            "end_date": period_end.isoformat(),
        },
        "historical_period_data": {},
        "latest_before_or_within_period": {},
        "current_snapshot": {},
        "undated": {},
    }
    snapshots = []
    for name, value in (data or {}).items():
        if name.startswith("_title_") or value is None or value == [] or value == {}:
            continue

        # Las listas se clasifican registro a registro. Así una sola medición
        # posterior no arrastra por error todo el bloque histórico a snapshot.
        raw_records = value if isinstance(value, list) else [value]
        bucketed = {
            "historical_period_data": [],
            "latest_before_or_within_period": [],
            "current_snapshot": [],
            "undated": [],
        }
        dated_before = []
        for raw_record in raw_records:
            dates = _metric_dates(raw_record)
            normalised = _normalise_training_metric(name, raw_record)
            if normalised is None or normalised == {} or normalised == []:
                continue
            if dates and min(dates) > period_end:
                bucketed["current_snapshot"].append(normalised)
                snapshots.append({
                    "metric": name,
                    "effective_dates": [
                        metric_date.isoformat()
                        for metric_date in dates
                        if metric_date > period_end
                    ],
                })
            elif dates and any(metric_date > period_end for metric_date in dates):
                # Un objeto indivisible que mezcla fechas se conserva como
                # snapshot y queda documentado. Las listas ya se separaron.
                bucketed["current_snapshot"].append(normalised)
                snapshots.append({
                    "metric": name,
                    "effective_dates": [
                        metric_date.isoformat()
                        for metric_date in dates
                        if metric_date > period_end
                    ],
                })
            elif any(period_start <= metric_date <= period_end for metric_date in dates):
                bucketed["historical_period_data"].append(normalised)
            elif dates and max(dates) < period_start:
                dated_before.append((max(dates), normalised))
            else:
                bucketed["undated"].append(normalised)

        if dated_before:
            dated_before.sort(key=lambda item: item[0])
            bucketed["latest_before_or_within_period"].append(
                dated_before[-1][1]
            )

        for bucket_name, records in bucketed.items():
            if not records:
                continue
            result[bucket_name][name] = (
                records if isinstance(value, list) else records[0]
            )
    return _strip_empty(result), snapshots


def _compact_personal_records(data, include_free_text=False):
    if not isinstance(data, dict):
        return None
    records = []
    for item in _as_list(data.get("personal_records")):
        if not isinstance(item, dict):
            continue
        record_date = _pick(
            item,
            "activityStartDateTimeLocal",
            "prStartTimeLocal",
        )
        records.append(_strip_empty({
            "record_type": _pick(item, "prTypeLabelKey", "typeId"),
            "value": _number(item.get("value")),
            "date": (
                record_date[:10]
                if isinstance(record_date, str)
                else None
            ),
            "activity_name": (
                item.get("activityName") if include_free_text else None
            ),
        }))
    active_goals = _sanitize_compact(data.get("active_goals") or [])
    if not include_free_text:
        private_goal_text_keys = {
            "activityname",
            "comment",
            "comments",
            "description",
            "goaldescription",
            "goalname",
            "name",
            "note",
            "notes",
            "title",
        }

        def remove_goal_text(value):
            if isinstance(value, dict):
                return {
                    key: remove_goal_text(nested)
                    for key, nested in value.items()
                    if re.sub(
                        r"[^a-z0-9]",
                        "",
                        str(key).lower(),
                    ) not in private_goal_text_keys
                }
            if isinstance(value, list):
                return [remove_goal_text(item) for item in value]
            return value

        active_goals = remove_goal_text(active_goals)
    return _strip_empty({
        "personal_records": records,
        "active_goals": active_goals,
    })


def _compact_hydration(data):
    if not isinstance(data, dict):
        return None
    actual_intake = _number(data.get("valueInML"))
    activity_intake = _number(data.get("activityIntakeInML"))
    sweat_loss = _number(data.get("sweatLossInML"))
    has_real_data = any(
        value is not None and value > 0
        for value in (actual_intake, activity_intake, sweat_loss)
    ) or bool(data.get("lastEntryTimestampLocal"))
    if not has_real_data:
        return None
    return _strip_empty({
        "intake_ml": actual_intake,
        "activity_intake_ml": activity_intake,
        "sweat_loss_ml": sweat_loss,
        "last_entry_local": data.get("lastEntryTimestampLocal"),
    })


def _compact_nutrition(data):
    if not isinstance(data, dict):
        return None
    food_log = data.get("food_log", {})
    if not isinstance(food_log, dict):
        return None
    foods = food_log.get("loggedFoodsWithServingSizes")
    if not isinstance(foods, list) or not foods:
        return None
    return _strip_empty({
        "daily_totals": _sanitize_compact(
            {
                key: value
                for key, value in food_log.items()
                if key not in ("loggedFoodsWithServingSizes", "dailyNutritionGoals")
            }
        ),
        "logged_foods": _sanitize_compact(foods),
    })


def _compact_gear_section(
    data,
    reference_secret=None,
    include_free_text=False,
):
    if not isinstance(data, dict):
        return []
    stats_by_id = {}
    for detail in _as_list(data.get("gear_details")):
        if not isinstance(detail, dict):
            continue
        gear_id = _pick(detail, "_uuid", "uuid", "gearUUID")
        if gear_id is not None:
            stats_by_id[str(gear_id)] = detail.get("stats", {})
    return _compact_gear_items(
        data.get("gear_list"),
        stats_by_id,
        reference_secret=reference_secret,
        include_free_text=include_free_text,
    )


def _average_present(items, path):
    values = []
    for item in items:
        value = item
        for key in path:
            if not isinstance(value, dict):
                value = None
                break
            value = value.get(key)
        numeric = _number(value)
        if numeric is not None:
            values.append(numeric)
    if not values:
        return None
    return round(sum(values) / len(values), 1)


def _sport_family(sport):
    value = str(sport or "").lower()
    if any(part in value for part in ("run", "jog")):
        return "running"
    if any(part in value for part in ("cycl", "bike", "biking")):
        return "cycling"
    if any(part in value for part in ("strength", "weight_training")):
        return "strength"
    return "other"


def _weekly_summary(activities, daily_records):
    activities = activities or []
    daily_records = daily_records or []
    summary = {
        "running_sessions": 0,
        "running_distance_m": 0,
        "running_duration_s": 0,
        "cycling_sessions": 0,
        "cycling_distance_m": 0,
        "cycling_duration_s": 0,
        "strength_sessions": 0,
        "strength_duration_s": 0,
        "other_sessions": 0,
        "total_training_duration_s": 0,
        "longest_run_distance_m": 0,
        "total_elevation_gain_m": 0,
        "weekly_training_load": 0,
        "self_evaluated_activities": 0,
    }
    zone_seconds = {}
    valid_heart_rate_seconds = 0.0
    missing_heart_rate_seconds = 0.0
    valid_unclassified_heart_rate_seconds = 0.0
    for activity in activities:
        family = _sport_family(activity.get("sport"))
        duration = _number(activity.get("duration_s")) or 0
        distance = _number(activity.get("distance_m")) or 0
        elevation = _number(activity.get("elevation_gain_m")) or 0
        load = _number(activity.get("training_load")) or 0
        summary["total_training_duration_s"] += duration
        summary["total_elevation_gain_m"] += elevation
        summary["weekly_training_load"] += load
        if activity.get("self_evaluation"):
            summary["self_evaluated_activities"] += 1
        if family == "running":
            summary["running_sessions"] += 1
            summary["running_distance_m"] += distance
            summary["running_duration_s"] += duration
            summary["longest_run_distance_m"] = max(
                summary["longest_run_distance_m"],
                distance,
            )
        elif family == "cycling":
            summary["cycling_sessions"] += 1
            summary["cycling_distance_m"] += distance
            summary["cycling_duration_s"] += duration
        elif family == "strength":
            summary["strength_sessions"] += 1
            summary["strength_duration_s"] += duration
        else:
            summary["other_sessions"] += 1

        for zone in activity.get("hr_zones", []):
            number = zone.get("zone")
            seconds = _number(zone.get("duration_s"))
            if number is not None and seconds is not None:
                zone_seconds[number] = zone_seconds.get(number, 0) + seconds
        quality = activity.get("heart_rate_distribution_quality", {})
        valid_heart_rate_seconds += (
            _number(quality.get("valid_heart_rate_duration_s")) or 0
        )
        missing_heart_rate_seconds += (
            _number(quality.get("missing_heart_rate_duration_s")) or 0
        )
        valid_unclassified_heart_rate_seconds += (
            _number(quality.get("valid_hr_unclassified_duration_s")) or 0
        )

    summary.update({
        "average_resting_heart_rate_bpm": _average_present(
            daily_records,
            ["resting_heart_rate_bpm"],
        ),
        "average_sleep_s": _average_present(
            daily_records,
            ["sleep", "total_sleep_s"],
        ),
        "average_hrv_ms": _average_present(
            daily_records,
            ["hrv", "overnight_average_ms"],
        ),
        "average_stress": _average_present(
            daily_records,
            ["average_stress"],
        ),
    })
    total_zone_seconds = sum(zone_seconds.values())
    distribution = []
    for number in sorted(zone_seconds):
        seconds = zone_seconds[number]
        distribution.append({
            "zone": number,
            "duration_s": round(seconds, 1),
            "percentage": (
                round(seconds * 100.0 / total_zone_seconds, 1)
                if total_zone_seconds
                else None
            ),
        })
    return _strip_empty({
        "weekly_summary": {
            key: round(value, 1) if isinstance(value, float) else value
            for key, value in summary.items()
        },
        "heart_rate_distribution": distribution,
        "heart_rate_distribution_quality": {
            "valid_heart_rate_duration_s": valid_heart_rate_seconds,
            "missing_heart_rate_duration_s": missing_heart_rate_seconds,
            "valid_hr_unclassified_duration_s": (
                valid_unclassified_heart_rate_seconds
            ),
            "heart_rate_zone_coverage_pct": (
                total_zone_seconds * 100.0 / valid_heart_rate_seconds
                if valid_heart_rate_seconds
                else None
            ),
            "distribution_total_pct": (
                sum(item.get("percentage", 0) for item in distribution)
                if distribution
                else None
            ),
        },
    })


def _round_activity_series(series):
    if not isinstance(series, dict):
        return series
    descriptors = series.get("metric_descriptors")
    samples = series.get("samples")
    if not isinstance(descriptors, list) or not isinstance(samples, list):
        return series
    fields = [
        descriptor.get("field") if isinstance(descriptor, dict) else None
        for descriptor in descriptors
    ]
    rounded_samples = []
    for sample in samples:
        if not isinstance(sample, list):
            rounded_samples.append(sample)
            continue
        row = []
        for index, value in enumerate(sample):
            field = fields[index] if index < len(fields) else None
            numeric = _number(value)
            if numeric is None:
                row.append(value)
            elif field == "timestamp":
                row.append(round(numeric))
            elif field and any(part in field for part in ("heart_rate", "power")):
                row.append(round(numeric))
            elif field and any(part in field for part in (
                "duration", "distance", "elevation", "temperature",
                "cadence", "ground_contact", "vertical_oscillation",
                "stride_length",
            )):
                row.append(round(numeric, 1))
            elif field and "speed" in field:
                row.append(round(numeric, 4))
            else:
                row.append(value)
        rounded_samples.append(row)
    return {
        "metric_descriptors": _round_compact_output(descriptors),
        "samples": rounded_samples,
    }


def _round_compact_output(data, key=None, parent=None):
    """Redondea solo la representación compacta, nunca la caché ni cálculos."""
    if isinstance(data, dict):
        rounded = {}
        for child_key, value in data.items():
            if child_key == "activity_series":
                rounded[child_key] = _round_activity_series(value)
            else:
                rounded[child_key] = _round_compact_output(
                    value,
                    child_key,
                    data,
                )
        return rounded
    if isinstance(data, list):
        return [_round_compact_output(item, key, parent) for item in data]
    numeric = _number(data)
    if numeric is None:
        return data

    name = str(key or "").lower()
    if parent and "speed_source_unit" in parent and name == "speed_m_s":
        return round(numeric, 4)
    if any(part in name for part in (
        "distance", "duration", "pace", "temperature", "percentage",
        "_pct", "training_effect", "training_load", "elevation",
    )):
        return round(numeric, 1)
    if any(part in name for part in (
        "heart_rate", "_bpm", "power_w", "maximum_power",
        "average_power", "normalized_power",
    )):
        return round(numeric)
    if "speed_m_s" in name:
        return round(numeric, 3)
    return data


def _json(data):
    """Serializa JSON y elimina vacíos en modo compacto.

    En modo dividido, cada entrada superior ocupa una línea para facilitar la
    indexación. Los objetos interiores permanecen compactos.
    """
    if _compact_mode:
        data = _strip_empty(data)
        data = _round_compact_output(data)
    if _split_mode:
        # Una línea por clave o elemento superior para facilitar el análisis.
        if isinstance(data, dict) and data:
            lines = []
            for k, v in data.items():
                lines.append(f"  {json.dumps(k)}: {json.dumps(v, default=str, ensure_ascii=False)}")
            return "{\n" + ",\n".join(lines) + "\n}"
        elif isinstance(data, list) and data:
            lines = [f"  {json.dumps(item, default=str, ensure_ascii=False)}" for item in data]
            return "[\n" + ",\n".join(lines) + "\n]"
    indent = None if _compact_mode else 2
    return json.dumps(data, indent=indent, default=str, ensure_ascii=False)


def _normalise_output_filename(filename: str) -> str:
    """Valida el nombre del archivo de texto solicitado.

    Solo se admite un nombre, nunca una ruta, para mantener la salida dentro de
    la carpeta elegida con --output.
    """
    name = filename.strip()
    if not name:
        raise ValueError("el nombre del archivo no puede estar vacío")
    if name in {".", ".."} or "/" in name or "\\" in name:
        raise ValueError("utiliza solo un nombre, sin carpetas")
    if any(char in name for char in '<>:"|?*'):
        raise ValueError("el nombre contiene caracteres no permitidos en Windows")

    if not name.lower().endswith(".txt"):
        name += ".txt"

    reserved = {
        "CON", "PRN", "AUX", "NUL",
        *(f"COM{i}" for i in range(1, 10)),
        *(f"LPT{i}" for i in range(1, 10)),
    }
    if Path(name).stem.upper() in reserved:
        raise ValueError("ese nombre está reservado por Windows")
    return name


def _normalise_output_stem(filename: str) -> str:
    """Valida un nombre base compartido por TXT y XLSX."""
    name = filename.strip()
    if not name:
        raise ValueError("el nombre del archivo no puede estar vacío")
    if name in {".", ".."} or "/" in name or "\\" in name:
        raise ValueError("utiliza solo un nombre, sin carpetas")
    if any(char in name for char in '<>:"|?*'):
        raise ValueError("el nombre contiene caracteres no permitidos en Windows")
    suffix = Path(name).suffix.casefold()
    if suffix in {".txt", ".xlsx"}:
        name = name[: -len(suffix)]
    if not name or name.endswith((" ", ".")):
        raise ValueError("el nombre no puede terminar en un espacio o un punto")
    reserved = {
        "CON", "PRN", "AUX", "NUL",
        *(f"COM{i}" for i in range(1, 10)),
        *(f"LPT{i}" for i in range(1, 10)),
    }
    if Path(name).name.upper() in reserved:
        raise ValueError("ese nombre está reservado por Windows")
    return name


def _inclusive_review_start(end_date: date, review_weeks: int) -> date:
    """Calcula un intervalo inclusivo de exactamente N semanas."""
    return end_date - timedelta(days=review_weeks * 7 - 1)


def _relative_manifest_paths(paths, output_directory):
    """Devuelve rutas portables sin publicar la carpeta personal del equipo."""
    output_root = Path(output_directory).resolve()
    relative_paths = []
    for path in paths:
        resolved = Path(path).resolve()
        try:
            relative = resolved.relative_to(output_root)
        except ValueError:
            raise RuntimeError(
                "El manifiesto solo puede incluir archivos de la carpeta de salida."
            ) from None
        relative_paths.append(relative.as_posix())
    return relative_paths


_WINDOWS_TIMEZONE_MAP = {
    "Romance Standard Time": "Europe/Madrid",
}


def _windows_timezone_key():
    if os.name != "nt":
        return None
    try:
        import winreg
        path = r"SYSTEM\CurrentControlSet\Control\TimeZoneInformation"
        with winreg.OpenKey(winreg.HKEY_LOCAL_MACHINE, path) as key:
            return winreg.QueryValueEx(key, "TimeZoneKeyName")[0]
    except (ImportError, OSError):
        return None


def _resolve_timezone(configured=None):
    """Resuelve una zona IANA configurable y estable para fechas históricas."""
    candidate = configured or os.getenv("GARMIN_EXPORT_TIMEZONE")
    if candidate:
        try:
            ZoneInfo(candidate)
        except ZoneInfoNotFoundError as exc:
            raise ValueError(
                f"zona horaria IANA desconocida: {candidate}"
            ) from exc
        return candidate

    windows_key = _windows_timezone_key()
    if windows_key in _WINDOWS_TIMEZONE_MAP:
        return _WINDOWS_TIMEZONE_MAP[windows_key]

    system_zone = getattr(datetime.now().astimezone().tzinfo, "key", None)
    if system_zone:
        try:
            ZoneInfo(system_zone)
            return system_zone
        except ZoneInfoNotFoundError:
            pass
    return "UTC"


def _timezone_metadata(timezone_name, period_end, current=None):
    zone = ZoneInfo(timezone_name)
    exported_at = (
        current.astimezone(zone)
        if isinstance(current, datetime)
        else datetime.now(zone)
    )
    historical = datetime(
        period_end.year,
        period_end.month,
        period_end.day,
        12,
        tzinfo=zone,
    )
    offset = historical.strftime("%z")
    if len(offset) == 5:
        offset = f"{offset[:3]}:{offset[3:]}"
    return exported_at, offset


def _section(md: list, title: str, data, level: int = 3):
    """Añade un bloque JSON con título; omite valores nulos."""
    if data is None:
        return
    md.append(f"{title}\n")
    md.append(f"{_json(data)}\n")


def _section_nodata(md: list, title: str):
    """Indica que una categoría completa no tiene datos."""
    md.append("No hay datos disponibles.\n")


def _chunked_date_call(fn, start: date, end: date, label: str, chunk_days: int = 365):
    """Consulta un intervalo por bloques anuales y combina los resultados.

    Algunos endpoints rechazan periodos superiores a un año con un error 400.
    Esta función los divide y reúne las respuestas en una sola lista.
    """
    all_results = []
    chunk_start = start
    while chunk_start <= end:
        # Restar uno porque ambos límites de la API son inclusivos.
        chunk_end = min(
            chunk_start + timedelta(days=chunk_days - 1),
            end,
        )
        result = safe_call(fn, chunk_start.isoformat(), chunk_end.isoformat(),
                           label=f"{label}_{chunk_start}")
        if result is not None:
            if isinstance(result, list):
                all_results.extend(result)
            else:
                all_results.append(result)
        chunk_start = chunk_end + timedelta(days=1)
    return all_results if all_results else None


# ---------------------------------------------------------------------------
# Caché para reanudar exportaciones interrumpidas sin repetir el trabajo.
# ---------------------------------------------------------------------------
_CACHE_ENVELOPE_KEY = "__garmin_export_cache__"
_CACHE_FORMAT_VERSION = 2


class ExportCache:
    """Caché JSON para respuestas diarias, actividades y secciones.

    Se guarda en {output_dir}/.cache/ y utiliza fechas o identificadores como
    claves. Los datos históricos permanecen entre ejecuciones.
    """

    def __init__(
        self,
        out_dir: Path,
        enabled: bool = True,
        cache_dir: Optional[Path] = None,
    ):
        self.enabled = enabled
        self.cache_dir = (
            Path(cache_dir)
            if cache_dir is not None
            else out_dir / ".cache"
        )
        self.daily_dir = self.cache_dir / "daily"
        self.activity_dir = self.cache_dir / "activities"
        self.section_dir = self.cache_dir / "sections"
        self.hits = 0
        self.misses = 0

        if not enabled:
            return

        self.daily_dir.mkdir(parents=True, exist_ok=True)
        self.activity_dir.mkdir(parents=True, exist_ok=True)
        self.section_dir.mkdir(parents=True, exist_ok=True)

        existing_files = list(self.daily_dir.glob("*.json"))
        daily_health = sum(1 for f in existing_files if f.name[0].isdigit())
        daily_hydration = sum(1 for f in existing_files if f.name.startswith("hydration_"))
        daily_nutrition = sum(1 for f in existing_files if f.name.startswith("nutrition_"))
        existing_acts = len(list(self.activity_dir.glob("*.json")))
        existing_sects = len(list(self.section_dir.glob("*.json")))
        total = len(existing_files) + existing_acts + existing_sects
        if total:
            parts = []
            if daily_health:
                parts.append(f"{daily_health} días de salud")
            if daily_hydration:
                parts.append(f"{daily_hydration} días de hidratación")
            if daily_nutrition:
                parts.append(f"{daily_nutrition} días de nutrición")
            if existing_acts:
                parts.append(f"{existing_acts} actividades")
            if existing_sects:
                parts.append(f"{existing_sects} secciones")
            log.info(f"Caché: {', '.join(parts)}")

    def _wipe(self):
        """Elimina la caché antigua."""
        import shutil
        if self.cache_dir.exists():
            shutil.rmtree(self.cache_dir, ignore_errors=True)

    def _read_entry(
        self,
        path: Path,
        *,
        accept_legacy: bool,
    ) -> tuple[Optional[dict], set[str], bool]:
        """Lee una entrada y separa sus datos de los metadatos de integridad."""
        if not self.enabled:
            return None, set(), False
        if path.exists():
            try:
                raw = json.loads(path.read_text(encoding="utf-8"))
                if not isinstance(raw, dict):
                    raise ValueError("La entrada de caché no es un objeto.")
                envelope = raw.get(_CACHE_ENVELOPE_KEY)
                if isinstance(envelope, dict):
                    payload = raw.get("data")
                    if (
                        envelope.get("version") != _CACHE_FORMAT_VERSION
                        or not isinstance(payload, dict)
                    ):
                        self.misses += 1
                        return None, set(), False
                    complete_keys = {
                        str(key)
                        for key in envelope.get("complete_keys", [])
                        if isinstance(key, str)
                    }
                    if envelope.get("complete") is False:
                        self.misses += 1
                        return None, complete_keys, False
                    self.hits += 1
                    return payload, complete_keys, True
                if accept_legacy:
                    # Los datos se conservan, pero ninguna clave se considera
                    # verificada: la primera ejecución con v3 la actualizará.
                    self.hits += 1
                    return raw, set(), False
            except (json.JSONDecodeError, OSError, ValueError):
                pass
        self.misses += 1
        return None, set(), False

    @staticmethod
    def _write_entry(
        path: Path,
        data: dict,
        *,
        complete: bool,
        complete_keys: Optional[set[str]] = None,
    ) -> None:
        """Escribe una entrada con una marca explícita de integridad."""
        keys = complete_keys if complete_keys is not None else set(data)
        envelope = {
            _CACHE_ENVELOPE_KEY: {
                "version": _CACHE_FORMAT_VERSION,
                "complete": bool(complete),
                "complete_keys": sorted(str(key) for key in keys),
            },
            "data": data,
        }
        path.write_text(
            json.dumps(envelope, default=str, ensure_ascii=False),
            encoding="utf-8",
        )

    def get_day_entry(self, ds: str) -> tuple[Optional[dict], set[str]]:
        """Devuelve los datos diarios y las claves confirmadas por Garmin."""
        if not self.enabled:
            return None, set()
        path = self.daily_dir / f"{ds}.json"
        data, complete_keys, _ = self._read_entry(
            path,
            accept_legacy=True,
        )
        return data, complete_keys

    def get_day(self, ds: str) -> Optional[dict]:
        """Compatibilidad: devuelve únicamente los datos de una entrada diaria."""
        data, _ = self.get_day_entry(ds)
        return data

    def put_day(
        self,
        ds: str,
        data: dict,
        *,
        complete_keys: Optional[set[str]] = None,
    ):
        if not self.enabled:
            return
        path = self.daily_dir / f"{ds}.json"
        keys = complete_keys if complete_keys is not None else set(data)
        self._write_entry(
            path,
            data,
            complete=True,
            complete_keys=keys,
        )

    def get_activity(self, activity_id) -> Optional[dict]:
        if not self.enabled:
            return None
        path = self.activity_dir / f"{activity_id}.json"
        data, _, _ = self._read_entry(path, accept_legacy=False)
        return data

    def put_activity(self, activity_id, data: dict, *, complete: bool = True):
        if not self.enabled:
            return
        path = self.activity_dir / f"{activity_id}.json"
        self._write_entry(path, data, complete=complete)

    def get_section(self, name: str) -> Optional[dict]:
        """Obtiene de caché una sección completa."""
        if not self.enabled:
            return None
        path = self.section_dir / f"{name}.json"
        data, _, _ = self._read_entry(path, accept_legacy=False)
        return data

    def section_needs_refresh(self, name: str, max_age_days: int) -> bool:
        """Indica si una sección no existe o ha superado su vigencia."""
        if not self.enabled:
            return True
        path = self.section_dir / f"{name}.json"
        try:
            age_seconds = max(0.0, time.time() - path.stat().st_mtime)
        except OSError:
            return True
        return age_seconds >= timedelta(days=max_age_days).total_seconds()

    def put_section(self, name: str, data: dict, *, complete: bool = True):
        if not self.enabled:
            return
        path = self.section_dir / f"{name}.json"
        self._write_entry(path, data, complete=complete)

    def summary(self) -> str:
        total = self.hits + self.misses
        if total == 0:
            return "Caché: sin consultas"
        pct = (self.hits / total) * 100
        return f"Caché: {self.hits} reutilizados, {self.misses} nuevos ({pct:.0f}% reutilizado)"


# ---------------------------------------------------------------------------
# Exportador
# ---------------------------------------------------------------------------
class GarminExporter:
    def __init__(self, api: Garmin, out_dir: Path, days: int, max_activities: int,
                 fetch_all: bool = False, cache: Optional[ExportCache] = None,
                 update_mode: bool = False,
                 explicit_start_date: Optional[date] = None,
                 explicit_end_date: Optional[date] = None,
                 output_filename: Optional[str] = None,
                 include_activity_details: bool = False,
                 timezone_name: Optional[str] = None,
                 output_format: str = "txt",
                 report_type: str = "history",
                 race_context: Optional[dict] = None,
                 journal: Optional[list] = None,
                  review_weeks: Optional[int] = None,
                  selected_activity_ref: Optional[str] = None,
                  include_free_text: bool = False,
                  manifest_path: Optional[Path] = None,
                  run_id: Optional[str] = None):
        self.api = api
        self.out = out_dir
        self.max_activities = max_activities
        self.fetch_all = fetch_all
        self.update_mode = update_mode
        self.explicit_start_date = explicit_start_date
        self.explicit_end_date = explicit_end_date
        self.output_filename = output_filename
        self.output_format = output_format
        self.report_type = report_type
        self.race_context = race_context
        self.journal = journal or []
        self.review_weeks = review_weeks
        self.selected_activity_ref = selected_activity_ref
        self.include_free_text = include_free_text
        self.manifest_path = Path(manifest_path) if manifest_path else None
        self.run_id = run_id
        self.include_activity_details = (
            include_activity_details or report_type == "activity"
        )
        self.timezone_name = _resolve_timezone(timezone_name)
        self.cache = cache or ExportCache(out_dir, enabled=False)
        self.reference_secret = load_or_create_reference_secret(
            self.cache.cache_dir
        )
        self.today = explicit_end_date or date.today()
        self.errors: list[str] = []
        self.endpoint_failures: set[tuple[str, str, str]] = set()
        self._endpoint_failure_events = 0
        self._endpoint_failures_lock = threading.Lock()
        self.md: list[str] = []
        self.semantic_model: dict[str, Any] = {}
        self.written_files: list[Path] = []
        self.sensitive_identifiers: set[str] = set()
        self.sensitive_private_values: set[str] = set()
        self.sensitive_activity_names: set[str] = set()
        self.update_base_date: Optional[str] = None  # end date of base export
        self.compact_daily_records = []
        self.compact_activities = []
        self.compact_training = {}
        self.data_quality = {
            "missing_critical_data": [],
            "warnings": [],
            "endpoint_errors": [],
            "series_validation_errors": [],
            "temporal_warnings": [],
            "current_snapshots_detected": [],
            "unit_conversions": [],
            "privacy_filters_applied": self._privacy_filter_descriptions(),
            "duplicate_sources_removed": [],
        }

        if update_mode:
            base_end = self._find_latest_export_end_date()
            if base_end:
                self.update_base_date = base_end.isoformat()
                # Solapar un día para recoger datos que hayan llegado tarde.
                self.start_date = base_end - timedelta(days=1)
                log.info(f"Modo de actualización: la última exportación termina el {base_end}; "
                         f"descargando desde {self.start_date}")
            else:
                log.warning("No se encontró una exportación anterior; se utilizará --days")
                self.start_date = self.today - timedelta(days=max(days - 1, 0))
        elif explicit_start_date is not None:
            self.start_date = explicit_start_date
        elif fetch_all:
            self.start_date = self._detect_start_date()
        else:
            self.start_date = self.today - timedelta(days=max(days - 1, 0))
        if self.start_date > self.today:
            raise ValueError("la fecha inicial no puede ser posterior a la fecha final")
        self.days = (self.today - self.start_date).days + 1

    def _range_cache_key(self, section_name: str) -> str:
        """Separa las cachés dependientes de fechas por intervalo inclusivo."""
        return f"{section_name}_{self.start_date}_{self.today}"

    def _privacy_filter_descriptions(self):
        return [
            "Las credenciales, tokens y cookies nunca se incorporan desde el almacén de sesión.",
            "Los títulos de actividades y todas las métricas deportivas derivadas se conservan.",
            "Se eliminan identidad, contacto, direcciones e identificadores personales de la fuente.",
            "Se conservan coordenadas, ubicaciones deportivas, tracks y polilíneas.",
        ]

    def _quality_add(self, category, message):
        values = self.data_quality.setdefault(category, [])
        if message not in values:
            values.append(message)

    def _record_endpoint_failure(
        self,
        section: str,
        endpoint: str,
        reason: str,
    ) -> None:
        """Conserva solo nombres técnicos controlados, nunca mensajes crudos."""
        safe_section = re.sub(
            r"[^A-Za-z0-9 &_-]",
            "",
            str(section),
        )[:80] or "Section"
        safe_endpoint = re.sub(
            r"[^A-Za-z0-9_]",
            "",
            str(endpoint),
        )[:80] or "endpoint"
        safe_reason = re.sub(
            r"[^A-Za-z0-9_. ()-]",
            "",
            str(reason),
        )[:100] or "Error"
        with self._endpoint_failures_lock:
            self._endpoint_failure_events += 1
            self.endpoint_failures.add(
                (safe_section, safe_endpoint, safe_reason)
            )

    def _partial_sections(self) -> list[str]:
        sections = {
            error.split(":", 1)[0]
            for error in self.errors
        }
        with self._endpoint_failures_lock:
            sections.update(
                section
                for section, _, _ in self.endpoint_failures
            )
        return sorted(sections)

    def _endpoint_failure_count(self) -> int:
        """Cuenta cada llamada fallida para decidir si una caché es reutilizable."""
        with self._endpoint_failures_lock:
            return self._endpoint_failure_events

    def _is_partial(self) -> bool:
        with self._endpoint_failures_lock:
            return bool(self.errors or self.endpoint_failures)

    def _remember_sensitive_payload(self, payload):
        """Recuerda valores privados observados para comprobar la salida final."""
        def visit(value, parents=()):
            if isinstance(value, dict):
                for key, nested in value.items():
                    normal = re.sub(r"[^a-z0-9]", "", str(key).lower())
                    if (
                        is_personal_data_key(key, parents)
                        and not isinstance(nested, (dict, list, bool))
                        and nested is not None
                    ):
                        candidate = str(nested).strip()
                        if len(candidate) >= 4:
                            self.sensitive_private_values.add(candidate)
                    visit(nested, (*parents, normal))
            elif isinstance(value, list):
                for nested in value:
                    visit(nested, parents)

        visit(payload)

    def _remember_compact(self, key: str, value):
        """Guarda el objeto que compartirán los renderizadores TXT y XLSX."""
        if _compact_mode:
            self.semantic_model[key] = value

    def _finalize_semantic_model(self):
        """Añade los cálculos v3 y ejecuta el control final de privacidad."""
        if not _compact_mode:
            return
        extensions = build_report_extensions(
            self.compact_activities,
            self.compact_daily_records,
            self.start_date,
            self.today,
            race_context=self.race_context,
            journal=self.journal,
        )
        self.compact_activities = extensions["activities"]
        self.semantic_model["activities"] = self.compact_activities
        self.semantic_model["period_summary"] = extensions["period_summary"]
        self.semantic_model["weekly_timeline"] = extensions["weekly_timeline"]
        self.semantic_model["race_analysis"] = extensions["race_analysis"]
        self.semantic_model["suggested_prompts"] = extensions["prompts"]
        if self.race_context:
            self.semantic_model["race_context"] = self.race_context
        if self.journal:
            self.semantic_model["journal"] = self.journal

        with self._endpoint_failures_lock:
            endpoint_failures = sorted(self.endpoint_failures)
        for error in self.errors:
            section = error.split(":", 1)[0]
            self._quality_add(
                "endpoint_errors",
                f"La sección {section} no terminó correctamente.",
            )
        for section, endpoint, reason in endpoint_failures:
            self._quality_add(
                "endpoint_errors",
                (
                    f"La consulta {endpoint} de la sección {section} "
                    f"no terminó correctamente ({reason})."
                ),
            )

        quality = build_quality_report(
            self.compact_daily_records,
            self.compact_activities,
            self.start_date,
            self.today,
            legacy_quality=self.data_quality,
        )
        activities_with_unmapped_data = [
            activity
            for activity in self.compact_activities
            if isinstance(activity, dict) and activity.get("unmapped_sport_data")
        ]
        unmapped_temporal_fields = sorted({
            descriptor.get("source_field")
            for activity in activities_with_unmapped_data
            for descriptor in (
                (
                    (
                        activity.get("unmapped_sport_data") or {}
                    ).get("details") or {}
                ).get("unmapped_activity_series") or {}
            ).get("metric_descriptors", [])
            if isinstance(descriptor, dict) and descriptor.get("source_field")
        })
        quality["compact_data_reduction"] = {
            "raw_activity_copy_exported": False,
            "normalised_sports_data_exported_once": True,
            "activities_with_unmapped_sport_data": len(
                activities_with_unmapped_data
            ),
            "unmapped_temporal_source_fields": unmapped_temporal_fields,
            "method": (
                "Los campos normalizados se exportan una sola vez; solo se "
                "conservan aparte los datos deportivos aún no reconocidos."
            ),
        }
        exported_gear = list(_as_list(self.semantic_model.get("gear")))
        for activity in self.compact_activities:
            if isinstance(activity, dict):
                exported_gear.extend(_as_list(activity.get("gear")))
        privacy = quality.setdefault("privacy", {})
        privacy.update({
            "mode": "redact_personal_identifiers",
            "garmin_activity_ids_exported": False,
            "garmin_gear_ids_exported": False,
            "activity_titles_exported_by_default": True,
            "exact_activity_times_exported_by_default": (
                self.include_free_text
            ),
            "coordinates_and_locations_exported": True,
            "titles_preserved": True,
            "coordinates_preserved_when_available": True,
            "altitude_profiles_preserved_when_available": True,
            "complete_laps_and_splits_preserved_when_available": True,
            "fields_preserved": [
                "activity_titles",
                "sports_metrics",
                "altitude_and_elevation_profiles",
                "laps_and_splits",
                "grade_adjusted_pace",
                "coordinates",
                "tracks",
                "sport_locations",
                "polylines",
            ],
            "fields_removed": [
                "credentials",
                "tokens",
                "cookies",
                "personal_identity",
                "personal_contact",
                "personal_addresses",
                "personal_identifiers",
            ] + (
                []
                if self.include_free_text
                else ["activity_descriptions_notes_and_exact_times"]
            ),
            "transformations_applied": ["personal_fields_redacted"],
        })
        exported_series_fields = {
            descriptor.get("field")
            for activity in self.compact_activities
            if isinstance(activity, dict)
            for descriptor in (
                (activity.get("activity_series") or {}).get(
                    "metric_descriptors",
                    [],
                )
            )
            if isinstance(descriptor, dict)
        }
        privacy["coordinates_present_in_export"] = any(
            isinstance(activity, dict) and bool(activity.get("coordinates"))
            for activity in self.compact_activities
        ) or bool({"latitude_deg", "longitude_deg"} & exported_series_fields)
        privacy["titles_present_in_export"] = any(
            isinstance(activity, dict) and bool(activity.get("name"))
            for activity in self.compact_activities
        )
        privacy["altitude_profiles_present_in_export"] = bool(
            {"elevation_raw"} & exported_series_fields
        )
        privacy["laps_or_splits_present_in_export"] = any(
            isinstance(activity, dict) and bool(activity.get("laps"))
            for activity in self.compact_activities
        )
        privacy["gear_names_exported"] = any(
            isinstance(item, dict) and bool(item.get("gear_name"))
            for item in exported_gear
        )
        privacy["gear_user_provided_text_exported"] = any(
            isinstance(item, dict)
            and (
                item.get("gear_name_user_provided") is True
                or item.get("model_user_provided") is True
            )
            for item in exported_gear
        )
        metadata = self.semantic_model.setdefault("export_metadata", {})
        is_partial = self._is_partial()
        metadata["export_status"] = "partial" if is_partial else "completed"
        if is_partial:
            failed_sections = self._partial_sections()
            metadata["failed_sections"] = failed_sections
            quality["export_errors"] = [
                {
                    "section": section,
                    "message": "La sección no terminó correctamente.",
                }
                for section in failed_sections
            ]
        if self.errors:
            fatal_sections = sorted({
                error.split(":", 1)[0]
                for error in self.errors
            })
            quality.setdefault("issues", []).extend(
                {
                    "code": "EXPORT_SECTION_ERROR",
                    "severity": "error",
                    "scope": section,
                    "message": "La sección no terminó correctamente.",
                }
                for section in fatal_sections
            )
        audit_model = {
            key: value
            for key, value in self.semantic_model.items()
            if key != "data_quality"
        }
        if self.errors:
            audit_model["export_errors"] = self.errors
        identifier_values = [
            value
            for value in self.sensitive_identifiers
            if len(value.strip()) >= 6
        ]
        private_values = [
            value
            for value in self.sensitive_private_values
            if len(value.strip()) >= 4
        ]
        audit = privacy_audit(
            audit_model,
            forbidden_values=private_values,
            forbidden_identifiers=identifier_values,
        )
        quality["privacy_audit"] = audit
        if not audit.get("passed"):
            paths = audit.get("forbidden_key_paths", [])
            if paths:
                log.error(
                    "La auditoría encontró campos personales sin filtrar en: "
                    + ", ".join(paths[:8])
                )
            if audit.get("forbidden_values_detected"):
                value_paths = audit.get("forbidden_value_paths", [])
                if value_paths:
                    log.error(
                        "Los valores no permitidos aparecieron en estos campos: "
                        + ", ".join(value_paths[:8])
                    )
                log.error(
                    "La auditoría encontró valores personales o identificadores "
                    "en campos que debían estar filtrados."
                )
            raise RuntimeError(
                "La comprobación de privacidad detectó campos no permitidos; "
                "no se ha escrito ningún archivo."
            )
        self.semantic_model["data_quality"] = quality

    def _render_compact_text(self) -> str:
        """Renderiza el modelo semántico v3 como texto plano indexable."""
        exported_at = (
            (self.semantic_model.get("export_metadata") or {}).get("exported_at")
            or datetime.now().isoformat()
        )
        lines = [
            "Exportación privada de entrenamiento de Garmin Connect",
            f"Exportado: {exported_at}",
            (
                f"Intervalo de fechas: {self.start_date} a {self.today} "
                f"({self.days} días)"
            ),
            (
                f"Formato: compacto semántico v{_COMPACT_SCHEMA_VERSION}; "
                "texto plano con objetos JSON."
            ),
            (
                "Finalidad: revisar una preparación deportiva con una IA. "
                "Nada se sube automáticamente."
            ),
            "",
            "Índice de contenidos",
        ]
        section_info = [
            ("Export Metadata", "export_metadata", "Versión, intervalo, zona horaria y privacidad."),
            ("Race Context", "race_context", "Carrera, objetivo y disponibilidad aportados por la persona."),
            ("Profile", "profile", "Contexto deportivo reducido sin identidad."),
            ("Period Summary", "period_summary", "Totales de todo el intervalo solicitado."),
            ("Weekly Timeline", "weekly_timeline", "Una fila por semana ISO, incluidas semanas vacías o parciales."),
            ("Daily Health", "daily_health", "Salud y recuperación por día."),
            ("Activities", "activities", "Actividades privadas, clasificación, vueltas, zonas y equipamiento."),
            ("Blood Pressure", "blood_pressure", "Mediciones reales de presión arterial."),
            ("Body Composition", "body_composition", "Peso y composición corporal con unidades."),
            ("Training Metrics", "training_metrics", "Métricas de Garmin separadas por relación temporal."),
            ("Goals and Records", "goals_and_records", "Récords y objetivos activos sin títulos privados."),
            ("Gear", "gear", "Equipamiento con referencias locales."),
            ("Hydration", "hydration", "Días con hidratación real registrada."),
            ("Nutrition", "nutrition", "Días con nutrición real registrada."),
            ("Journal", "journal", "Anotaciones opcionales aportadas por la persona."),
            ("Race Analysis", "race_analysis", "Comparaciones y métricas derivadas auditables."),
            ("Suggested Prompts", "suggested_prompts", "Instrucciones listas para copiar en una IA."),
            ("Data Quality", "data_quality", "Cobertura, límites, transformaciones y privacidad."),
        ]
        included = [
            item
            for item in section_info
            if item[1] in self.semantic_model
            and self.semantic_model[item[1]] not in (None, {}, [])
        ]
        for index, (title, _, description) in enumerate(included, 1):
            lines.append(f"  {index}. {title} -- {description}")
        lines.append("")
        for title, key, description in included:
            lines.extend([
                title,
                f'Schema: "{description}"',
                _json(self.semantic_model[key]),
                "",
            ])
        if self.errors:
            lines.append("Errores durante la exportación")
            lines.extend(f"- {error}" for error in self.errors)
            lines.append("")
        return "\n".join(lines).rstrip() + "\n"

    def _find_latest_export_end_date(self) -> Optional[date]:
        """Obtiene el final más reciente sin depender del nombre del archivo."""
        found: list[tuple[date, str]] = []
        filename_fallbacks: list[tuple[date, str]] = []

        def remember(raw_date, source: str) -> None:
            try:
                parsed = date.fromisoformat(str(raw_date)[:10])
            except (TypeError, ValueError):
                return
            if parsed <= self.today:
                found.append((parsed, source))

        # El lanzador reutiliza un manifiesto estable por perfil. Es la fuente
        # preferida para nombres personalizados y exportaciones solo XLSX.
        if self.manifest_path and self.manifest_path.exists():
            try:
                manifest = json.loads(
                    self.manifest_path.read_text(encoding="utf-8")
                )
                if (
                    isinstance(manifest, dict)
                    and manifest.get("report_type") != "activity"
                    and manifest.get("status") != "partial"
                    and manifest.get("export_status") != "partial"
                ):
                    remember(manifest.get("end_date"), self.manifest_path.name)
            except (OSError, json.JSONDecodeError):
                pass

        timestamp_pattern = re.compile(
            r"garmin_export_(\d{4}-\d{2}-\d{2}_\d{6})"
        )
        range_pattern = re.compile(
            r"(?:Date range|Intervalo de fechas):\s*(\d{4}-\d{2}-\d{2})"
            r"\s+(?:to|a)\s+(\d{4}-\d{2}-\d{2})"
        )
        metadata_end_pattern = re.compile(
            r'"(?:requested_end_date|end_date)"\s*:\s*"'
            r"(\d{4}-\d{2}-\d{2})"
        )
        report_type_pattern = re.compile(
            r'"report_type"\s*:\s*"([^"]+)"'
        )
        export_status_pattern = re.compile(
            r'"export_status"\s*:\s*"([^"]+)"'
        )
        for text_path in self.out.glob("*.txt"):
            if text_path.name.casefold().startswith("analisis_actividad_"):
                continue
            try:
                header = text_path.read_text(encoding="utf-8")[:20_000]
            except OSError:
                continue
            report_match = report_type_pattern.search(header)
            if report_match and report_match.group(1) == "activity":
                continue
            status_match = export_status_pattern.search(header)
            if status_match and status_match.group(1) == "partial":
                continue
            match = range_pattern.search(header)
            if match:
                remember(match.group(2), text_path.name)
                continue
            match = metadata_end_pattern.search(header)
            if match:
                remember(match.group(1), text_path.name)
                continue
            timestamp_match = timestamp_pattern.search(text_path.name)
            if timestamp_match:
                try:
                    fallback = datetime.strptime(
                        timestamp_match.group(1),
                        "%Y-%m-%d_%H%M%S",
                    ).date()
                    filename_fallbacks.append((fallback, text_path.name))
                except ValueError:
                    pass

        # Los libros v3 guardan el final solicitado en RESUMEN. La lectura es
        # opcional y en modo de solo lectura para no cargar sus series.
        xlsx_paths = list(self.out.glob("*.xlsx"))
        if xlsx_paths:
            try:
                from openpyxl import load_workbook
            except ImportError:
                load_workbook = None
            if load_workbook is not None:
                for xlsx_path in xlsx_paths:
                    if xlsx_path.name.casefold().startswith(
                        "analisis_actividad_"
                    ):
                        continue
                    workbook = None
                    try:
                        workbook = load_workbook(
                            xlsx_path,
                            read_only=True,
                            data_only=True,
                        )
                        sheet = workbook["RESUMEN"]
                        summary_values = {}
                        for row in sheet.iter_rows(
                            min_row=2,
                            max_col=2,
                            values_only=True,
                        ):
                            if row and isinstance(row[0], str):
                                summary_values[row[0]] = row[1]
                        if (
                            str(summary_values.get("export.report_type", ""))
                            .strip()
                            .casefold()
                            == "activity"
                        ):
                            continue
                        if (
                            str(summary_values.get("export.status", ""))
                            .strip()
                            .casefold()
                            == "partial"
                        ):
                            continue
                        remember(
                            summary_values.get("period.end_date"),
                            xlsx_path.name,
                        )
                    except Exception:
                        # Un XLSX ajeno o incompleto no debe impedir actualizar.
                        pass
                    finally:
                        if workbook is not None:
                            workbook.close()

        candidates = found or [
            item for item in filename_fallbacks if item[0] <= self.today
        ]
        if not candidates:
            return None
        latest_end, source = max(candidates, key=lambda item: item[0])
        log.info(
            f"Última exportación encontrada: {source} "
            f"(termina el {latest_end})"
        )
        return latest_end

    def _detect_start_date(self) -> date:
        """Calcula hasta dónde se remonta el historial de la cuenta.

        Busca la actividad más antigua y añade una semana previa para recoger
        datos de salud anteriores. Si no puede determinarlo, usa cinco años.
        """
        log.info("Detectando el periodo disponible en la cuenta...")

        # Solicitar la actividad más antigua con orden ascendente.
        oldest = safe_call(
            self.api.get_activities_by_date,
            "2000-01-01", self.today.isoformat(), None, "asc",
            label="oldest_activity",
        )
        if oldest and isinstance(oldest, list) and len(oldest) > 0:
            first_act = oldest[0]
            start_str = first_act.get("startTimeLocal", "")[:10]
            if start_str:
                try:
                    d = date.fromisoformat(start_str)
                    # Añadir una semana para recoger salud previa a la actividad.
                    d = d - timedelta(days=7)
                    log.info(f"Actividad más antigua encontrada: {start_str}")
                    log.info(f"Se exportará desde: {d}")
                    return d
                except ValueError:
                    pass

        # Alternativa: cinco años.
        fallback = self.today - timedelta(days=365 * 5)
        log.info(f"No se pudo detectar el dato más antiguo; se utilizará {fallback}")
        return fallback

    def run(self):
        now = datetime.now()
        suffix = ""
        if self.update_mode:
            suffix = "_update"
        if self.output_filename:
            output_stem = _normalise_output_stem(self.output_filename)
        elif self.report_type == "preparation":
            output_stem = "revision_preparacion_actual"
        elif self.report_type == "activity":
            output_stem = f"analisis_actividad_{self.today}"
        elif self.explicit_start_date is not None and self.explicit_end_date is not None:
            output_stem = f"garmin_datos_{self.start_date}_a_{self.today}"
        else:
            output_stem = (
                f"garmin_export_{now.strftime('%Y-%m-%d_%H%M%S')}{suffix}"
            )
        txt_filename = f"{output_stem}.txt"
        xlsx_filename = f"{output_stem}.xlsx"

        log.info(f"Periodo: {self.start_date} a {self.today} ({self.days} días)")
        if self.update_mode:
            log.info(f"Modo: actualización (datos nuevos desde {self.update_base_date})")
        if _compact_mode:
            log.info("Modo: compacto (archivo más pequeño para herramientas de IA)")
            log.info(
                "Privacidad: identidad e identificadores personales ocultos "
                "automáticamente"
            )
            if self.include_activity_details:
                log.info("Actividades: se incluirá el máximo detalle temporal registrado")
        if _split_mode:
            log.info("Modo: dividido (varios archivos de menos de 500.000 palabras)")
        if self.fetch_all and not self.update_mode:
            log.info("Modo: --all (historial completo)")
            log.info("Máximo de actividades: sin límite")
        elif self.explicit_start_date is not None:
            log.info(f"Modo: fecha inicial concreta ({self.explicit_start_date})")
            log.info("Actividades: todas las incluidas en el periodo elegido")
        else:
            log.info(f"Máximo de actividades: {self.max_activities}")
        print()

        if self.update_mode:
            self.md.append("Exportación de datos de Garmin Connect — Actualización\n")
            self.md.append(f"Exportado: {now.isoformat()}")
            self.md.append(f"Actualización de datos desde: {self.update_base_date}")
            self.md.append(
                f"Intervalo de fechas: {self.start_date} a {self.today} "
                f"({self.days} días)"
            )
            self.md.append(
                "Este archivo solo contiene datos nuevos. Cárgalo junto a la "
                "exportación base.\n"
            )
        else:
            self.md.append("Exportación de datos de Garmin Connect\n")
            self.md.append(f"Exportado: {now.isoformat()}")
            self.md.append(
                f"Intervalo de fechas: {self.start_date} a {self.today} "
                f"({self.days} días)"
            )
            if self.explicit_start_date is not None:
                self.md.append("Actividades: todas las del intervalo seleccionado")
            else:
                self.md.append(f"Máximo de actividades: {self.max_activities}")
        if _compact_mode:
            activity_detail_note = (
                "series de actividad incluidas a la máxima resolución registrada"
                if self.include_activity_details
                else "series temporales de actividad omitidas"
            )
            self.md.append(
                "Formato: compacto semántico (identidad oculta automáticamente, "
                f"JSON en una línea, {activity_detail_note})"
            )
            self.md.append(
                "Finalidad: análisis de entrenamiento de resistencia asistido por IA; "
                "las unidades de origen se conservan o se nombran al convertirlas.\n"
            )
        else:
            self.md.append(
                "Formato: completo (JSON original con todos los campos)\n"
            )

        if _compact_mode:
            sections = [
                ("Export Metadata", self.export_metadata),
                ("Profile", self.export_profile),
                ("Daily Health", self.export_daily_health),
                ("Blood Pressure", self.export_blood_pressure),
                ("Activities", self.export_activities),
                ("Body Composition", self.export_body_composition),
                ("Training Metrics", self.export_training),
                ("Goals and Records", self.export_goals),
                ("Gear", self.export_gear),
                ("Hydration", self.export_hydration),
                ("Nutrition", self.export_nutrition),
            ]
        else:
            sections = [
                ("Profile", self.export_profile),
                ("Daily Health", self.export_daily_health),
                ("Blood Pressure", self.export_blood_pressure),
                ("Activities", self.export_activities),
                ("Body Composition", self.export_body_composition),
                ("Training Metrics", self.export_training),
                ("Goals and Records", self.export_goals),
                ("Trends", self.export_trends),
                ("Golf", self.export_golf),
                ("Gear", self.export_gear),
                ("Training Plans", self.export_training_plans),
                ("Workouts", self.export_workouts),
                ("Hydration", self.export_hydration),
                ("Nutrition", self.export_nutrition),
                ("Women's Health", self.export_womens_health),
            ]

        # Índice para facilitar la orientación de una IA.
        if _compact_mode:
            toc_info = [
                ("Export Metadata", "Intervalo, zona horaria, modo y versión del esquema"),
                ("Profile", "Contexto privado: edad, sexo, altura, unidades, zona horaria y reloj principal"),
                ("Daily Health", "Una fila por día: recuperación, sueño, VFC, estrés, batería corporal y actividad"),
                ("Blood Pressure", "Solo mediciones reales de presión arterial; se omite si está vacío"),
                ("Activities", "Actividades normalizadas con vueltas, zonas, autoevaluación, equipamiento y series opcionales"),
                ("Body Composition", "Mediciones del intervalo con unidades explícitas"),
                ("Training Metrics", "Valores históricos separados de fotografías actuales y métricas sin fecha"),
                ("Goals and Records", "Récords y objetivos activos; sin insignias ni objetivos pasados"),
                ("Gear", "Zapatillas, bicicletas y otro equipamiento sin datos personales"),
                ("Hydration", "Solo días con ingesta o pérdida de sudor"),
                ("Nutrition", "Solo días con alimentos realmente registrados"),
                ("Weekly Summary", "Totales calculados y distribución por zonas de pulso"),
                ("Data Quality", "Datos ausentes, avisos temporales, conversiones y privacidad"),
            ]
        else:
            toc_info = [
                ("Profile", "Información, ajustes, dispositivos, alarmas y tipos de actividad"),
                ("Daily Health", "Pasos, pulso, sueño, estrés, batería corporal, SpO2, VFC y respiración por día"),
                ("Blood Pressure", "Mediciones de presión arterial del intervalo"),
                ("Activities", "Resumen, vueltas, zonas, ejercicios, tiempo y series por actividad"),
                ("Body Composition", "Peso, IMC, grasa, músculo, hueso, agua y pesajes"),
                ("Training Metrics", "VO2 máx., edad física, preparación, estado, umbral, FTP y predicciones"),
                ("Goals and Records", "Récords, insignias y objetivos activos o pasados"),
                ("Trends", "Agregados semanales, pasos, pisos y progreso"),
                ("Golf", "Rondas, tarjetas y golpes"),
                ("Gear", "Equipamiento, estadísticas y valores predeterminados"),
                ("Training Plans", "Planes activos y pasados con sus detalles"),
                ("Workouts", "Entrenamientos guardados con su estructura completa"),
                ("Hydration", "Ingesta de líquidos por día"),
                ("Nutrition", "Alimentos, comidas y ajustes de nutrición por día"),
                ("Women's Health", "Calendario menstrual y resumen de embarazo"),
            ]
        self.md.append("Índice de contenidos\n")
        if self.update_mode:
            self.md.append(
                f"Este archivo contiene datos nuevos desde {self.update_base_date}."
            )
            self.md.append(
                "Cárgalo junto a la exportación base para disponer del historial."
            )
        else:
            self.md.append(
                "Este archivo contiene la exportación solicitada de salud y "
                "entrenamiento de Garmin Connect."
            )
        if _compact_mode:
            self.md.append(
                "Cada sección incluida contiene un bloque JSON semántico con "
                "una descripción del esquema."
            )
            self.md.append(
                "Se omiten secciones opcionales vacías y catálogos sin registros."
            )
        else:
            self.md.append(
                "Cada sección tiene subsecciones con bloques JSON titulados."
            )
            self.md.append(
                "Todos los datos son respuestas JSON originales de Garmin Connect."
            )
        self.md.append(
            "Las secciones vacías indican: No hay datos disponibles.\n"
        )
        for i, (name, desc) in enumerate(toc_info, 1):
            self.md.append(f"  {i}. {name} -- {desc}")
        self.md.append("")

        for name, fn in sections:
            log.info(f"Exportando {name}...")
            _set_safe_call_failure_handler(
                lambda endpoint, reason, section=name:
                    self._record_endpoint_failure(
                        section,
                        endpoint,
                        reason,
                    )
            )
            try:
                try:
                    fn()
                    log.info(f"  Completado: {name}")
                except KeyboardInterrupt:
                    log.info(f"\n  Interrumpido durante {name}; se guardará la exportación parcial")
                    self.errors.append(
                        f"{name}: interrumpido por la persona usuaria (datos parciales)"
                    )
                    break
                except Exception as e:
                    reason = _safe_exception_reason(e)
                    self.errors.append(f"{name}: fallo técnico ({reason})")
                    log.error(f"  Fallo en {name}: {reason}")
                    log.debug(
                        "La traza detallada se ha omitido para no exponer datos "
                        "devueltos por Garmin."
                    )
            finally:
                _set_safe_call_failure_handler(None)

        if self.errors:
            self.md.append("\nErrores durante la exportación\n")
            for err in self.errors:
                self.md.append(f"- {err}")
            self.md.append("")
        if not _compact_mode:
            with self._endpoint_failures_lock:
                endpoint_failures = sorted(self.endpoint_failures)
            if endpoint_failures:
                self.md.append("\nConsultas incompletas\n")
                for section, endpoint, reason in endpoint_failures:
                    self.md.append(
                        f"- {section}: {endpoint} no terminó correctamente "
                        f"({reason})"
                    )
                self.md.append("")

        # Las estadísticas se muestran solo en la consola.

        needs_text_output = self.output_format in {"txt", "both"}
        full_text = None
        if _compact_mode:
            self._finalize_semantic_model()
            if needs_text_output:
                full_text = self._render_compact_text()
        elif needs_text_output:
            full_text = "\n".join(self.md)

        written: list[Path] = []
        if needs_text_output:
            assert full_text is not None
            if _split_mode:
                written.extend(self._write_split(full_text, txt_filename))
            else:
                out_path = self.out / txt_filename
                atomic_write_text(out_path, full_text)
                written.append(out_path)

        if self.output_format in {"xlsx", "both"}:
            if not _compact_mode:
                raise ValueError("La salida XLSX requiere el modo compacto.")
            xlsx_path = self.out / xlsx_filename
            log.info("Creando Excel opcional...")
            xlsx_started = time.perf_counter()
            render_xlsx(self.semantic_model, xlsx_path)
            log.info(
                "Excel creado en %.1f segundos.",
                time.perf_counter() - xlsx_started,
            )
            written.append(xlsx_path)

        self.written_files = written
        print()
        result_label = (
            "Exportación parcial creada"
            if self._is_partial()
            else "Exportación completada"
        )
        if len(written) == 1:
            log.info(f"{result_label}: {written[0]}")
        else:
            log.info(
                f"{result_label}: {len(written)} archivos en {self.out}"
            )
        total_kb = sum(path.stat().st_size for path in written) / 1024
        log.info(
            f"Tamaño total: {total_kb:.0f} KB"
            + (f" en {len(written)} archivos" if len(written) > 1 else "")
        )

        if self.manifest_path:
            atomic_write_json(
                self.manifest_path,
                {
                    "schema_version": _COMPACT_SCHEMA_VERSION,
                    "run_id": self.run_id,
                    "status": (
                        "partial"
                        if self._is_partial()
                        else "completed"
                    ),
                    "report_type": self.report_type,
                    "output_format": self.output_format,
                    "file_stem": output_stem,
                    "start_date": self.start_date.isoformat(),
                    "end_date": self.today.isoformat(),
                    "files": _relative_manifest_paths(written, self.out),
                    "warnings": len(
                        (self.semantic_model.get("data_quality") or {}).get(
                            "issues",
                            [],
                        )
                    ),
                    "errors": self._partial_sections(),
                },
            )

        log.info(f"Llamadas a la API: {_limiter.call_count}")
        log.info(self.cache.summary())
        if self._is_partial():
            log.warning(
                f"{len(self._partial_sections())} secciones quedaron incompletas"
            )
        return written

    def _write_split(self, full_text: str, base_filename: str) -> list:
        """Divide la exportación en archivos bajo el límite de palabras.

        Corta por secciones y fragmenta las demasiado grandes por fechas o
        elementos. Utiliza TXT para mejorar la compatibilidad con herramientas
        RAG como NotebookLM.
        """
        # Nombres técnicos de sección que aparecen en una línea independiente.
        section_names_list = [
            "Export Metadata", "Race Context", "Profile", "Period Summary",
            "Weekly Timeline", "Daily Health", "Blood Pressure",
            "Activities", "Body Composition", "Race Analysis",
            "Suggested Prompts", "Journal",
            "Training Metrics", "Goals and Records", "Trends", "Golf",
            "Gear", "Training Plans", "Workouts", "Hydration", "Nutrition",
            "Weekly Summary", "Data Quality", "Women's Health",
            "Errors During Export", "Errores durante la exportación",
        ]
        # Crear una expresión que corte por líneas con nombres conocidos.
        escaped = [re.escape(n) for n in section_names_list]
        split_pattern = r'(?=\n(?:' + '|'.join(escaped) + r')\n)'
        parts = re.split(split_pattern, full_text)

        # La primera parte contiene título, fecha, formato e índice.
        file_header = parts[0] if parts else ""
        header_words = _word_count(file_header)
        raw_sections = parts[1:] if len(parts) > 1 else []

        # Fragmentar secciones grandes y conservar las pequeñas.
        section_chunks = []  # Lista de pares (nombre visible, texto).
        for sec_text in raw_sections:
            # Obtener el nombre desde la primera línea no vacía.
            sec_name = sec_text.strip().split('\n')[0].strip()
            wc = _word_count(sec_text)

            if wc <= _SPLIT_WORD_LIMIT * 0.85:
                section_chunks.append((sec_name, sec_text))
            else:
                sub = self._split_oversized_section(sec_text, sec_name)
                section_chunks.extend(sub)

        # Empaquetado secuencial dentro del límite.
        files = []  # Lista de listas con pares (nombre, texto).
        current_file = []
        current_words = header_words

        for name, text in section_chunks:
            chunk_words = _word_count(text)
            if current_words + chunk_words > _SPLIT_WORD_LIMIT and current_file:
                files.append(current_file)
                current_file = []
                current_words = header_words
            current_file.append((name, text))
            current_words += chunk_words

        if current_file:
            files.append(current_file)

        # Escribir cada archivo.
        total = len(files)
        written = []
        for i, file_sections in enumerate(files, 1):
            section_names = [n for n, _ in file_sections]

            header = (
                f"Exportación de datos de Garmin Connect — Parte {i} de "
                f"{total}\n\n"
            )
            header += (
                f"Secciones de este archivo: {', '.join(section_names)}\n"
            )
            header += (
                f"Carga las {total} partes en el mismo cuaderno para disponer "
                "de todos los datos.\n\n"
            )

            content = header + "\n".join(text for _, text in file_sections)

            suffix = f"_split_part{i}of{total}"
            fname = base_filename.replace(".txt", f"{suffix}.txt")
            path = self.out / fname
            atomic_write_text(path, content)

            wc = _word_count(content)
            size_kb = path.stat().st_size / 1024
            written.append(path)
            log.info(f"  Parte {i}/{total}: {fname} ({wc:,} palabras, {size_kb:.0f} KB)")

        return written

    def _split_oversized_section(self, sec_text: str, sec_name: str) -> list:
        """Fragmenta el JSON de una sección demasiado grande.

        Los objetos se dividen por grupos de claves o fechas; las listas, por
        cantidad de elementos.
        """
        # Localizar la línea de esquema y el bloque JSON.
        match = re.search(
            r'(Schema:[^\n]*\n)\s*(\{.*\}|\[.*\])',
            sec_text, re.DOTALL,
        )
        if not match:
            return [(sec_name, sec_text)]

        schema_line = match.group(1)
        json_str = match.group(2)

        try:
            data = json.loads(json_str)
        except (json.JSONDecodeError, TypeError):
            return [(sec_name, sec_text)]

        target_words = int(_SPLIT_WORD_LIMIT * 0.8)

        if isinstance(data, dict) and data:
            # Crear grupos según las palabras reales de cada clave.
            keys = list(data.keys())
            groups = []
            cur_keys = []
            cur_data = {}
            cur_words = 0

            for k in keys:
                item_str = json.dumps({k: data[k]}, default=str, ensure_ascii=False)
                item_words = _word_count(item_str)

                if cur_words + item_words > target_words and cur_keys:
                    groups.append((cur_keys[0], cur_keys[-1], cur_data))
                    cur_keys = []
                    cur_data = {}
                    cur_words = 0

                cur_keys.append(k)
                cur_data[k] = data[k]
                cur_words += item_words

            if cur_keys:
                groups.append((cur_keys[0], cur_keys[-1], cur_data))

            results = []
            for idx, (first, last, chunk_data) in enumerate(groups, 1):
                part_name = f"{sec_name} (Parte {idx} de {len(groups)}: {first} a {last})"
                text = f"{part_name}\n\n{schema_line}\n{_json(chunk_data)}\n"
                results.append((part_name, text))
            return results

        elif isinstance(data, list) and data:
            # Crear bloques según el número real de palabras de cada elemento.
            groups = []
            cur_items = []
            cur_start = 1
            cur_words = 0

            for i, item in enumerate(data):
                item_str = json.dumps(item, default=str, ensure_ascii=False)
                item_words = _word_count(item_str)

                if cur_words + item_words > target_words and cur_items:
                    groups.append((cur_start, cur_start + len(cur_items) - 1, cur_items))
                    cur_items = []
                    cur_start = i + 1
                    cur_words = 0

                cur_items.append(item)
                cur_words += item_words

            if cur_items:
                groups.append((cur_start, cur_start + len(cur_items) - 1, cur_items))

            results = []
            for idx, (start, end, chunk_data) in enumerate(groups, 1):
                part_name = f"{sec_name} (Parte {idx} de {len(groups)}: elementos {start}-{end})"
                text = f"{part_name}\n\n{schema_line}\n{_json(chunk_data)}\n"
                results.append((part_name, text))
            return results

        return [(sec_name, sec_text)]

    # ===================================================================
    # Metadatos del compacto semántico
    # ===================================================================
    def export_metadata(self):
        if not _compact_mode:
            return
        current, historical_offset = _timezone_metadata(
            self.timezone_name,
            self.today,
        )
        self.md.append("\nExport Metadata\n")
        self.md.append(
            'Schema: "Metadatos versionados de esta exportación compacta semántica."'
        )
        metadata = {
            "exported_at": current.isoformat(),
            "requested_start_date": self.start_date.isoformat(),
            "requested_end_date": self.today.isoformat(),
            "timezone": self.timezone_name,
            "utc_offset": historical_offset,
            "utc_offset_date": self.today.isoformat(),
            "mode": "compact",
            "report_type": self.report_type,
            "activity_series_mode": (
                "full" if self.include_activity_details else "none"
            ),
            "export_variant": (
                "compact_with_full_activity_series"
                if self.include_activity_details
                else "compact_summary"
            ),
            "free_text_included": self.include_free_text,
            "privacy_mode": "redact_personal_identifiers",
            "schema_version": _COMPACT_SCHEMA_VERSION,
        }
        self._remember_compact("export_metadata", metadata)
        self.md.append(_json({"export_metadata": metadata}))
        self.md.append("")

    # ===================================================================
    # Perfil
    # ===================================================================
    def export_profile(self):
        cached = self.cache.get_section("profile")
        if cached is not None and not self.update_mode:
            data = cached
        else:
            failures_before_cache = self._endpoint_failure_count()
            data = {}
            data["full_name"] = safe_call(self.api.get_full_name, label="full_name")
            data["unit_system"] = safe_call(self.api.get_unit_system, label="unit_system")
            data["user_profile"] = safe_call(self.api.get_user_profile, label="user_profile")
            data["profile_settings"] = safe_call(self.api.get_userprofile_settings, label="profile_settings")
            data["devices"] = safe_call(self.api.get_devices, label="devices")
            data["primary_device"] = safe_call(self.api.get_primary_training_device, label="primary_device")
            data["device_alarms"] = safe_call(self.api.get_device_alarms, label="device_alarms")
            data["last_used_device"] = safe_call(self.api.get_device_last_used, label="last_used_device")
            data["activity_types"] = safe_call(self.api.get_activity_types, label="activity_types")
            self.cache.put_section(
                "profile",
                data,
                complete=(
                    self._endpoint_failure_count()
                    == failures_before_cache
                ),
            )

        if _compact_mode:
            self._remember_sensitive_payload(data)
            profile = _compact_profile(
                data,
                self.today,
                timezone_name=self.timezone_name,
            )
            self.md.append("\nProfile\n")
            self.md.append(
                'Schema: "Contexto deportivo respetuoso con la privacidad. Se excluyen '
                'fecha de nacimiento exacta, nombres, IDs de usuario o dispositivo, '
                'números de serie, URLs, alarmas, capacidades y catálogos."\n'
            )
            self._remember_compact("profile", profile)
            self.md.append(f"{_json({'profile': profile})}\n")
        else:
            self.md.append("\nProfile\n")
            for title, key in [("Nombre completo", "full_name"), ("Sistema de unidades", "unit_system"),
                               ("Perfil de usuario", "user_profile"), ("Ajustes del perfil", "profile_settings"),
                               ("Dispositivos", "devices"), ("Dispositivo principal de entrenamiento", "primary_device"),
                               ("Alarmas de dispositivos", "device_alarms"), ("Último dispositivo usado", "last_used_device"),
                               ("Tipos de actividad", "activity_types")]:
                _section(self.md, title, data.get(key))

    # ===================================================================
    # Salud diaria: una sección por día con las respuestas completas.
    # ===================================================================
    def export_daily_health(self):
        weeks = self.days / 7
        months = self.days / 30.44
        log.info(f"  {self.days} días para procesar ({months:.1f} meses / {weeks:.0f} semanas)")
        log.info(f"  Periodo: {self.start_date} a {self.today}")
        log.info("  13 llamadas por día, descargando 4 simultáneamente")

        self.md.append("\nDaily Health\n")

        # Claves de endpoints en el orden de presentación.
        endpoint_keys = [
            "summary", "heart_rate", "rhr", "sleep", "stress", "spo2",
            "respiration", "hrv", "body_battery", "bb_events",
            "intensity_min", "events", "lifestyle",
        ]

        display_names = {
            "summary": "Resumen diario", "heart_rate": "Frecuencia cardiaca",
            "rhr": "Frecuencia cardiaca en reposo", "sleep": "Sueño", "stress": "Estrés",
            "spo2": "Oxígeno en sangre (SpO2)", "respiration": "Respiración",
            "hrv": "Variabilidad de la frecuencia cardiaca", "body_battery": "Batería corporal",
            "bb_events": "Eventos de batería corporal", "intensity_min": "Minutos de intensidad",
            "events": "Eventos del día", "lifestyle": "Registro de estilo de vida",
        }

        def _fetch_endpoint(key, ds):
            """Consulta un endpoint para una fecha; se ejecuta en un hilo."""
            api = self.api
            if key == "summary":
                return _safe_call_with_status(
                    api.get_user_summary,
                    ds,
                    label=f"summary_{ds}",
                )
            elif key == "heart_rate":
                return _safe_call_with_status(
                    api.get_heart_rates,
                    ds,
                    label=f"hr_{ds}",
                )
            elif key == "rhr":
                return _safe_call_with_status(
                    api.get_rhr_day,
                    ds,
                    label=f"rhr_{ds}",
                )
            elif key == "sleep":
                payload, succeeded = _safe_call_with_status(
                    api.get_sleep_data,
                    ds,
                    label=f"sleep_{ds}",
                )
                _debug_health_payload("sleep", ds, payload, "api")
                return payload, succeeded
            elif key == "stress":
                return _safe_call_with_status(
                    api.get_all_day_stress,
                    ds,
                    label=f"stress_{ds}",
                )
            elif key == "spo2":
                return _safe_call_with_status(
                    api.get_spo2_data,
                    ds,
                    label=f"spo2_{ds}",
                )
            elif key == "respiration":
                return _safe_call_with_status(
                    api.get_respiration_data,
                    ds,
                    label=f"resp_{ds}",
                )
            elif key == "hrv":
                payload, succeeded = _safe_call_with_status(
                    api.get_hrv_data,
                    ds,
                    label=f"hrv_{ds}",
                )
                _debug_health_payload("hrv", ds, payload, "api")
                return payload, succeeded
            elif key == "body_battery":
                return _safe_call_with_status(
                    api.get_body_battery,
                    ds,
                    ds,
                    label=f"bb_{ds}",
                )
            elif key == "bb_events":
                return _safe_call_with_status(
                    api.get_body_battery_events,
                    ds,
                    label=f"bbe_{ds}",
                )
            elif key == "intensity_min":
                return _safe_call_with_status(
                    api.get_intensity_minutes_data,
                    ds,
                    label=f"im_{ds}",
                )
            elif key == "events":
                return _safe_call_with_status(
                    api.get_all_day_events,
                    ds,
                    label=f"events_{ds}",
                )
            elif key == "lifestyle":
                return _safe_call_with_status(
                    api.get_lifestyle_logging_data,
                    ds,
                    label=f"ll_{ds}",
                )
            return None, True

        t_start = time.time()
        cached_days = 0
        if _compact_mode:
            all_days = []

        for i in range(self.days):
            d = self.today - timedelta(days=i)
            ds = d.isoformat()

            # Consultar primero la caché.
            day_data, complete_keys = self.cache.get_day_entry(ds)
            if day_data is not None:
                cached_days += 1
                _debug_health_payload("sleep", ds, day_data.get("sleep"), "cache")
                _debug_health_payload("hrv", ds, day_data.get("hrv"), "cache")
            else:
                day_data = {}
                complete_keys = set()

            recent_cutoff = self.today - timedelta(days=2)
            keys_to_fetch = (
                endpoint_keys
                if self.update_mode or d >= recent_cutoff
                else [
                    key
                    for key in endpoint_keys
                    if key not in complete_keys
                ]
            )
            if keys_to_fetch:
                # Consultar en paralelo lo reciente o lo que falló anteriormente.
                with ThreadPoolExecutor(max_workers=4) as pool:
                    futures = {
                        pool.submit(_fetch_endpoint, key, ds): key
                        for key in keys_to_fetch
                    }
                    for future in as_completed(futures):
                        key = futures[future]
                        try:
                            value, succeeded = future.result()
                            if succeeded:
                                day_data[key] = value
                                complete_keys.add(key)
                        except Exception as exc:
                            self._record_endpoint_failure(
                                "Daily Health",
                                "worker",
                                _safe_exception_reason(exc),
                            )
                self.cache.put_day(
                    ds,
                    day_data,
                    complete_keys=complete_keys,
                )

            # Añadir los datos al documento.
            if _compact_mode:
                record = _compact_daily_record(
                    ds,
                    day_data,
                    timezone_name=self.timezone_name,
                    quality_callback=self._quality_add,
                    include_free_text=self.include_free_text,
                )
                all_days.append(record)
                self.compact_daily_records.append(record)
                sleep = record.get("sleep", {})
                if not sleep.get("valid_sleep"):
                    self._quality_add(
                        "missing_critical_data",
                        f"No hay sueño real disponible para {ds}.",
                    )
                if not record.get("hrv"):
                    self._quality_add(
                        "missing_critical_data",
                        f"No hay VFC disponible para {ds}.",
                    )
            else:
                self.md.append(f"{ds}\n")
                for key in endpoint_keys:
                    _section(self.md, display_names[key], day_data.get(key), 4)

            # Informar del progreso con más frecuencia al principio.
            done = i + 1
            report_interval = 5 if done <= 25 else 25
            if done % report_interval == 0 or done == 1 or done == self.days:
                elapsed = time.time() - t_start
                d_display = d.isoformat()
                if done > cached_days and (done - cached_days) > 0:
                    fetched = done - cached_days
                    per_day = elapsed / fetched
                    remaining_fetch = max(0, self.days - done)
                    eta_sec = remaining_fetch * per_day
                    eta_min = eta_sec / 60
                    log.info(f"  {done}/{self.days} días ({d_display}) | "
                             f"{cached_days} en caché | "
                             f"{_limiter.call_count} llamadas | quedan ~{eta_min:.0f} min")
                else:
                    log.info(f"  {done}/{self.days} días ({d_display}) | "
                             f"{cached_days} en caché (todos hasta ahora)")

        if _compact_mode:
            if all_days:
                all_days.sort(key=lambda item: item.get("date", ""))
                self.compact_daily_records = list(all_days)
                self._remember_compact("daily_health", all_days)
                self.md.append(
                    'Schema: "Un objeto semántico y privado por fecha. Se omiten series '
                    'de bienestar de alta frecuencia, campos duplicados y catálogos de '
                    'hábitos no registrados. La falta de sueño o VFC aparece en Data Quality."\n'
                )
                self.md.append(f"{_json(all_days)}\n")
                if any(
                    day.get("sleep", {}).get("sleep_need_s") is not None
                    for day in all_days
                ):
                    self._quality_add(
                        "unit_conversions",
                        "dailySleepDTO.sleepNeed.actual se convirtió de minutos a segundos.",
                    )
                if any(
                    day.get("sleep", {}).get("sleep_start_local")
                    for day in all_days
                ):
                    self._quality_add(
                        "unit_conversions",
                        "Los epochs de sueño de Garmin se interpretaron explícitamente como milisegundos UTC y se convirtieron a ISO 8601 con la zona IANA configurada.",
                    )
                self._quality_add(
                    "warnings",
                    "Sueño y VFC se consultaron con get_sleep_data y get_hrv_data; si Garmin no devolvió dailySleepDTO válido o hrvSummary, se marcó la ausencia sin inventar valores.",
                )
            else:
                _section_nodata(self.md, "Daily Health")

    # ===================================================================
    # Presión arterial: una respuesta por intervalo en JSON original.
    # ===================================================================
    def export_blood_pressure(self):
        cache_key = self._range_cache_key("blood_pressure")
        cached = self.cache.get_section(cache_key)
        if cached is not None and not self.update_mode:
            data = cached
        else:
            data = safe_call(
                self.api.get_blood_pressure,
                self.start_date.isoformat(),
                self.today.isoformat(),
                label="blood_pressure",
            )
            if data is not None:
                self.cache.put_section(cache_key, data)

        if _compact_mode:
            measurements = _find_blood_pressure_measurements(data)
            if not measurements:
                self._quality_add(
                    "warnings",
                    "El endpoint de presión arterial no devolvió mediciones reales.",
                )
                return
            measurements.sort(key=lambda item: str(item.get("timestamp") or ""))
            self._remember_compact("blood_pressure", measurements)
            self.md.append("\nBlood Pressure\n")
            self.md.append(
                'Schema: "Solo mediciones reales con valores sistólico y diastólico; '
                'el intervalo consultado no se trata como dato de salud."\n'
            )
            self.md.append(f"{_json({'measurements': measurements})}\n")
        else:
            self.md.append("\nBlood Pressure\n")
            if data is None:
                _section_nodata(self.md, "Blood Pressure")
            else:
                _section(self.md, "Mediciones de presión arterial", data)

    # ===================================================================
    # Actividades: datos completos para cada actividad.
    # ===================================================================
    def export_activities(self):
        self.md.append("\nActivities\n")

        if self.fetch_all or self.update_mode or self.explicit_start_date is not None:
            activities = safe_call(
                self.api.get_activities_by_date,
                self.start_date.isoformat(), self.today.isoformat(), None,
                label="activities_all",
            ) or []
        else:
            activities = safe_call(
                self.api.get_activities, 0, self.max_activities,
                label="activities_list",
            ) or []
            if not activities:
                activities = safe_call(
                    self.api.get_activities_by_date,
                    self.start_date.isoformat(), self.today.isoformat(), "",
                    label="activities_by_date",
                ) or []

        if self.selected_activity_ref:
            selected = str(self.selected_activity_ref)
            activities = [
                activity
                for activity in activities
                if private_reference(
                    "activity",
                    activity.get("activityId"),
                    self.reference_secret,
                )
                == selected
            ]
            if not activities:
                raise ValueError(
                    "No se encontró la actividad seleccionada en el periodo. "
                    "Actualiza la lista de actividades y vuelve a intentarlo."
                )

        self.md.append(f"Total de actividades encontradas: {len(activities)}\n")
        calls_per_activity = 11 if _compact_mode else 10
        log.info(
            f"  {len(activities)} actividades encontradas, hasta "
            f"{calls_per_activity} llamadas por actividad = "
            f"{len(activities) * calls_per_activity:,} llamadas"
        )

        t_start = time.time()
        cached_acts = 0
        if _compact_mode:
            all_activities = []

        for i, act in enumerate(activities):
            failures_before_activity = self._endpoint_failure_count()
            aid = act.get("activityId", i)
            name = act.get("activityName") or "Sin nombre"
            if _compact_mode:
                self._remember_sensitive_payload(act)
            if act.get("activityId") is not None:
                self.sensitive_identifiers.add(str(act.get("activityId")))
            if act.get("activityName"):
                self.sensitive_activity_names.add(str(act.get("activityName")))
            atype = (act.get("activityType", {}).get("typeKey", "?")
                     if isinstance(act.get("activityType"), dict)
                     else str(act.get("activityType", "?")))
            start = act.get("startTimeLocal", "")

            if not _compact_mode:
                self.md.append(f"Actividad {aid}: {name}\n")
                self.md.append(f"Tipo: {atype} | Fecha: {start}\n")

            # Consultar la caché.
            act_data = self.cache.get_activity(aid)
            if act_data is not None:
                cached_acts += 1
                # La lista reciente puede contener metadatos ya corregidos.
                act_data["summary"] = act
                cache_changed = True
                activity_date = None
                if isinstance(start, str):
                    try:
                        activity_date = date.fromisoformat(start[:10])
                    except ValueError:
                        pass
                recent_cutoff = self.today - timedelta(
                    days=_RECENT_ACTIVITY_REFRESH_DAYS
                )
                is_recent = (
                    activity_date is not None
                    and activity_date >= recent_cutoff
                )
                if is_recent:
                    refreshed_detail, succeeded = _safe_call_with_status(
                        self.api.get_activity,
                        aid,
                        label=f"act_refresh_{aid}",
                    )
                    if succeeded:
                        act_data["detail"] = refreshed_detail
                        cache_changed = True

                wants_full_series = (
                    self.include_activity_details or not _compact_mode
                )
                cached_detail_limit = act_data.get(
                    "_details_maxchart_requested"
                )
                full_series_was_requested = (
                    isinstance(cached_detail_limit, int)
                    and not isinstance(cached_detail_limit, bool)
                    and cached_detail_limit >= _MAX_ACTIVITY_CHART_POINTS
                )
                should_refresh_series = (
                    wants_full_series
                    and (
                        self.selected_activity_ref is not None
                        or is_recent
                        or not full_series_was_requested
                    )
                )
                if should_refresh_series:
                    refreshed_series, succeeded = _safe_call_with_status(
                        self.api.get_activity_details,
                        aid,
                        maxchart=_MAX_ACTIVITY_CHART_POINTS,
                        label=f"details_refresh_{aid}",
                    )
                    if succeeded:
                        act_data["details"] = refreshed_series
                        act_data[
                            "_details_maxchart_requested"
                        ] = _MAX_ACTIVITY_CHART_POINTS
                        cache_changed = True

                if _compact_mode:
                    should_refresh_gear = (
                        "gear" not in act_data
                        or is_recent
                    )
                    if (
                        should_refresh_gear
                        and hasattr(self.api, "get_activity_gear")
                    ):
                        refreshed_gear, succeeded = _safe_call_with_status(
                            self.api.get_activity_gear,
                            aid,
                            label=f"activity_gear_{aid}",
                        )
                        if succeeded:
                            act_data["gear"] = refreshed_gear
                            cache_changed = True
                    elif should_refresh_gear:
                        self._quality_add(
                            "warnings",
                            "La versión instalada no permite consultar el equipamiento por actividad.",
                        )
                if cache_changed:
                    self.cache.put_activity(
                        aid,
                        act_data,
                        complete=(
                            self._endpoint_failure_count()
                            == failures_before_activity
                        ),
                    )
            else:
                # Consultar la API y guardar en caché.
                act_data = {"summary": act}
                act_data["detail"] = safe_call(self.api.get_activity, aid, label=f"act_{aid}")
                act_data["splits"] = safe_call(self.api.get_activity_splits, aid, label=f"splits_{aid}")
                act_data["split_summaries"] = safe_call(self.api.get_activity_split_summaries, aid, label=f"ss_{aid}")
                act_data["typed_splits"] = safe_call(self.api.get_activity_typed_splits, aid, label=f"typed_splits_{aid}")
                act_data["weather"] = safe_call(self.api.get_activity_weather, aid, label=f"wx_{aid}")
                act_data["hr_zones"] = safe_call(self.api.get_activity_hr_in_timezones, aid, label=f"hrz_{aid}")
                act_data["power_zones"] = safe_call(self.api.get_activity_power_in_timezones, aid, label=f"pwrz_{aid}")
                act_data["exercise_sets"] = safe_call(self.api.get_activity_exercise_sets, aid, label=f"sets_{aid}")
                detail_limit = (
                    _MAX_ACTIVITY_CHART_POINTS
                    if self.include_activity_details or not _compact_mode
                    else 2000
                )
                details, details_succeeded = _safe_call_with_status(
                    self.api.get_activity_details,
                    aid,
                    maxchart=detail_limit,
                    label=f"details_{aid}",
                )
                act_data["details"] = details
                if details_succeeded:
                    act_data["_details_maxchart_requested"] = detail_limit
                if _compact_mode and hasattr(self.api, "get_activity_gear"):
                    act_data["gear"] = safe_call(
                        self.api.get_activity_gear,
                        aid,
                        label=f"activity_gear_{aid}",
                    )
                self.cache.put_activity(
                    aid,
                    act_data,
                    complete=(
                        self._endpoint_failure_count()
                        == failures_before_activity
                    ),
                )

            if _compact_mode:
                for gear_item in _as_list(act_data.get("gear")):
                    if not isinstance(gear_item, dict):
                        continue
                    gear_id = _pick(
                        gear_item,
                        "uuid",
                        "gearUUID",
                        "gearId",
                        "gearPk",
                        "id",
                    )
                    if gear_id is not None:
                        self.sensitive_identifiers.add(str(gear_id))
                compact_activity = _compact_activity(
                    act_data,
                    include_series=self.include_activity_details,
                    quality_callback=self._quality_add,
                    reference_secret=self.reference_secret,
                    include_free_text=self.include_free_text,
                )
                if compact_activity:
                    all_activities.append(compact_activity)
                    self.compact_activities.append(compact_activity)
            else:
                _section(self.md, "Resumen de la actividad", act_data.get("summary"), 4)
                _section(self.md, "Detalle completo de la actividad", act_data.get("detail"), 4)
                _section(self.md, "Vueltas", act_data.get("splits"), 4)
                _section(self.md, "Resúmenes de vueltas", act_data.get("split_summaries"), 4)
                _section(self.md, "Vueltas tipificadas", act_data.get("typed_splits"), 4)
                _section(self.md, "Meteorología", act_data.get("weather"), 4)
                _section(self.md, "Zonas de frecuencia cardiaca", act_data.get("hr_zones"), 4)
                _section(self.md, "Zonas de potencia", act_data.get("power_zones"), 4)
                _section(self.md, "Series de ejercicios", act_data.get("exercise_sets"), 4)
                _section(self.md, "Detalle de las series temporales", act_data.get("details"), 4)

            done = i + 1
            if done % 10 == 0 or done == len(activities):
                elapsed = time.time() - t_start
                if done > cached_acts:
                    fetched = done - cached_acts
                    per_act = elapsed / fetched
                    remaining = max(0, len(activities) - done)
                    eta_min = (remaining * per_act) / 60
                    log.info(f"  {done}/{len(activities)} actividades | {cached_acts} en caché | "
                             f"{_limiter.call_count} llamadas | quedan ~{eta_min:.0f} min")
                else:
                    log.info(f"  {done}/{len(activities)} actividades | {cached_acts} en caché")

        if _compact_mode:
            if all_activities:
                all_activities.sort(key=lambda item: item.get("date", ""))
                self.compact_activities = list(all_activities)
                self._remember_compact("activities", all_activities)
                self.md.append(
                    'Schema: "Actividades normalizadas con título, elevación, desnivel, '
                    'GAP/RAP, vueltas, zonas, equipamiento y autoevaluación. El GPS, '
                    'track, ubicaciones y fuente original dependen del modo de privacidad; '
                    'las métricas deportivas derivadas se conservan siempre."\n'
                )
                self.md.append(f"{_json(all_activities)}\n")
                self._quality_add(
                    "duplicate_sources_removed",
                    "Las vueltas usan splits.lapDTOs y typed_splits.splits solo como alternativa.",
                )
                self._quality_add(
                    "duplicate_sources_removed",
                    "Se unificaron summary y detail.summaryDTO de cada actividad.",
                )
                if any(
                    activity.get("self_evaluation") is not None
                    for activity in all_activities
                ):
                    self._quality_add(
                        "unit_conversions",
                        "directWorkoutRpe se interpretó en pasos de diez (10–100) y se normalizó a perceived_exertion_1_10; el valor raw queda en self_evaluation.",
                    )
                if any(
                    activity.get("average_temperature_source_unit") == "fahrenheit"
                    for activity in all_activities
                ):
                    self._quality_add(
                        "unit_conversions",
                        "weather.temp se convirtió de Fahrenheit a Celsius; las temperaturas directas del sensor se conservaron como Celsius.",
                    )
            else:
                _section_nodata(self.md, "Activities")

    # ===================================================================
    # Composición corporal
    # ===================================================================
    def export_body_composition(self):
        cache_key = self._range_cache_key("body_comp")
        cached = self.cache.get_section(cache_key)
        if cached is not None and not self.update_mode:
            data = cached
        else:
            failures_before_cache = self._endpoint_failure_count()
            data = {}
            data["body_comp"] = _chunked_date_call(self.api.get_body_composition,
                                                   self.start_date, self.today, "body_comp")
            data["weigh_ins"] = _chunked_date_call(self.api.get_weigh_ins,
                                                   self.start_date, self.today, "weigh_ins")
            self.cache.put_section(
                cache_key,
                data,
                complete=(
                    self._endpoint_failure_count()
                    == failures_before_cache
                ),
            )

        if _compact_mode:
            measurements = _compact_body_composition(data)
            if not measurements:
                return
            measurements.sort(key=lambda item: str(item.get("date") or ""))
            self._remember_compact("body_composition", measurements)
            self.md.append("\nBody Composition\n")
            self.md.append(
                'Schema: "Mediciones únicas del intervalo. Peso, masa muscular y masa '
                'ósea de Garmin se convierten de gramos a kilogramos."\n'
            )
            self.md.append(
                f"{_json({'measurements': measurements})}\n"
            )
            self._quality_add(
                "unit_conversions",
                "Peso, masa muscular y masa ósea: gramos de Garmin convertidos a kilogramos.",
            )
        else:
            self.md.append("\nBody Composition\n")
            _section(self.md, "Composición corporal", data.get("body_comp"))
            _section(self.md, "Pesajes", data.get("weigh_ins"))

    # ===================================================================
    # Métricas de entrenamiento
    # ===================================================================
    def export_training(self):
        self.md.append("\nTraining Metrics\n")

        cache_key = self._range_cache_key("training")
        cached = self.cache.get_section(cache_key)
        if cached is not None and not self.update_mode:
            data = cached
        else:
            failures_before_cache = self._endpoint_failure_count()
            today_s = self.today.isoformat()
            start_s = self.start_date.isoformat()

            items = [
                ("training_readiness", "Preparación para entrenar",
                 safe_call(self.api.get_training_readiness, today_s, label="training_readiness")),
                ("morning_readiness", "Preparación matinal para entrenar",
                 safe_call(self.api.get_morning_training_readiness, today_s, label="morning_readiness")),
                ("training_status", "Estado de entrenamiento",
                 safe_call(self.api.get_training_status, today_s, label="training_status")),
                ("max_metrics", "VO2 máximo y métricas máximas",
                 safe_call(self.api.get_max_metrics, today_s, label="max_metrics")),
                ("fitness_age", "Edad física",
                 safe_call(self.api.get_fitnessage_data, today_s, label="fitness_age")),
                ("lactate_threshold", "Umbral de lactato",
                 safe_call(self.api.get_lactate_threshold, label="lactate_threshold")),
                ("cycling_ftp", "FTP de ciclismo",
                 safe_call(self.api.get_cycling_ftp, label="cycling_ftp")),
                ("intensity_min", "Minutos de intensidad",
                 safe_call(self.api.get_intensity_minutes_data, today_s, label="intensity_min")),
                ("hill_score", "Puntuación de pendientes",
                 _chunked_date_call(self.api.get_hill_score, self.start_date, self.today, "hill_score")),
                ("endurance_score", "Puntuación de resistencia",
                 _chunked_date_call(self.api.get_endurance_score, self.start_date, self.today, "endurance_score")),
                ("running_tolerance", "Tolerancia de carrera",
                 _chunked_date_call(self.api.get_running_tolerance, self.start_date, self.today, "running_tolerance")),
                ("race_predictions", "Predicciones de carrera",
                 safe_call(self.api.get_race_predictions, label="race_predictions")),
            ]

            data = {}
            for key, title, result in items:
                data[key] = result
                data[f"_title_{key}"] = title
            self.cache.put_section(
                cache_key,
                data,
                complete=(
                    self._endpoint_failure_count()
                    == failures_before_cache
                ),
            )

        if _compact_mode:
            compact_data, snapshots = _compact_training(
                data,
                self.start_date,
                self.today,
            )
            self.compact_training = compact_data
            self._remember_compact("training_metrics", compact_data)
            self.md.append(
                'Schema: "Métricas clasificadas como datos del periodo, último dato '
                'anterior, fotografía actual posterior o sin fecha. Se eliminan '
                'identidad e identificadores de dispositivos."\n'
            )
            self.md.append(f"{_json(compact_data)}\n")
            for snapshot in snapshots:
                message = (
                    f"{snapshot['metric']} se separó como current_snapshot "
                    f"porque contiene fechas posteriores al periodo: "
                    f"{', '.join(snapshot['effective_dates'])}."
                )
                self._quality_add("current_snapshots_detected", message)
            self._quality_add(
                "warnings",
                "recoveryTime y hrvWeeklyAverage se conservan como valores raw hasta confirmar sus unidades.",
            )
            if "lactate_threshold" in json.dumps(compact_data, ensure_ascii=False):
                self._quality_add(
                    "unit_conversions",
                    "La velocidad del endpoint lactate-threshold se interpretó como décimas de m/s, se multiplicó por 10 y se conserva también como speed_raw con su unidad de origen.",
                )
        else:
            for key in ["training_readiness", "morning_readiness", "training_status",
                        "max_metrics", "fitness_age", "lactate_threshold", "cycling_ftp",
                        "intensity_min", "hill_score", "endurance_score", "running_tolerance",
                        "race_predictions"]:
                _section(self.md, data.get(f"_title_{key}", key), data.get(key))

    # ===================================================================
    # Objetivos y récords
    # ===================================================================
    def export_goals(self):
        self.md.append("\nGoals and Records\n")

        cached = self.cache.get_section("goals")
        if cached is not None and not self.update_mode:
            data = cached
        else:
            failures_before_cache = self._endpoint_failure_count()
            data = {}
            data["personal_records"] = safe_call(self.api.get_personal_record, label="personal_records")
            data["badges"] = safe_call(self.api.get_earned_badges, label="badges")
            data["active_goals"] = safe_call(self.api.get_goals, "active", 0, 100, label="active_goals")
            data["past_goals"] = safe_call(self.api.get_goals, "past", 0, 100, label="past_goals")
            self.cache.put_section(
                "goals",
                data,
                complete=(
                    self._endpoint_failure_count()
                    == failures_before_cache
                ),
            )

        if _compact_mode:
            compact_data = _compact_personal_records(
                data,
                include_free_text=self.include_free_text,
            )
            self._remember_compact("goals_and_records", compact_data or {})
            self.md.append(
                'Schema: "Récords personales y objetivos activos reducidos. Se omiten '
                'insignias y objetivos pasados en las exportaciones semanales para IA."\n'
            )
            self.md.append(f"{_json(compact_data or {})}\n")
        else:
            _section(self.md, "Récords personales", data.get("personal_records"))
            _section(self.md, "Insignias conseguidas", data.get("badges"))
            _section(self.md, "Objetivos activos", data.get("active_goals"))
            _section(self.md, "Objetivos anteriores", data.get("past_goals"))

    # ===================================================================
    # Tendencias
    # ===================================================================
    def export_trends(self):
        self.md.append("\nTrends\n")

        cache_key = self._range_cache_key("trends")
        cached = self.cache.get_section(cache_key)
        if cached is not None and not self.update_mode:
            data = cached
        else:
            failures_before_cache = self._endpoint_failure_count()
            start_s = self.start_date.isoformat()
            today_s = self.today.isoformat()

            data = {}
            data["daily_steps"] = safe_call(self.api.get_daily_steps, start_s, today_s, label="daily_steps")
            data["weekly_steps"] = safe_call(self.api.get_weekly_steps, today_s, 52, label="weekly_steps")
            data["weekly_stress"] = safe_call(self.api.get_weekly_stress, today_s, 52, label="weekly_stress")
            data["weekly_im"] = _chunked_date_call(self.api.get_weekly_intensity_minutes,
                                                    self.start_date, self.today, "weekly_im")
            data["floors"] = safe_call(self.api.get_floors, start_s, label="floors")

            for metric in ("distance", "duration", "elevationGain", "calories"):
                result = safe_call(
                    self.api.get_progress_summary_between_dates,
                    start_s, today_s, metric, True,
                    label=f"progress_{metric}",
                )
                data[f"progress_{metric}"] = result

            # El endpoint de batería corporal por intervalo devuelve 400.
            # Los valores diarios ya están en las secciones de salud.
            data["bb_range"] = None
            self.cache.put_section(
                cache_key,
                data,
                complete=(
                    self._endpoint_failure_count()
                    == failures_before_cache
                ),
            )

        if _compact_mode:
            compact_data = {k: v for k, v in data.items() if k != "bb_range"}
            self.md.append('Schema: "daily_steps, weekly_steps (52 semanas), weekly_stress (52 semanas), weekly_im (minutos de intensidad), floors, progress_distance, progress_duration, progress_elevationGain y progress_calories."\n')
            self.md.append(f"{_json(compact_data)}\n")
        else:
            _section(self.md, "Pasos diarios", data.get("daily_steps"))
            _section(self.md, "Pasos semanales (52 semanas)", data.get("weekly_steps"))
            _section(self.md, "Estrés semanal (52 semanas)", data.get("weekly_stress"))
            _section(self.md, "Minutos de intensidad semanales", data.get("weekly_im"))
            _section(self.md, "Pisos subidos", data.get("floors"))
            for metric in ("distance", "duration", "elevationGain", "calories"):
                _section(self.md, f"Progreso: {metric}", data.get(f"progress_{metric}"))

    # ===================================================================
    # Golf
    # ===================================================================
    def export_golf(self):
        self.md.append("\nGolf\n")

        cached = self.cache.get_section("golf")
        if cached is not None and not self.update_mode:
            data = cached
        else:
            failures_before_cache = self._endpoint_failure_count()
            data = {}
            summary = safe_call(self.api.get_golf_summary, label="golf_summary")
            data["summary"] = summary

            scorecards = []
            if summary and isinstance(summary, list):
                for item in summary:
                    sc_id = item.get("scorecardId") or item.get("id")
                    if not sc_id:
                        continue
                    sc = {"_id": sc_id}
                    sc["detail"] = safe_call(self.api.get_golf_scorecard, sc_id, label=f"golf_sc_{sc_id}")
                    sc["shots"] = safe_call(self.api.get_golf_shot_data, sc_id, label=f"golf_shots_{sc_id}")
                    scorecards.append(sc)

            data["scorecards"] = scorecards
            self.cache.put_section(
                "golf",
                data,
                complete=(
                    self._endpoint_failure_count()
                    == failures_before_cache
                ),
            )

        if not data.get("summary") and not data.get("scorecards"):
            _section_nodata(self.md, "Golf")
        elif _compact_mode:
            self.md.append('Schema: "summary: lista de rondas. scorecards: matriz {_id, detail, shots} con los datos de cada ronda."\n')
            self.md.append(f"{_json(data)}\n")
        else:
            _section(self.md, "Resumen de golf", data.get("summary"))
            for sc in data.get("scorecards", []):
                _section(self.md, f"Tarjeta {sc.get('_id', '?')}", sc.get("detail"))
                _section(self.md, f"Datos de golpes {sc.get('_id', '?')}", sc.get("shots"))

    # ===================================================================
    # Equipamiento
    # ===================================================================
    def export_gear(self):
        cached = self.cache.get_section("gear")
        cached_has_result = (
            isinstance(cached, dict)
            and cached.get("gear_list") is not None
        )
        should_refresh = (
            not cached_has_result
            or self.cache.section_needs_refresh(
                "gear",
                _GEAR_REFRESH_DAYS,
            )
        )
        if cached_has_result and not should_refresh:
            data = cached
        else:
            # Los endpoints de equipamiento necesitan el número de perfil.
            profile, profile_succeeded = _safe_call_with_status(
                self.api.get_user_profile,
                label="gear_profile",
            )
            profile_num = None
            if profile and isinstance(profile, dict):
                profile_num = str(
                    profile.get("profileNumber")
                    or profile.get("userProfileNumber")
                    or profile.get("id")
                    or ""
                )

            data = {}
            refresh_complete = profile_succeeded and bool(profile_num)
            if profile_num:
                data["gear_list"], list_succeeded = _safe_call_with_status(
                    self.api.get_gear,
                    profile_num,
                    label="gear_list",
                )
                data["gear_defaults"], defaults_succeeded = (
                    _safe_call_with_status(
                        self.api.get_gear_defaults,
                        profile_num,
                        label="gear_defaults",
                    )
                )
                gear_list_valid = isinstance(data["gear_list"], list)
                refresh_complete = (
                    refresh_complete
                    and list_succeeded
                    and gear_list_valid
                    and defaults_succeeded
                )
            else:
                data["gear_list"] = None
                data["gear_defaults"] = None

            gear_details = []
            if data["gear_list"] and isinstance(data["gear_list"], list):
                for item in data["gear_list"]:
                    uuid = item.get("uuid") or item.get("gearUUID")
                    if not uuid:
                        continue
                    g = {"_uuid": uuid}
                    g["stats"], stats_succeeded = _safe_call_with_status(
                        self.api.get_gear_stats,
                        uuid,
                        label=f"gear_stats_{uuid}",
                    )
                    refresh_complete = refresh_complete and stats_succeeded
                    gear_details.append(g)

            data["gear_details"] = gear_details
            if refresh_complete:
                self.cache.put_section("gear", data, complete=True)
            elif cached_has_result:
                # Una incidencia temporal nunca debe destruir ni ocultar la
                # última fotografía completa de equipamiento.
                data = cached
            else:
                self.cache.put_section("gear", data, complete=False)

        if _compact_mode:
            self._remember_sensitive_payload(data.get("gear_list"))
            for gear_item in _as_list(data.get("gear_list")):
                if not isinstance(gear_item, dict):
                    continue
                gear_id = _pick(
                    gear_item,
                    "uuid",
                    "gearUUID",
                    "gearId",
                    "gearPk",
                    "id",
                )
                if gear_id is not None:
                    self.sensitive_identifiers.add(str(gear_id))
            gear = _compact_gear_section(
                data,
                reference_secret=self.reference_secret,
                include_free_text=self.include_free_text,
            )
            if not gear:
                self._quality_add(
                    "warnings",
                    "Garmin no devolvió una lista global de equipamiento.",
                )
                return
            _enrich_activity_gear_from_catalog(
                self.compact_activities,
                gear,
            )
            if self.compact_activities:
                self._remember_compact(
                    "activities",
                    self.compact_activities,
                )
            self._remember_compact("gear", gear)
            self.md.append("\nGear\n")
            self.md.append(
                'Schema: "Lista privada de equipamiento con nombre, fabricante y '
                'modelo para interpretar las actividades. El texto personalizado se '
                'marca como user_provided. Se excluyen IDs reales, valores '
                'predeterminados, datos de dispositivos y campos técnicos."\n'
            )
            self.md.append(f"{_json({'gear': gear})}\n")
        else:
            self.md.append("\nGear\n")
            if not data.get("gear_list"):
                _section_nodata(self.md, "Gear")
            else:
                _section(self.md, "Lista de equipamiento", data.get("gear_list"))
                _section(self.md, "Equipamiento predeterminado", data.get("gear_defaults"))
                for g in data.get("gear_details", []):
                    _section(self.md, f"Estadísticas del equipamiento: {g.get('_uuid', '?')}", g.get("stats"))

    # ===================================================================
    # Planes de entrenamiento
    # ===================================================================
    def export_training_plans(self):
        self.md.append("\nTraining Plans\n")

        cached = self.cache.get_section("training_plans")
        if cached is not None and not self.update_mode:
            data = cached
        else:
            failures_before_cache = self._endpoint_failure_count()
            data = {}
            plans = safe_call(self.api.get_training_plans, label="training_plans")
            data["plans"] = plans

            plan_details = []
            if plans and isinstance(plans, list):
                for item in plans:
                    pid = item.get("trainingPlanId") or item.get("id")
                    if not pid:
                        continue
                    p = {"_id": pid}
                    # Probar primero el plan estándar y después el adaptativo.
                    detail = safe_call(self.api.get_training_plan_by_id, pid, label=f"plan_{pid}")
                    if detail is None:
                        detail = safe_call(self.api.get_adaptive_training_plan_by_id, pid,
                                           label=f"adaptive_plan_{pid}")
                    p["detail"] = detail
                    plan_details.append(p)

            data["plan_details"] = plan_details
            self.cache.put_section(
                "training_plans",
                data,
                complete=(
                    self._endpoint_failure_count()
                    == failures_before_cache
                ),
            )

        if not data.get("plans"):
            _section_nodata(self.md, "Training Plans")
        elif _compact_mode:
            self.md.append('Schema: "plans: lista de planes de entrenamiento. plan_details: matriz {_id, detail} con el detalle de cada plan."\n')
            self.md.append(f"{_json(data)}\n")
        else:
            _section(self.md, "Training Plans", data.get("plans"))
            for p in data.get("plan_details", []):
                _section(self.md, f"Plan: {p.get('_id', '?')}", p.get("detail"))

    # ===================================================================
    # Entrenamientos
    # ===================================================================
    def export_workouts(self):
        self.md.append("\nWorkouts\n")

        cached = self.cache.get_section("workouts")
        if cached is not None and not self.update_mode:
            data = cached
        else:
            failures_before_cache = self._endpoint_failure_count()
            data = {}
            workout_list = safe_call(self.api.get_workouts, 0, 1000, label="workouts")
            data["workout_list"] = workout_list

            workout_details = []
            if workout_list and isinstance(workout_list, list):
                for item in workout_list:
                    wid = item.get("workoutId") or item.get("id")
                    if not wid:
                        continue
                    w = {"_id": wid}
                    w["detail"] = safe_call(self.api.get_workout_by_id, wid, label=f"workout_{wid}")
                    workout_details.append(w)

            data["workout_details"] = workout_details
            self.cache.put_section(
                "workouts",
                data,
                complete=(
                    self._endpoint_failure_count()
                    == failures_before_cache
                ),
            )

        if not data.get("workout_list"):
            _section_nodata(self.md, "Workouts")
        elif _compact_mode:
            self.md.append('Schema: "workout_list: definiciones de entrenamientos guardados. workout_details: matriz {_id, detail} con el detalle de cada entrenamiento."\n')
            self.md.append(f"{_json(data)}\n")
        else:
            _section(self.md, "Lista de entrenamientos", data.get("workout_list"))
            for w in data.get("workout_details", []):
                _section(self.md, f"Entrenamiento: {w.get('_id', '?')}", w.get("detail"))

    # ===================================================================
    # Hidratación: consulta diaria concurrente con caché.
    # ===================================================================
    def export_hydration(self):
        if not _compact_mode:
            self.md.append("\nHydration\n")
        log.info(f"  {self.days} días para comprobar")

        # Separar los días ya almacenados de los pendientes.
        days_list = []
        cached_results = {}
        cached_complete_keys = {}
        uncached_dates = []
        recent_cutoff = self.today - timedelta(days=2)
        for i in range(self.days):
            d = self.today - timedelta(days=i)
            ds = d.isoformat()
            days_list.append(ds)
            cached, complete_keys = self.cache.get_day_entry(
                f"hydration_{ds}"
            )
            if cached is not None:
                cached_results[ds] = cached
            cached_complete_keys[ds] = complete_keys
            if (
                cached is None
                or self.update_mode
                or d >= recent_cutoff
                or "hydration" not in complete_keys
            ):
                uncached_dates.append(ds)

        log.info(f"  {len(cached_results)} en caché, {len(uncached_dates)} por descargar")

        # Descargar en paralelo los días no almacenados.
        fetched_results = {}
        if uncached_dates:
            api = self.api
            def _fetch_hydration(ds):
                result, succeeded = _safe_call_with_status(
                    api.get_hydration_data,
                    ds,
                    label=f"hydration_{ds}",
                )
                return ds, result, succeeded

            t_start = time.time()
            done = 0
            with ThreadPoolExecutor(max_workers=4) as pool:
                for ds, result, succeeded in pool.map(
                    lambda ds: _fetch_hydration(ds),
                    uncached_dates,
                ):
                    data = dict(cached_results.get(ds) or {})
                    complete_keys = set(cached_complete_keys.get(ds) or set())
                    if succeeded:
                        data["hydration"] = result
                        complete_keys.add("hydration")
                    self.cache.put_day(
                        f"hydration_{ds}",
                        data,
                        complete_keys=complete_keys,
                    )
                    fetched_results[ds] = data
                    done += 1
                    if done % 50 == 0 or done == len(uncached_dates):
                        elapsed = time.time() - t_start
                        remaining = len(uncached_dates) - done
                        eta = (elapsed / done * remaining / 60) if done else 0
                        log.info(f"    {done}/{len(uncached_dates)} descargados | quedan aproximadamente {eta:.0f} min")

        # Escribir los resultados en orden cronológico.
        if _compact_mode:
            all_days = {}
            for ds in days_list:
                day_data = fetched_results.get(ds) or cached_results.get(ds)
                compact = (
                    _compact_hydration(day_data.get("hydration"))
                    if day_data
                    else None
                )
                if compact:
                    all_days[ds] = compact
            if all_days:
                self._remember_compact("hydration", dict(sorted(all_days.items())))
                self.md.append("\nHydration\n")
                self.md.append(
                    'Schema: "Solo fechas con ingesta real, ingesta durante actividad '
                    'o pérdida de sudor. El objetivo diario no cuenta como consumo."\n'
                )
                self.md.append(f"{_json(all_days)}\n")
        else:
            has_data = False
            for ds in days_list:
                day_data = fetched_results.get(ds) or cached_results.get(ds)
                if day_data and day_data.get("hydration"):
                    has_data = True
                    self.md.append(f"{ds}\n")
                    _section(self.md, "Hydration", day_data["hydration"], 4)
            if not has_data:
                _section_nodata(self.md, "Hydration")

    # ===================================================================
    # Nutrición: consulta diaria concurrente con caché.
    # ===================================================================
    def export_nutrition(self):
        if not _compact_mode:
            self.md.append("\nNutrition\n")
        log.info(f"  {self.days} días para comprobar")

        # Separar los días ya almacenados de los pendientes.
        days_list = []
        cached_results = {}
        cached_complete_keys = {}
        uncached_dates = []
        recent_cutoff = self.today - timedelta(days=2)
        for i in range(self.days):
            d = self.today - timedelta(days=i)
            ds = d.isoformat()
            days_list.append(ds)
            cached, complete_keys = self.cache.get_day_entry(
                f"nutrition_{ds}"
            )
            if cached is not None:
                cached_results[ds] = cached
            cached_complete_keys[ds] = complete_keys
            required_keys = (
                {"food_log", "meals"}
                if _compact_mode
                else {"food_log", "meals", "settings"}
            )
            if (
                cached is None
                or self.update_mode
                or d >= recent_cutoff
                or not required_keys.issubset(complete_keys)
            ):
                uncached_dates.append(ds)

        log.info(f"  {len(cached_results)} en caché, {len(uncached_dates)} por descargar")

        # Descargar en paralelo: tres llamadas por día.
        fetched_results = {}
        if uncached_dates:
            api = self.api
            def _fetch_nutrition(ds):
                existing = dict(cached_results.get(ds) or {})
                complete_keys = set(cached_complete_keys.get(ds) or set())
                values = {}
                calls = {
                    "food_log": (
                        api.get_nutrition_daily_food_log,
                        f"food_{ds}",
                    ),
                    "meals": (
                        api.get_nutrition_daily_meals,
                        f"meals_{ds}",
                    ),
                }
                if not _compact_mode:
                    calls["settings"] = (
                        api.get_nutrition_daily_settings,
                        f"nutr_set_{ds}",
                    )
                for key, (function, label) in calls.items():
                    value, succeeded = _safe_call_with_status(
                        function,
                        ds,
                        label=label,
                    )
                    if succeeded:
                        values[key] = value
                        complete_keys.add(key)
                existing.update(values)
                if _compact_mode:
                    existing.setdefault("settings", None)
                return ds, existing, complete_keys

            t_start = time.time()
            done = 0
            with ThreadPoolExecutor(max_workers=4) as pool:
                for ds, data, complete_keys in pool.map(
                    lambda ds: _fetch_nutrition(ds),
                    uncached_dates,
                ):
                    self.cache.put_day(
                        f"nutrition_{ds}",
                        data,
                        complete_keys=complete_keys,
                    )
                    fetched_results[ds] = data
                    done += 1
                    if done % 50 == 0 or done == len(uncached_dates):
                        elapsed = time.time() - t_start
                        remaining = len(uncached_dates) - done
                        eta = (elapsed / done * remaining / 60) if done else 0
                        log.info(f"    {done}/{len(uncached_dates)} descargados | quedan aproximadamente {eta:.0f} min")

        # Escribir los resultados en orden cronológico.
        if _compact_mode:
            all_days = {}
            for ds in days_list:
                day_data = fetched_results.get(ds) or cached_results.get(ds)
                compact = _compact_nutrition(day_data) if day_data else None
                if compact:
                    all_days[ds] = compact
            if all_days:
                self._remember_compact("nutrition", dict(sorted(all_days.items())))
                self.md.append("\nNutrition\n")
                self.md.append(
                    'Schema: "Solo fechas con alimentos registrados. Se omiten '
                    'contenedores de comidas, objetivos y ajustes diarios repetidos."\n'
                )
                self.md.append(f"{_json(all_days)}\n")
        else:
            has_data = False
            for ds in days_list:
                day_data = fetched_results.get(ds) or cached_results.get(ds)
                if day_data and any(day_data.get(k) for k in ("food_log", "meals", "settings")):
                    has_data = True
                    self.md.append(f"{ds}\n")
                    _section(self.md, "Registro de alimentos", day_data.get("food_log"), 4)
                    _section(self.md, "Comidas", day_data.get("meals"), 4)
                    _section(self.md, "Ajustes de nutrición", day_data.get("settings"), 4)
            if not has_data:
                _section_nodata(self.md, "Nutrition")

    # ===================================================================
    # Secciones calculadas del compacto semántico
    # ===================================================================
    def export_weekly_summary(self):
        if not _compact_mode:
            return
        self.md.append("\nWeekly Summary\n")
        self.md.append(
            'Schema: "Calculado únicamente con Activities y Daily Health de esta '
            'exportación. Las actividades nunca se cuentan desde el bienestar diario."\n'
        )
        self.md.append(
            f"{_json(_weekly_summary(self.compact_activities, self.compact_daily_records))}\n"
        )

    def export_data_quality(self):
        if not _compact_mode:
            return
        if self.errors:
            for error in self.errors:
                section = error.split(":", 1)[0]
                self._quality_add(
                    "endpoint_errors",
                    f"La sección {section} no terminó correctamente.",
                )
        self.md.append("\nData Quality\n")
        self.md.append(
            'Schema: "Limitaciones, datos críticos ausentes, clasificación temporal, '
            'unidades, filtros de privacidad y eliminación de duplicados."\n'
        )
        self.md.append(f"{_json({'data_quality': self.data_quality})}\n")

    # ===================================================================
    # Salud femenina
    # ===================================================================
    def export_womens_health(self):
        self.md.append("\nWomen's Health\n")

        cache_key = self._range_cache_key("womens_health")
        cached = self.cache.get_section(cache_key)
        if cached is not None and not self.update_mode:
            data = cached
        else:
            failures_before_cache = self._endpoint_failure_count()
            data = {}
            data["pregnancy"] = safe_call(self.api.get_pregnancy_summary, label="pregnancy")

            # El calendario menstrual devuelve 400 si no está activado.
            # Probar un día y consultar el historial solo si responde.
            probe = safe_call(self.api.get_menstrual_calendar_data,
                              self.today.isoformat(),
                              self.today.isoformat(),
                              label="menstrual_probe")
            if probe is not None:
                data["menstrual_calendar"] = _chunked_date_call(
                    self.api.get_menstrual_calendar_data,
                    self.start_date, self.today, "menstrual_cal")
            else:
                data["menstrual_calendar"] = None
            self.cache.put_section(
                cache_key,
                data,
                complete=(
                    self._endpoint_failure_count()
                    == failures_before_cache
                ),
            )

        if not any(data.get(k) for k in ("pregnancy", "menstrual_calendar")):
            _section_nodata(self.md, "Women's Health")
        elif _compact_mode:
            self.md.append('Schema: "pregnancy: resumen del seguimiento del embarazo. menstrual_calendar: historial de ciclos. Estas funciones deben activarse en Garmin."\n')
            self.md.append(f"{_json(data)}\n")
        else:
            _section(self.md, "Resumen del embarazo", data.get("pregnancy"))
            _section(self.md, "Calendario menstrual", data.get("menstrual_calendar"))


# ---------------------------------------------------------------------------
# Interfaz de línea de comandos
# ---------------------------------------------------------------------------
class SpanishArgumentParser(argparse.ArgumentParser):
    """Presenta en español los rótulos generales que argparse deja en inglés."""

    def format_usage(self):
        return super().format_usage().replace("usage: ", "uso: ", 1)

    def format_help(self):
        return super().format_help().replace("usage: ", "uso: ", 1)


def main():
    parser = SpanishArgumentParser(
        description="Exporta los datos de salud y actividad de Garmin Connect a texto con JSON",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        add_help=False,
        epilog="""
Ejemplos:
  python garmin_export.py --login                   # Iniciar sesión y guardar tokens
  python garmin_export.py                           # Últimos 30 días y 100 actividades
  python garmin_export.py --all                     # Historial completo
  python garmin_export.py --all --compact           # Historial completo en un archivo más pequeño
  python garmin_export.py --all --split             # Dividir para cargarlo en NotebookLM
  python garmin_export.py --update                  # Exportar solo los datos nuevos
  python garmin_export.py --all --no-cache          # Descargar todo de nuevo sin usar caché
  python garmin_export.py --start-date 2025-01-01   # Exportar desde una fecha concreta
  python garmin_export.py --start-date 2025-01-01 --end-date 2025-03-31
  python garmin_export.py --start-date 2025-01-01 --end-date 2025-03-31 --activity-details
  python garmin_export.py --days 365                # Un año de datos diarios
  python garmin_export.py --days 90 --activities 500
  python garmin_export.py --delay 1.0               # Ritmo más lento

Inicio de sesión:
  Ejecuta --login y escribe las credenciales cuando se soliciten.
  Los tokens se guardan localmente durante aproximadamente un año.
""",
    )
    parser._optionals.title = "opciones"
    parser.add_argument(
        "-h",
        "--help",
        action="help",
        help="Mostrar esta ayuda y salir",
    )
    parser.add_argument("--all", action="store_true", help="Exportar el historial completo")
    parser.add_argument("--days", type=int, default=30, help="Días de datos diarios (predeterminado: 30)")
    parser.add_argument("--activities", type=int, default=100, help="Número máximo de actividades (predeterminado: 100)")
    parser.add_argument("--start-date", type=str, default=None,
                        help="Exportar desde esta fecha (AAAA-MM-DD), incluidas todas las actividades del periodo")
    parser.add_argument("--end-date", type=str, default=None,
                        help="Final inclusivo del periodo (AAAA-MM-DD); requiere --start-date")
    parser.add_argument("--output", type=str, default="export", help="Carpeta de salida (predeterminada: export)")
    parser.add_argument("--filename", type=str, default=None,
                        help="Nombre base del archivo dentro de --output")
    parser.add_argument("--tokenstore", type=str, default=None, help="Ruta donde se guardan los tokens")
    parser.add_argument(
        "--force-login",
        action="store_true",
        help="Ignorar la sesión existente y solicitar de nuevo las credenciales",
    )
    parser.add_argument(
        "--ignore-credential-env",
        action="store_true",
        help=(
            "No leer credenciales de .env ni de variables de entorno; "
            "recomendado para perfiles separados"
        ),
    )
    parser.add_argument(
        "--non-interactive-auth",
        action="store_true",
        help=(
            "Utilizar solo los tokens guardados y fallar si la sesión necesita "
            "renovarse"
        ),
    )
    parser.add_argument(
        "--cache-dir",
        type=str,
        default=None,
        help="Carpeta privada de caché; permite aislarla por perfil",
    )
    parser.add_argument(
        "--report",
        choices=("preparation", "activity", "history"),
        default="history",
        help="Tipo de informe: preparación, una actividad o histórico",
    )
    parser.add_argument(
        "--race-review",
        action="store_true",
        help="Alias de --report preparation; crea una fotografía autosuficiente",
    )
    parser.add_argument(
        "--race-context",
        type=str,
        default=None,
        help="JSON local con carrera, objetivo y disponibilidad",
    )
    parser.add_argument(
        "--journal",
        type=str,
        default=None,
        help="JSON local con anotaciones opcionales",
    )
    parser.add_argument(
        "--review-weeks",
        type=int,
        default=None,
        help="Semanas de la revisión de preparación (4 a 52)",
    )
    parser.add_argument(
        "--format",
        choices=("txt", "xlsx", "both"),
        default="txt",
        help="Formato de salida: txt, xlsx o ambos",
    )
    parser.add_argument(
        "--xlsx",
        action="store_true",
        help="Añadir XLSX además del TXT (alias de --format both)",
    )
    parser.add_argument(
        "--include-free-text",
        action="store_true",
        help="Incluir descripciones, notas y otros textos libres mediante consentimiento explícito",
    )
    parser.add_argument(
        "--activity-id",
        type=str,
        default=None,
        help="Referencia privada de la única actividad que se analizará",
    )
    parser.add_argument(
        "--list-activities",
        type=str,
        default=None,
        metavar="RUTA_JSON",
        help="Guardar un catálogo privado de actividades y salir",
    )
    parser.add_argument(
        "--manifest",
        type=str,
        default=None,
        help="Guardar un manifiesto JSON del resultado para el lanzador",
    )
    parser.add_argument(
        "--run-id",
        type=str,
        default=None,
        help="Identificador local de ejecución que se copiará al manifiesto",
    )
    parser.add_argument(
        "--timezone",
        type=str,
        default=None,
        help="Zona horaria IANA (por ejemplo, Europe/Madrid)",
    )
    parser.add_argument("--delay", type=float, default=0.15, help="Espera base entre llamadas en segundos (predeterminada: 0.15)")
    parser.add_argument("--no-cache", action="store_true", help="No utilizar la caché y descargar todo de nuevo")
    parser.add_argument("--compact", action="store_true",
                        help="Crear un archivo más pequeño para herramientas de IA")
    parser.add_argument("--activity-details", action="store_true",
                        help="Con --compact, conservar el máximo detalle temporal de las actividades")
    parser.add_argument("--split", action="store_true",
                        help="Dividir la salida en archivos de menos de 500.000 palabras. Activa --compact")
    parser.add_argument("--update", action="store_true",
                        help="Exportar solo los datos nuevos desde la última exportación. Activa --compact")
    parser.add_argument("--login", action="store_true", help="Iniciar sesión, guardar tokens y salir")
    parser.add_argument(
        "--check-session",
        action="store_true",
        help="Comprobar los tokens guardados sin pedir credenciales y salir",
    )
    parser.add_argument("--verbose", action="store_true", help="Mostrar información técnica detallada")

    args = parser.parse_args()

    if args.days < 1:
        parser.error("--days debe ser al menos 1")
    if args.activities < 1:
        parser.error("--activities debe ser al menos 1")
    if args.review_weeks is not None and not 4 <= args.review_weeks <= 52:
        parser.error("--review-weeks debe estar entre 4 y 52")
    if args.login and args.check_session:
        parser.error("--login no se puede combinar con --check-session")
    if args.login and args.non_interactive_auth:
        parser.error("--login no se puede combinar con --non-interactive-auth")
    if args.force_login and not args.login:
        parser.error("--force-login requiere --login")
    if args.run_id and not re.fullmatch(r"[A-Za-z0-9_-]{8,80}", args.run_id):
        parser.error("--run-id contiene caracteres no permitidos")
    if args.activity_id and not re.fullmatch(
        r"activity_[0-9a-f]{12}",
        args.activity_id.strip(),
    ):
        parser.error(
            "--activity-id solo admite la referencia privada creada por el programa, "
            "con formato activity_ seguido de 12 caracteres hexadecimales"
        )
    if args.check_session:
        args.ignore_credential_env = True
        args.non_interactive_auth = True

    report_type = "preparation" if args.race_review else args.report
    if report_type in {"preparation", "activity"}:
        args.compact = True
    if report_type == "activity":
        if not args.activity_id:
            parser.error("--report activity requiere --activity-id")
        args.activity_details = True
    if report_type == "preparation" and (args.all or args.update):
        parser.error("la revisión de preparación no se combina con --all ni --update")

    output_format = "both" if args.xlsx else args.format
    if output_format in {"xlsx", "both"}:
        args.compact = True
    if args.split and output_format == "xlsx":
        parser.error("--split requiere que la salida incluya TXT")

    explicit_start_date = None
    explicit_end_date = None
    reference_today = date.today()
    if args.end_date:
        try:
            explicit_end_date = date.fromisoformat(args.end_date)
        except ValueError:
            parser.error("--end-date debe utilizar el formato AAAA-MM-DD")
        if explicit_end_date > reference_today:
            parser.error("--end-date no puede ser una fecha futura")
    if args.start_date:
        try:
            explicit_start_date = date.fromisoformat(args.start_date)
        except ValueError:
            parser.error("--start-date debe utilizar el formato AAAA-MM-DD")
        if explicit_start_date > reference_today:
            parser.error("--start-date no puede ser una fecha futura")
        if args.all or args.update:
            parser.error("--start-date no se puede combinar con --all ni --update")
    if explicit_end_date is not None:
        if (
            explicit_start_date is None
            and report_type != "preparation"
            and not args.list_activities
        ):
            parser.error("--end-date requiere --start-date")
        if (
            explicit_start_date is not None
            and explicit_end_date < explicit_start_date
        ):
            parser.error("--end-date no puede ser anterior a --start-date")

    context_reference_date = explicit_end_date or reference_today
    try:
        race_context_raw = load_local_json(
            Path(args.race_context) if args.race_context else None,
            "contexto de carrera",
        )
        race_context, default_review_weeks = normalise_race_context(
            race_context_raw,
            context_reference_date,
            include_free_text=args.include_free_text,
        )
        journal_raw = load_local_json(
            Path(args.journal) if args.journal else None,
            "diario",
        )
        journal = normalise_journal(
            journal_raw,
            include_free_text=args.include_free_text,
        )
    except ValueError as exc:
        parser.error(str(exc))

    review_weeks = args.review_weeks or default_review_weeks
    if report_type == "preparation" and explicit_start_date is None:
        explicit_end_date = explicit_end_date or reference_today
        explicit_start_date = _inclusive_review_start(
            explicit_end_date,
            review_weeks,
        )

    output_filename = None
    if args.filename:
        try:
            output_filename = _normalise_output_stem(args.filename)
        except ValueError as exc:
            parser.error(f"--filename: {exc}")

    try:
        timezone_name = _resolve_timezone(args.timezone)
    except ValueError as exc:
        parser.error(f"--timezone: {exc}")

    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)

    global _limiter, _compact_mode, _split_mode, _update_mode
    _limiter = RateLimiter(base_delay=args.delay)
    _update_mode = args.update
    _split_mode = args.split
    if _split_mode or _update_mode:
        args.compact = True  # --split y --update activan --compact.
    _compact_mode = args.compact

    print()
    print("  ExportaGarmin — Tus datos de Garmin, ordenados y preparados para la IA")
    print(f"  {'-' * 38}")
    print()

    if args.tokenstore:
        tokenstore = args.tokenstore
    elif args.ignore_credential_env:
        tokenstore = "~/.garminconnect"
    else:
        tokenstore = os.getenv("GARMINTOKENS", "~/.garminconnect")

    try:
        api = authenticate(
            tokenstore,
            use_credential_environment=not args.ignore_credential_env,
            interactive=not args.non_interactive_auth,
            force_login=args.force_login,
        )
    except GarminConnectTooManyRequestsError:
        log.error("Demasiadas solicitudes. Espera unos minutos y vuelve a intentarlo.")
        sys.exit(1)
    except SystemExit:
        raise
    except Exception as e:
        log.error(f"El inicio de sesión ha fallado: {_friendly_login_error(e)}")
        log.debug(f"Tipo de fallo de acceso: {_safe_exception_reason(e)}")
        sys.exit(1)

    if args.login:
        print()
        log.info("Inicio de sesión correcto: los tokens están guardados. Ya puedes realizar exportaciones.")
        sys.exit(0)
    if args.check_session:
        print()
        log.info("Sesión comprobada correctamente.")
        sys.exit(0)

    out = Path(args.output)
    out.mkdir(parents=True, exist_ok=True)

    use_cache = not getattr(args, 'no_cache', False)
    cache = ExportCache(
        out,
        enabled=use_cache,
        cache_dir=Path(args.cache_dir) if args.cache_dir else None,
    )
    if use_cache:
        log.info("Caché: activada (permite continuar exportaciones interrumpidas)")
    else:
        log.info("Caché: desactivada (--no-cache)")

    if args.list_activities:
        catalog_end = explicit_end_date or reference_today
        catalog_start = (
            explicit_start_date
            or catalog_end - timedelta(days=max(args.days - 1, 0))
        )
        activities = safe_call(
            api.get_activities_by_date,
            catalog_start.isoformat(),
            catalog_end.isoformat(),
            None,
            label="activity_catalog",
        ) or []
        secret = load_or_create_reference_secret(cache.cache_dir)
        catalog = [
            activity_catalog_entry(activity, secret)
            for activity in activities
            if isinstance(activity, dict)
        ]
        catalog.sort(key=lambda item: item.get("date", ""), reverse=True)
        atomic_write_json(
            Path(args.list_activities),
            {
                "schema_version": _COMPACT_SCHEMA_VERSION,
                "start_date": catalog_start.isoformat(),
                "end_date": catalog_end.isoformat(),
                "activities": catalog,
            },
        )
        log.info(f"Catálogo privado creado: {len(catalog)} actividades")
        sys.exit(0)

    exporter = GarminExporter(api, out, args.days, args.activities,
                              fetch_all=getattr(args, 'all', False),
                              cache=cache,
                              update_mode=_update_mode,
                              explicit_start_date=explicit_start_date,
                              explicit_end_date=explicit_end_date,
                              output_filename=output_filename,
                              include_activity_details=args.activity_details,
                              timezone_name=timezone_name,
                              output_format=output_format,
                              report_type=report_type,
                              race_context=race_context,
                              journal=journal,
                              review_weeks=review_weeks,
                              selected_activity_ref=args.activity_id,
                              include_free_text=args.include_free_text,
                              manifest_path=(
                                  Path(args.manifest)
                                  if args.manifest
                                  else None
                              ),
                              run_id=args.run_id)
    try:
        exporter.run()
    except KeyboardInterrupt:
        print()
        log.info("Proceso interrumpido: los datos de la caché están guardados. Ejecuta de nuevo para continuar.")
        sys.exit(130)


if __name__ == "__main__":
    main()

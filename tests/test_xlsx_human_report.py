import copy
import json
import tempfile
import unittest
import zipfile
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook

from training_analysis import (
    SCHEMA_VERSION,
    XLSX_PRESENTATION_VERSION,
    build_quality_report,
    build_report_extensions,
    render_xlsx,
)


class HumanXlsxReportTests(unittest.TestCase):
    """Contrato del Excel humano con cantidades ficticias equivalentes."""

    @classmethod
    def setUpClass(cls):
        from tests.test_compact_reduction import CompactReductionRegressionTests

        source_case = CompactReductionRegressionTests(methodName="runTest")
        _, activities = source_case._dataset()
        cls.start = source_case.START
        cls.end = source_case.END

        target_running_distance = 327_579.6
        running = [item for item in activities if item.get("sport") == "running"]
        cycling = [item for item in activities if item.get("sport") == "cycling"]
        # Dos registros accidentales de pocos segundos en un mismo día: uno de
        # carrera y otro de ciclismo. Se conservan, pero no alteran indicadores.
        micro_date = running[0]["date"]
        for index, activity in enumerate((running[0], cycling[0])):
            activity["date"] = micro_date
            activity["distance_m"] = 0.0
            activity["duration_s"] = 5.0 - index
            activity["moving_duration_s"] = 5.0 - index
        running[1]["sport"] = "trail_running"
        valid_running = running[1:]
        per_run = round(target_running_distance / len(valid_running), 6)
        assigned = 0.0
        for index, activity in enumerate(valid_running):
            distance = (
                round(target_running_distance - assigned, 6)
                if index == len(valid_running) - 1
                else per_run
            )
            activity["distance_m"] = distance
            assigned += distance

        for activity_index, activity in enumerate(activities):
            activity["hr_zones"] = [
                {
                    "zone": zone,
                    "duration_s": 300 + zone,
                    "low_boundary_bpm": 90 + zone * 15,
                }
                for zone in range(1, 6)
            ]
            power_count = 5 if activity_index < 17 else 4
            activity["power_zones"] = [
                {
                    "zone": zone,
                    "duration_s": 180 + zone,
                    "low_boundary_w": 80 + zone * 40,
                }
                for zone in range(1, power_count + 1)
            ]
            activity.pop("self_evaluation", None)
            activity["training_load"] = 60.0 + activity_index
            activity["training_effect_label"] = "AEROBIC_BASE"
            activity["start_time_bucket"] = "morning"
        valid_activities = [
            activity for activity in activities
            if activity not in (running[0], cycling[0])
        ]
        for activity in valid_activities[:43]:
            activity["self_evaluation"] = {
                "perceived_exertion_1_10": 5.0,
                "feeling": "good",
            }
        # Los dos registros breves tienen autoevaluación a propósito: así se
        # demuestra que quedan fuera tanto del numerador como del denominador.
        for activity in (running[0], cycling[0]):
            activity["self_evaluation"] = {
                "perceived_exertion_1_10": 2.0,
                "feeling": "strong",
            }
        activities[1]["training_effect_label"] = "Other"

        gear = []
        for index in range(29):
            reference = f"gear_{index + 1:012x}"
            gear.append({
                "gear_ref": reference,
                "gear_name": (
                    "Mis Zapatillas Ñandú"
                    if index == 0
                    else f"Material ficticio {index + 1}"
                ),
                "manufacturer": "Marca ficticia",
                "model": f"Modelo {index + 1}",
                "type": "Shoes" if index < 20 else "Bike",
                "status": "retired" if index == 28 else "active",
                "total_distance_m": 100_000 + index * 1_000,
            })
        for index, activity in enumerate(activities):
            activity["gear_refs"] = [gear[index % len(gear)]["gear_ref"]]

        days = []
        for offset in range(112):
            day = cls.start + timedelta(days=offset)
            days.append({
                "date": day.isoformat(),
                "steps": 0 if offset == 0 else 8_000 + offset,
                "distance_m": 6_000 + offset,
                "active_calories_kcal": 500 + offset,
                "total_calories_kcal": 2_200 + offset,
                "moderate_intensity_minutes": 20,
                "vigorous_intensity_minutes": 10,
                "resting_heart_rate_bpm": 48 + offset % 5,
                "average_stress": 30 + offset % 8,
                "maximum_stress": 70,
                "body_battery_high": 80,
                "body_battery_low": 20,
                "body_battery_charged": 55,
                "body_battery_drained": 45,
                "average_spo2_pct": 96.0,
                "lowest_spo2_pct": 92.0,
                "average_waking_respiration_brpm": 14.2,
                "average_sleep_respiration_brpm": 12.8,
                "sleep": {
                    "valid_sleep": True,
                    "sleep_start_local": f"{day.isoformat()}T23:00:00+02:00",
                    "sleep_end_local": f"{(day + timedelta(days=1)).isoformat()}T07:00:00+02:00",
                    "total_sleep_s": 28_000 + offset,
                    "awake_s": 800,
                    "light_sleep_s": 15_000,
                    "deep_sleep_s": 6_000,
                    "rem_sleep_s": 7_000,
                    "nap_time_s": 0,
                    "sleep_score": None if offset == 0 else 82,
                    "sleep_score_qualifier": "GOOD",
                    "average_sleep_heart_rate_bpm": 46,
                    "average_sleep_stress": 18,
                    "average_sleep_spo2_pct": 95.0,
                },
                "hrv": {
                    "overnight_average_ms": 52,
                    "highest_five_min_average_ms": 70,
                    "weekly_average_ms": 50,
                    "baseline_balanced_low_ms": 42,
                    "baseline_balanced_high_ms": 62,
                    "status": "BALANCED",
                },
                "lifestyle_logs": (
                    [{
                        "date": day.isoformat(),
                        "behaviour": "travel",
                        "status": True,
                    }]
                    if offset == 10 else []
                ),
            })

        race_context = {
            "event": {
                "label": "Maratón ficticio de Valencia",
                "distance_type": "marathon",
                "distance_m": 42_195.0,
                "race_date": "2026-12-06",
                "days_remaining": 129,
                "weeks_remaining": 18.4,
            },
            "goal": {
                "target_time_s": 16_200,
                "target_pace_s_per_km": 383.9,
            },
            "availability": {
                "available_days_per_week": 5,
                "long_run_day": "Domingo",
                "restrictions": "Sin restricciones ficticias.",
            },
        }
        journal = [
            {
                "date": activities[index]["date"],
                "activity_ref": activities[index]["activity_ref"],
                "note": "Nota ficticia = no es fórmula" if index == 0 else f"Nota ficticia {index + 1}",
                "intended_session_type": "easy" if index % 2 == 0 else "interval",
                "source": "user_provided",
            }
            for index in range(6)
        ]
        journal[5]["date"] = (cls.start + timedelta(days=90)).isoformat()
        journal[0]["intended_session_type"] = None
        journal.append({
            "date": journal[0]["date"],
            "activity_ref": journal[0]["activity_ref"],
            "note": None,
            "intended_session_type": "Prueba",
            "source": "user_provided",
        })
        journal.append({
            "date": journal[0]["date"],
            "activity_ref": journal[0]["activity_ref"],
            "note": None,
            "intended_session_type": None,
            "source": "user_provided",
        })
        extensions = build_report_extensions(
            activities,
            days,
            cls.start,
            cls.end,
            race_context,
            journal,
        )
        classified = extensions["activities"]

        def set_classification(index, session_type, name, source="user_provided"):
            activity = classified[index]
            activity["name"] = name
            activity["garmin_event_type"] = (
                "race" if session_type == "race" else "uncategorized"
            )
            activity["classification"] = {
                "type": session_type,
                "source": source,
                "confidence": 1.0 if source == "user_provided" else 0.0,
                "evidence": (
                    ["manual_intended_session_type"]
                    if source == "user_provided" else []
                ),
            }

        # Distribución exacta del caso ficticio: 44 sin clasificar, 4 de
        # intervalos, 3 tempo, 6 fáciles, 3 largas, 1 competición,
        # 29 cruzadas y 1 fuerza.
        set_classification(0, "unknown", "Registro breve de carrera", "insufficient_evidence")
        for index in range(1, 5):
            set_classification(index, "interval", f"S{index} Intervalos - Repeticiones")
        for index in range(5, 8):
            set_classification(index, "tempo", f"S{index - 4} Tempo")
        for index in range(8, 14):
            set_classification(index, "easy", f"Carrera fácil {index - 7}")
        for index in range(14, 17):
            set_classification(index, "long_run", f"Tirada larga {index - 13}")
        set_classification(17, "race", "Rock and roll 1/2 maratón 2026")
        set_classification(18, "unknown", "Entrenamiento en cinta", "insufficient_evidence")
        set_classification(19, "unknown", "Base", "insufficient_evidence")
        for index in range(20, 60):
            set_classification(index, "unknown", f"Carrera continua {index}", "insufficient_evidence")
        set_classification(60, "unknown", "Registro breve de ciclismo", "insufficient_evidence")
        for index in range(61, 76):
            set_classification(index, "cross_training", f"Ciclismo {index - 60}")
        set_classification(76, "strength", "Fuerza")
        for index in range(77, 91):
            set_classification(index, "cross_training", f"Caminar {index - 76}")

        # Las 52 carreras no estructuradas contienen resúmenes agregados de
        # Garmin. No deben aparecer en INTERVALOS.
        for index in [0, *range(8, 60)]:
            activity = classified[index]
            activity["interval_summaries"] = [
                {
                    "interval_type": "INTERVAL_ACTIVE",
                    "interval_count": 1,
                    "distance_m": activity.get("distance_m"),
                    "duration_s": activity.get("duration_s"),
                    "moving_duration_s": activity.get("moving_duration_s"),
                },
                {
                    "interval_type": "RWD_RUN",
                    "interval_count": 1,
                    "distance_m": activity.get("distance_m"),
                    "duration_s": activity.get("duration_s"),
                    "moving_duration_s": activity.get("moving_duration_s"),
                },
            ]

        # Siete sesiones realmente estructuradas. Los totales duplicados y la
        # pausa automática se incluyen en la entrada para comprobar el filtro.
        for index in range(1, 8):
            activity = classified[index]
            activity["interval_summaries"] = [
                {
                    "interval_type": "INTERVAL_WARMUP",
                    "interval_count": 1,
                    "distance_m": 1_500.0,
                    "duration_s": 600.0,
                    "moving_duration_s": 590.0,
                },
                {
                    "interval_type": "INTERVAL_ACTIVE",
                    "interval_count": 6,
                    "distance_m": 50.0 if index == 1 else 4_800.0,
                    "duration_s": 1_440.0,
                    "moving_duration_s": 1.0 if index == 1 else 1_420.0,
                    "average_pace_s_per_km": 17_023.0 if index == 1 else 300.0,
                    "moving_pace_s_per_km": 46_663.0 if index == 1 else 296.0,
                    "best_pace_s_per_km": 9_999.0 if index == 1 else 280.0,
                    "grade_adjusted_pace_s_per_km": 8_888.0 if index == 1 else 298.0,
                },
                {
                    "interval_type": "INTERVAL_ACTIVE",
                    "interval_count": 6,
                    "distance_m": 50.1 if index == 1 else 4_803.0,
                    "duration_s": 1_441.0,
                    "moving_duration_s": 1.0 if index == 1 else 1_421.0,
                },
                {
                    "interval_type": "INTERVAL_RECOVERY",
                    "interval_count": 5,
                    "distance_m": 800.0,
                    "duration_s": 600.0,
                    "moving_duration_s": 580.0,
                },
                {
                    "interval_type": "INTERVAL_COOLDOWN",
                    "interval_count": 1,
                    "distance_m": 1_000.0,
                    "duration_s": 480.0,
                    "moving_duration_s": 470.0,
                },
                {
                    "interval_type": "RWD_RUN",
                    "interval_count": 1,
                    "distance_m": activity.get("distance_m"),
                    "duration_s": activity.get("duration_s"),
                },
                {
                    "interval_type": "RUN",
                    "interval_count": 1,
                    "distance_m": (activity.get("distance_m") or 0) + 30.0,
                    "duration_s": (activity.get("duration_s") or 0) + 20.0,
                },
                {
                    "interval_type": "RWD_STAND",
                    "interval_count": 1,
                    "distance_m": 0.0,
                    "duration_s": 5.0,
                },
            ]

        short_laps = 0
        for activity in extensions["activities"]:
            for lap in activity.get("laps", []) or []:
                if short_laps >= 69:
                    break
                lap.update({
                    "distance_m": 50.0,
                    "duration_s": 30.0,
                    "moving_duration_s": 30.0,
                    "average_pace_s_per_km": 600.0,
                    "best_pace_s_per_km": 500.0,
                    "partial_lap": True,
                })
                short_laps += 1
            if short_laps >= 69:
                break
        if short_laps != 69:
            raise AssertionError("No se prepararon las 69 vueltas cortas ficticias")

        # Seis centinelas negativos de Garmin deben contar como ausencias solo
        # en la vista humana; el modelo de datos original permanece intacto.
        for offset in (2, 20, 40, 60, 80, 100):
            days[offset]["average_stress"] = -1
        quality = build_quality_report(
            days,
            extensions["activities"],
            cls.start,
            cls.end,
            legacy_quality={
                "unit_conversions": [
                    "raw sleep epochs convertidos a fecha local.",
                    "lactate-threshold speed convertida a m/s.",
                    "totalAscent y totalDescent interpretados en metros.",
                ],
                "duplicate_sources_removed": [
                    "splits.vueltas, typed_splits.splits y summary/detail unificados.",
                    "split_summaries duplicados eliminados.",
                ],
            },
        )
        quality["warnings"].append(
            "endpoint de sueño no disponible entre 2026-05-02/2026-05-04."
        )
        quality["transformations"].append(
            "summaryDTO, lapDTOs y speed_raw se conservaron en la trazabilidad técnica."
        )
        composition = [
            {
                "date": (cls.start + timedelta(days=index * 6)).isoformat(),
                "weight_kg": 72.0 - index * 0.05,
                "bmi": 22.3,
                "body_fat_pct": 14.2,
                "body_water_pct": 58.1,
                "muscle_mass_kg": 57.0,
                "bone_mass_kg": 3.2,
            }
            for index in range(18)
        ]
        pressure = [
            {
                "timestamp": f"{(cls.start + timedelta(days=index * 15)).isoformat()}T08:30:00+02:00",
                "systolic_mmhg": 118 + index,
                "diastolic_mmhg": 72 + index,
                "pulse_bpm": 52 + index,
            }
            for index in range(7)
        ]
        cls.model = {
            "export_metadata": {
                "schema_version": SCHEMA_VERSION,
                "export_status": "completed",
                "exported_at": "2026-08-07T09:15:00+02:00",
            },
            "race_context": race_context,
            "profile": {"primary_watch": "Garmin ficticio 955"},
            "daily_health": days,
            "gear": gear,
            "journal": journal,
            "blood_pressure": pressure,
            "body_composition": composition,
            "data_quality": quality,
            **extensions,
        }
        cls.temp_dir = tempfile.TemporaryDirectory()
        cls.output = Path(cls.temp_dir.name) / "revision_ficticia_maraton.xlsx"
        render_xlsx(cls.model, cls.output)

    @classmethod
    def tearDownClass(cls):
        cls.temp_dir.cleanup()

    def open_book(self, **kwargs):
        return load_workbook(self.output, **kwargs)

    @staticmethod
    def table_values(sheet):
        table = next(iter(sheet.tables.values()))
        start, end = table.ref.split(":")
        from openpyxl.utils.cell import range_boundaries
        min_col, min_row, max_col, max_row = range_boundaries(table.ref)
        headers = [
            sheet.cell(min_row, column).value
            for column in range(min_col, max_col + 1)
        ]
        rows = []
        for row_index in range(min_row + 1, max_row + 1):
            rows.append({
                header: sheet.cell(row_index, column_index).value
                for header, column_index in zip(
                    headers, range(min_col, max_col + 1)
                )
            })
        return rows

    def test_exact_equivalent_counts_and_totals_are_preserved(self):
        workbook = self.open_book(data_only=False)
        try:
            self.assertEqual(91, len(self.table_values(workbook["ACTIVIDADES"])))
            intervals = self.table_values(workbook["INTERVALOS"])
            self.assertEqual(7, len({row["Actividad"] for row in intervals}))
            self.assertEqual(28, len(intervals))
            self.assertEqual(364, len(self.table_values(workbook["VUELTAS"])))
            self.assertEqual(836, len(self.table_values(workbook["ZONAS"])))
            self.assertEqual(17, len(self.table_values(workbook["SEMANAS"])))
            self.assertEqual(112, len(self.table_values(workbook["SALUD DIARIA"])))
            self.assertEqual(29, len(self.table_values(workbook["EQUIPAMIENTO"])))
            measures = workbook["MEDIDAS"]
            self.assertEqual(2, len(measures.tables))
            from openpyxl.utils.cell import range_boundaries
            composition_bounds = range_boundaries(
                measures.tables["TablaComposicion"].ref
            )
            pressure_bounds = range_boundaries(
                measures.tables["TablaPresionArterial"].ref
            )
            self.assertEqual(18, composition_bounds[3] - composition_bounds[1])
            self.assertEqual(7, pressure_bounds[3] - pressure_bounds[1])
            self.assertGreaterEqual(measures.column_dimensions["A"].width, 19)
            self.assertIsInstance(
                measures.cell(pressure_bounds[1] + 1, 1).value,
                datetime,
            )
            activity_rows = self.table_values(workbook["ACTIVIDADES"])
            sports = {}
            for row in activity_rows:
                sports[row["Deporte"]] = sports.get(row["Deporte"], 0) + 1
            self.assertEqual(59, sports["Carrera"])
            self.assertEqual(1, sports["Carrera por montaña"])
            self.assertEqual(16, sports["Ciclismo"])
            self.assertEqual(1, sports["Fuerza"])
            self.assertEqual(14, sports["Caminar"])
            running_km = sum(
                row["Distancia (km)"]
                for row in activity_rows
                if row["Deporte"] in {"Carrera", "Carrera por montaña"}
            )
            self.assertAlmostEqual(327.5796, running_km, places=4)
            self.assertEqual(6, len(self.table_values(workbook["DIARIO"])))
        finally:
            workbook.close()

    def test_excel_preserves_a_case_with_91_original_activities(self):
        from tests.test_compact_reduction import CompactReductionRegressionTests

        source_case = CompactReductionRegressionTests(methodName="runTest")
        _, activities = source_case._dataset()
        extensions = build_report_extensions(
            activities, [], source_case.START, source_case.END
        )
        model = {
            "export_metadata": {"schema_version": SCHEMA_VERSION},
            **extensions,
        }
        with tempfile.TemporaryDirectory() as directory:
            output = Path(directory) / "91_actividades_ficticias.xlsx"
            render_xlsx(model, output)
            workbook = load_workbook(output, data_only=False)
            try:
                self.assertEqual(
                    91, len(self.table_values(workbook["ACTIVIDADES"]))
                )
            finally:
                workbook.close()

    def test_rendering_excel_does_not_modify_the_txt_model(self):
        original = copy.deepcopy(self.model)
        with tempfile.TemporaryDirectory() as directory:
            render_xlsx(
                self.model,
                Path(directory) / "modelo_sin_cambios.xlsx",
            )
        self.assertEqual(original, self.model)
        self.assertEqual("3.3.1", SCHEMA_VERSION)

    def test_visible_structure_is_human_and_technical_sheets_are_hidden(self):
        workbook = self.open_book(data_only=False)
        try:
            visible = [
                sheet.title for sheet in workbook.worksheets
                if sheet.sheet_state == "visible"
            ]
            self.assertEqual([
                "INICIO", "RESUMEN", "SEMANAS", "ACTIVIDADES",
                "INTERVALOS", "VUELTAS", "ZONAS", "SALUD DIARIA",
                "HÁBITOS", "MEDIDAS", "EQUIPAMIENTO", "DIARIO",
                "CALIDAD DATOS", "AYUDA",
            ], visible)
            hidden = [
                sheet.title for sheet in workbook.worksheets
                if sheet.sheet_state == "hidden"
            ]
            self.assertTrue(hidden)
            self.assertTrue(all(
                name.startswith("TÉCNICO - ") or name == "DATOS POR SEGUNDO"
                for name in hidden
            ))
            self.assertIn("TÉCNICO - ACTIVIDADES", hidden)
            self.assertIn("TÉCNICO - MAPEO", hidden)
            for sheet in workbook.worksheets:
                if sheet.sheet_state != "visible":
                    continue
                self.assertTrue(any(
                    cell.value is not None
                    for row in sheet.iter_rows()
                    for cell in row
                ), sheet.title)
                for table in sheet.tables.values():
                    min_col, min_row, max_col, _ = __import__(
                        "openpyxl.utils.cell", fromlist=["range_boundaries"]
                    ).range_boundaries(table.ref)
                    for column in range(min_col, max_col + 1):
                        header = str(sheet.cell(min_row, column).value or "")
                        self.assertNotIn("_", header)
                        self.assertNotIn(".", header)
            self.assertLessEqual(workbook["ACTIVIDADES"].max_column, 30)
            self.assertLessEqual(workbook["SEMANAS"].max_column, 35)
        finally:
            workbook.close()

    def test_dates_durations_paces_percentages_and_missing_values_are_real(self):
        workbook = self.open_book(data_only=False)
        try:
            activities = workbook["ACTIVIDADES"]
            headers = {cell.value: cell.column for cell in activities[1]}
            self.assertIsInstance(
                activities.cell(2, headers["Fecha"]).value,
                (datetime,),
            )
            self.assertNotIsInstance(
                activities.cell(2, headers["Duración"]).value,
                str,
            )
            self.assertEqual(
                "[m]:ss",
                activities.cell(2, headers["Ritmo medio (min/km)"]).number_format,
            )
            self.assertIsNone(
                activities.cell(2, headers["Potencia normalizada (W)"]).value
            )

            health = workbook["SALUD DIARIA"]
            health_headers = {cell.value: cell.column for cell in health[1]}
            self.assertEqual(0, health.cell(2, health_headers["Pasos"]).value)
            self.assertIsNone(
                health.cell(2, health_headers["Puntuación del sueño"]).value
            )
            self.assertIsInstance(
                health.cell(2, health_headers["Inicio del sueño"]).value,
                datetime,
            )

            zones = workbook["ZONAS"]
            zone_headers = {cell.value: cell.column for cell in zones[1]}
            percentage = zones.cell(2, zone_headers["Porcentaje"])
            self.assertGreaterEqual(percentage.value, 0)
            self.assertLessEqual(percentage.value, 1)
            self.assertEqual("0.0%", percentage.number_format)

            quality = workbook["CALIDAD DATOS"]
            quality_rows = self.table_values(quality)
            sleep_quality = next(
                row for row in quality_rows
                if row["Métrica"] == "Duración del sueño"
            )
            self.assertIsInstance(sleep_quality["Media"], timedelta)
            metric_column = {
                cell.value: cell.column for cell in quality[1]
            }
            self.assertEqual(
                "[h]:mm:ss",
                quality.cell(2, metric_column["Media"]).number_format,
            )
        finally:
            workbook.close()

    def test_no_visible_json_formula_or_formula_error(self):
        workbook = self.open_book(data_only=False)
        try:
            formula_errors = {"#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A"}
            for sheet in workbook.worksheets:
                if sheet.sheet_state != "visible":
                    continue
                for row in sheet.iter_rows():
                    for cell in row:
                        self.assertNotEqual("f", cell.data_type, f"{sheet.title}!{cell.coordinate}")
                        self.assertNotIn(cell.value, formula_errors)
                        if not isinstance(cell.value, str):
                            continue
                        stripped = cell.value.strip()
                        if stripped.startswith(("{", "[")):
                            try:
                                decoded = json.loads(stripped)
                            except json.JSONDecodeError:
                                continue
                            self.assertNotIsInstance(decoded, (dict, list))
        finally:
            workbook.close()

    def test_translations_session_types_and_microactivities_are_coherent(self):
        workbook = self.open_book(data_only=False)
        try:
            forbidden = {
                "strong", "trail_running", "trail running", "interval",
                "other", "none", "uncategorized",
            }
            for sheet in workbook.worksheets:
                if sheet.sheet_state != "visible":
                    continue
                for row in sheet.iter_rows():
                    for cell in row:
                        if isinstance(cell.value, str):
                            self.assertNotIn(
                                cell.value.strip().casefold(),
                                forbidden,
                                f"{sheet.title}!{cell.coordinate}",
                            )

            activities = self.table_values(workbook["ACTIVIDADES"])
            brief = [
                row for row in activities
                if row["Estado del registro"] == "Registro muy breve"
            ]
            self.assertEqual(2, len(brief))
            valid = [
                row for row in activities
                if row["Estado del registro"] == "Válido para análisis"
            ]
            self.assertEqual(89, len(valid))
            references = [row["ID de actividad"] for row in activities]
            self.assertEqual(91, len(set(references)))
            uncategorized_name = self.model["activities"][18]["name"]
            uncategorized = next(
                row for row in activities if row["Actividad"] == uncategorized_name
            )
            self.assertEqual("Sin clasificar", uncategorized["Tipo de sesión"])

            summary = workbook["RESUMEN"]
            indicators = {}
            for row in range(1, summary.max_row + 1):
                indicators[summary.cell(row, 1).value] = summary.cell(row, 2).value
                indicators[summary.cell(row, 4).value] = summary.cell(row, 5).value
            self.assertEqual(59, indicators["Sesiones de carrera"])
            self.assertEqual(
                self.model["period_summary"]["training"]["days_with_any_training"] - 1,
                indicators["Días entrenados"],
            )
            expected_load = sum(
                activity.get("training_load") or 0
                for activity in self.model["activities"]
                if not (
                    activity.get("duration_s") < 60
                    and activity.get("distance_m") < 100
                    and any(
                        token in str(activity.get("sport") or "").casefold()
                        for token in ("run", "cycl", "bike")
                    )
                )
            )
            self.assertEqual(expected_load, indicators["Carga Garmin total"])
            weekly = self.table_values(workbook["SEMANAS"])
            self.assertEqual(89, sum(row["Sesiones totales"] for row in weekly))

            session_counts = {}
            for activity in activities:
                session_type = activity["Tipo de sesión"]
                session_counts[session_type] = session_counts.get(session_type, 0) + 1
            self.assertEqual({
                "Sin clasificar": 44,
                "Intervalos": 4,
                "Tempo": 3,
                "Rodaje fácil": 6,
                "Tirada larga": 3,
                "Competición": 1,
                "Entrenamiento cruzado": 29,
                "Fuerza": 1,
            }, session_counts)

            visible_text = "\n".join(
                str(cell.value)
                for sheet in workbook.worksheets
                if sheet.sheet_state == "visible"
                for row in sheet.iter_rows()
                for cell in row if cell.value is not None
            )
            quality_text = "\n".join(
                str(cell.value)
                for row in workbook["CALIDAD DATOS"].iter_rows()
                for cell in row if cell.value is not None
            )
            self.assertIn("2 registros muy breves", quality_text)
            self.assertIn("no cuentan en sesiones", quality_text)
        finally:
            workbook.close()

    def test_intervals_contains_only_the_seven_structured_activities(self):
        workbook = self.open_book(data_only=False)
        try:
            intervals = self.table_values(workbook["INTERVALOS"])
            activities = {row["Actividad"] for row in intervals}
            self.assertEqual({
                "S1 Intervalos - Repeticiones",
                "S2 Intervalos - Repeticiones",
                "S3 Intervalos - Repeticiones",
                "S4 Intervalos - Repeticiones",
                "S1 Tempo", "S2 Tempo", "S3 Tempo",
            }, activities)
            self.assertNotIn("Entrenamiento en cinta", activities)
            self.assertNotIn("Base", activities)
            self.assertNotIn("Carrera continua 20", activities)

            by_activity = {}
            for row in intervals:
                by_activity.setdefault(row["Actividad"], []).append(row)
            for activity_rows in by_activity.values():
                self.assertLessEqual(
                    sum(
                        row["Nivel"] == "Total de actividad"
                        for row in activity_rows
                    ),
                    1,
                )
                self.assertEqual({
                    "Calentamiento", "Trabajo activo",
                    "Recuperación", "Enfriamiento",
                }, {row["Tipo de intervalo"] for row in activity_rows})
        finally:
            workbook.close()

    def test_self_evaluation_coverage_uses_only_valid_activities(self):
        workbook = self.open_book(data_only=False)
        try:
            quality = next(
                row for row in self.table_values(workbook["CALIDAD DATOS"])
                if row["Métrica"] == "Autoevaluación de actividades válidas"
            )
            self.assertEqual(43, quality["Datos disponibles"])
            self.assertEqual(89, quality["Datos esperados"])
            self.assertEqual(46, quality["Datos ausentes"])
            self.assertAlmostEqual(0.483, quality["Cobertura"], places=3)

            summary = workbook["RESUMEN"]
            indicators = {
                summary.cell(row, 1).value: summary.cell(row, 2).value
                for row in range(1, summary.max_row + 1)
            }
            indicators.update({
                summary.cell(row, 4).value: summary.cell(row, 5).value
                for row in range(1, summary.max_row + 1)
            })
            self.assertAlmostEqual(
                0.483, indicators["Cobertura de autoevaluación"], places=3
            )
        finally:
            workbook.close()

    def test_session_type_requires_reliable_evidence(self):
        workbook = self.open_book(data_only=False)
        try:
            activities = {
                row["Actividad"]: row["Tipo de sesión"]
                for row in self.table_values(workbook["ACTIVIDADES"])
            }
            self.assertEqual(
                "Sin clasificar", activities["Entrenamiento en cinta"]
            )
            self.assertEqual("Sin clasificar", activities["Base"])
            self.assertNotEqual(
                "Intervalos", activities["Rock and roll 1/2 maratón 2026"]
            )
            self.assertEqual(
                "Competición", activities["Rock and roll 1/2 maratón 2026"]
            )
            self.assertEqual(
                "Intervalos", activities["S1 Intervalos - Repeticiones"]
            )
            self.assertEqual("Tempo", activities["S1 Tempo"])

            interval_active = [
                row for row in self.table_values(workbook["INTERVALOS"])
                if row["Tipo de intervalo"] == "Trabajo activo"
            ]
            self.assertTrue(interval_active)
            self.assertTrue(all(
                row["Nivel"] == "Resumen de bloques"
                for row in interval_active
            ))
        finally:
            workbook.close()

    def test_negative_sentinels_are_missing_before_visible_calculations(self):
        workbook = self.open_book(data_only=False)
        try:
            # Se revisan las hojas de mediciones con dominios no negativos. Un
            # -100 % comparativo en RESUMEN sí puede ser un cálculo válido.
            for sheet_name in ("SEMANAS", "SALUD DIARIA", "CALIDAD DATOS"):
                sheet = workbook[sheet_name]
                for row in sheet.iter_rows():
                    for cell in row:
                        self.assertNotEqual(
                            -1, cell.value, f"{sheet.title}!{cell.coordinate}"
                        )

            health = self.table_values(workbook["SALUD DIARIA"])
            self.assertEqual(
                6,
                sum(row["Estrés medio"] is None for row in health),
            )
            stress = next(
                row for row in self.table_values(workbook["CALIDAD DATOS"])
                if row["Métrica"] == "Estrés medio"
            )
            self.assertEqual(106, stress["Datos disponibles"])
            self.assertEqual(112, stress["Datos esperados"])
            self.assertEqual(6, stress["Datos ausentes"])
            self.assertAlmostEqual(0.946, stress["Cobertura"], places=3)
            self.assertEqual(6, len(stress["Periodos sin datos"].split(", ")))
        finally:
            workbook.close()

    def test_quality_uses_friendly_names_and_spanish_date_ranges(self):
        workbook = self.open_book(data_only=False)
        try:
            quality_text = "\n".join(
                str(cell.value)
                for row in workbook["CALIDAD DATOS"].iter_rows()
                for cell in row if cell.value is not None
            )
            visible_text = "\n".join(
                str(cell.value)
                for sheet in workbook.worksheets
                if sheet.sheet_state == "visible"
                for row in sheet.iter_rows()
                for cell in row if cell.value is not None
            )
            for technical in (
                "raw", "epochs", "endpoint", "lactate-threshold",
                "splits.vueltas", "typed_splits.splits", "summary",
                "detail", "split_summaries", "totalAscent", "totalDescent",
                "summaryDTO", "lapDTOs", "speed_raw",
            ):
                self.assertNotIn(technical.casefold(), visible_text.casefold())
            self.assertNotRegex(quality_text, r"\b\d{4}-\d{2}-\d{2}\b")
            self.assertNotIn("02/05/2026/04/05/2026", quality_text)
            self.assertIn("02/05/2026–04/05/2026", quality_text)
            for explanation in (
                "Se convirtieron las marcas de tiempo del sueño a fecha y hora local.",
                "La velocidad del umbral de lactato se convirtió a metros por segundo.",
                "Se evitó repetir información de vueltas procedente de distintas fuentes de Garmin.",
                "Se unificaron los resúmenes generales de cada actividad.",
                "Se eliminaron resúmenes de intervalos duplicados.",
                "El ascenso y descenso acumulados se interpretaron en metros.",
            ):
                self.assertIn(explanation, quality_text)
        finally:
            workbook.close()

    def test_table_headers_have_fixed_human_height(self):
        workbook = self.open_book(data_only=False)
        try:
            for name in (
                "SEMANAS", "ACTIVIDADES", "ZONAS", "SALUD DIARIA",
                "EQUIPAMIENTO",
            ):
                sheet = workbook[name]
                for table in sheet.tables.values():
                    from openpyxl.utils.cell import range_boundaries
                    _, header_row, _, _ = range_boundaries(table.ref)
                    self.assertGreaterEqual(
                        sheet.row_dimensions[header_row].height, 30, name
                    )
                    self.assertLessEqual(
                        sheet.row_dimensions[header_row].height, 36, name
                    )
                    for cell in sheet[header_row]:
                        self.assertEqual("center", cell.alignment.vertical)
                        self.assertTrue(cell.alignment.wrap_text)
        finally:
            workbook.close()

    def test_short_laps_keep_rows_but_hide_useless_paces(self):
        workbook = self.open_book(data_only=False)
        try:
            laps = self.table_values(workbook["VUELTAS"])
            short_laps = [row for row in laps if row["Distancia (m)"] < 100]
            self.assertEqual(69, len(short_laps))
            for row in short_laps:
                self.assertIsNone(row["Ritmo medio (min/km)"])
                self.assertIsNone(row["Mejor ritmo (min/km)"])
                self.assertIsNotNone(row["Distancia (m)"])
                self.assertIsNotNone(row["Duración"])
                self.assertEqual("Sí", row["Vuelta parcial"])
        finally:
            workbook.close()

    def test_intervals_journal_dates_and_collapsed_columns(self):
        workbook = self.open_book(data_only=False)
        try:
            intervals = self.table_values(workbook["INTERVALOS"])
            short = next(row for row in intervals if row["Distancia (km)"] == 0.05)
            self.assertEqual("Trabajo activo", short["Tipo de intervalo"])
            self.assertEqual("Resumen de bloques", short["Nivel"])
            for header in (
                "Ritmo medio (min/km)", "Ritmo en movimiento (min/km)",
                "Mejor ritmo (min/km)",
                "Ritmo ajustado por pendiente (min/km)",
            ):
                self.assertIsNone(short[header])

            journal = self.table_values(workbook["DIARIO"])
            self.assertEqual(6, len(journal))
            self.assertTrue(all(
                row["Nota"] or row["Tipo de sesión previsto"] for row in journal
            ))
            self.assertEqual("Prueba", journal[0]["Tipo de sesión previsto"])
            mismatches = [row for row in journal if row["Aviso"]]
            self.assertEqual(1, len(mismatches))
            self.assertNotEqual(
                mismatches[0]["Fecha del diario"],
                mismatches[0]["Fecha de la actividad"],
            )

            primary = {
                "ACTIVIDADES": {
                    "Fecha", "Actividad", "Deporte", "Tipo de sesión",
                    "Estado del registro", "Distancia (km)", "Duración",
                    "Ritmo medio (min/km)", "FC media (lpm)", "Carga Garmin",
                    "Esfuerzo percibido (1–10)", "Sensación", "Equipamiento",
                },
                "SEMANAS": {
                    "Semana", "Estado de la semana", "Sesiones totales",
                    "Días entrenados", "Carrera (km)", "Tiempo corriendo",
                    "Tirada más larga (km)", "Carga Garmin",
                    "Esfuerzo percibido medio", "Sueño medio",
                    "VFC nocturna (ms)", "Pulso en reposo (lpm)",
                },
                "SALUD DIARIA": {
                    "Fecha", "Pasos", "Pulso en reposo (lpm)", "Estrés medio",
                    "Body Battery máximo", "Sueño total",
                    "Puntuación del sueño", "VFC nocturna media (ms)",
                },
            }
            for sheet_name, expected_visible in primary.items():
                sheet = workbook[sheet_name]
                visible_headers = {
                    cell.value for cell in sheet[1]
                    if not sheet.column_dimensions[cell.column_letter].hidden
                }
                self.assertEqual(expected_visible, visible_headers)
                self.assertTrue(any(
                    dimension.hidden and dimension.outlineLevel == 1
                    for dimension in sheet.column_dimensions.values()
                ))
        finally:
            workbook.close()

    def test_visible_quality_dates_identifiers_and_weight_axis_are_friendly(self):
        workbook = self.open_book(data_only=False)
        try:
            quality_text = "\n".join(
                str(cell.value)
                for row in workbook["CALIDAD DATOS"].iter_rows()
                for cell in row if cell.value is not None
            )
            for technical in (
                "get_sleep_data", "get_hrv_data", "dailySleepDTO",
                "hrvSummary", "gear_type", "lap_type", "session_type",
            ):
                self.assertNotIn(technical.casefold(), quality_text.casefold())
            self.assertNotRegex(quality_text, r"\b\d{4}-\d{2}-\d{2}\b")

            for sheet in workbook.worksheets:
                if sheet.sheet_state != "visible":
                    continue
                for table in sheet.tables.values():
                    from openpyxl.utils.cell import range_boundaries
                    min_col, min_row, max_col, max_row = range_boundaries(table.ref)
                    for column in range(min_col, max_col + 1):
                        header = str(sheet.cell(min_row, column).value or "")
                        dimension = sheet.column_dimensions[
                            sheet.cell(min_row, column).column_letter
                        ]
                        if "ID" in header:
                            self.assertTrue(dimension.hidden, f"{sheet.title}: {header}")
                        if dimension.hidden or not any(
                            word in header for word in ("Fecha", "Desde", "Hasta")
                        ):
                            continue
                        for row in range(min_row + 1, max_row + 1):
                            cell = sheet.cell(row, column)
                            if cell.value is not None:
                                self.assertIn("dd/mm", cell.number_format.casefold())

            hidden_technical = {
                sheet.title for sheet in workbook.worksheets
                if sheet.sheet_state == "hidden"
                and sheet.title.startswith("TÉCNICO - ")
            }
            self.assertEqual({
                "TÉCNICO - ACTIVIDADES",
                "TÉCNICO - MODELO",
                "TÉCNICO - MAPEO",
            }, hidden_technical)

            measures = workbook["MEDIDAS"]
            weight_chart = measures._charts[0]
            self.assertEqual(
                "'MEDIDAS'!$A$2:$A$19",
                weight_chart.series[0].cat.numRef.f,
            )
            self.assertEqual("dd/mm", weight_chart.x_axis.numFmt.formatCode)
            self.assertFalse(weight_chart.x_axis.delete)
            self.assertEqual(
                "'RESUMEN'!$N$2:$N$18",
                workbook["RESUMEN"]._charts[0].series[0].cat.numRef.f,
            )
        finally:
            workbook.close()

    def test_four_complete_weeks_compare_two_against_two(self):
        weeks = []
        for index in range(4):
            weeks.append({
                "iso_week": f"2026-W{index + 1:02d}",
                "status": "complete",
                "training_sessions_total": 2,
                "days_with_any_training": 2,
                "running_sessions": 2,
                "running_distance_m": 20_000 + index * 1_000,
                "running_duration_s": 7_200,
                "longest_run_distance_m": 12_000,
                "garmin_training_load_total": 100,
                "session_rpe_load_total": 200,
            })
        model = {
            "export_metadata": {"schema_version": SCHEMA_VERSION},
            "weekly_timeline": weeks,
            "period_summary": {},
            "data_quality": {},
        }
        with tempfile.TemporaryDirectory() as directory:
            output = Path(directory) / "cuatro_semanas.xlsx"
            render_xlsx(model, output)
            workbook = load_workbook(output, data_only=False)
            try:
                values = {
                    str(cell.value)
                    for row in workbook["RESUMEN"].iter_rows()
                    for cell in row if cell.value is not None
                }
                self.assertIn(
                    "Comparación de 2 semanas completas con las 2 anteriores",
                    values,
                )
                self.assertIn("2 semanas anteriores", values)
                self.assertIn("2 semanas recientes", values)
            finally:
                workbook.close()

        self.assertEqual("3.3.1", SCHEMA_VERSION)

    def test_user_names_help_tables_charts_and_visual_settings(self):
        workbook = self.open_book(data_only=False)
        try:
            activities = self.table_values(workbook["ACTIVIDADES"])
            self.assertEqual("Registro breve de carrera", activities[0]["Actividad"])
            gear = self.table_values(workbook["EQUIPAMIENTO"])
            self.assertEqual("Mis Zapatillas Ñandú", gear[0]["Nombre"])
            help_rows = self.table_values(workbook["AYUDA"])
            self.assertGreater(len(help_rows), 100)
            self.assertTrue(all(row["Qué significa"] for row in help_rows))
            self.assertFalse(any(
                "Campo del modelo semántico" in row["Qué significa"]
                for row in help_rows
            ))
            self.assertEqual(3, len(workbook["RESUMEN"]._charts))
            self.assertEqual(1, len(workbook["MEDIDAS"]._charts))
            self.assertTrue(all(
                chart.legend is None
                for chart in workbook["RESUMEN"]._charts[:2]
            ))
            self.assertIsNone(workbook["MEDIDAS"]._charts[0].legend)
            self.assertFalse(workbook["INICIO"].sheet_view.showGridLines)
            self.assertFalse(workbook["RESUMEN"].sheet_view.showGridLines)
            for name in (
                "SEMANAS", "ACTIVIDADES", "INTERVALOS", "VUELTAS",
                "ZONAS", "SALUD DIARIA", "EQUIPAMIENTO", "DIARIO",
                "CALIDAD DATOS", "AYUDA",
            ):
                self.assertTrue(workbook[name].tables, name)
                self.assertIsNotNone(workbook[name].freeze_panes, name)
            for name in ("ACTIVIDADES", "INTERVALOS", "VUELTAS", "EQUIPAMIENTO", "DIARIO"):
                sheet = workbook[name]
                hidden_columns = [
                    dimension
                    for dimension in sheet.column_dimensions.values()
                    if dimension.hidden
                ]
                self.assertTrue(hidden_columns, name)

            quality_text = "\n".join(
                str(cell.value)
                for row in workbook["CALIDAD DATOS"].iter_rows()
                for cell in row
                if cell.value is not None
            )
            self.assertNotIn("Data Quality.coverage", quality_text)
            self.assertNotIn("Valores aún sin traducción específica", quality_text)
            self.assertIn("Revisa la cobertura de datos", quality_text)
            self.assertIn("10/04/2026", quality_text)

            help_units = {
                row["Unidad"] for row in help_rows if row["Hoja"] == "INICIO"
            }
            self.assertNotIn("Según el campo", help_units)
        finally:
            workbook.close()

    def test_workbook_xml_is_valid_and_contains_no_personal_fixture(self):
        self.assertGreater(self.output.stat().st_size, 10_000)
        with zipfile.ZipFile(self.output) as archive:
            self.assertIsNone(archive.testzip())
            workbook_xml = archive.read("xl/workbook.xml").decode("utf-8")
            self.assertNotIn("\\users\\", workbook_xml.casefold())
            self.assertNotIn(".garminconnect", workbook_xml.casefold())
        workbook = self.open_book(data_only=False)
        try:
            start_values = {
                cell.value
                for row in workbook["INICIO"].iter_rows()
                for cell in row
                if isinstance(cell.value, str)
            }
            self.assertTrue(any(
                XLSX_PRESENTATION_VERSION in value for value in start_values
            ))
        finally:
            workbook.close()

    def test_large_series_are_omitted_only_from_xlsx_and_reported(self):
        samples = [[1.0]] * 25_001
        model = {
            "export_metadata": {
                "schema_version": SCHEMA_VERSION,
                "export_status": "completed",
            },
            "activities": [{
                "activity_ref": "activity_0123456789ab",
                "activity_series": {
                    "metric_descriptors": [{
                        "field": "elapsed_duration_raw",
                        "source_field": "sumElapsedDuration",
                        "source_unit": "second",
                    }],
                    "samples": samples,
                },
            }],
            "data_quality": {},
        }
        with tempfile.TemporaryDirectory() as directory:
            output = Path(directory) / "series_grandes.xlsx"
            render_xlsx(model, output)
            workbook = load_workbook(output, data_only=False)
            try:
                self.assertNotIn("DATOS POR SEGUNDO", workbook.sheetnames)
                quality_text = "\n".join(
                    str(cell.value)
                    for row in workbook["CALIDAD DATOS"].iter_rows()
                    for cell in row
                    if cell.value is not None
                )
                self.assertIn("25.001 muestras", quality_text)
                self.assertIn("TXT conserva las series completas", quality_text)
            finally:
                workbook.close()
        self.assertEqual(25_001, len(samples))


if __name__ == "__main__":
    unittest.main()
